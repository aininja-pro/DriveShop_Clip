import streamlit as st
import pandas as pd
import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
import time
import json
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, DataReturnMode, GridUpdateMode
from src.utils.logger import logger
from src.utils.sentiment_analysis import run_sentiment_analysis
from src.utils.fms_api import FMSAPIClient
from src.utils.auth_improved import ImprovedSupabaseAuth as SupabaseAuth
from src.dashboard.strategic_intelligence_json_display import display_strategic_intelligence_tab
from src.dashboard.message_pullthrough_clean import display_pullthrough_analysis_tab
from src.dashboard.oem_messaging_ui import display_oem_messaging_tab
from src.dashboard.historical_reprocessing import display_historical_reprocessing_tab
from src.dashboard.active_jobs_tab import display_active_jobs_tab, submit_job_to_queue
from src.dashboard.cooldown_management import display_cooldown_management_tab
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
from streamlit_extras.stylable_container import stylable_container

# Initialize environment (handles .env loading gracefully)
from src.config.env import init_environment
from src.utils.apify_healthcheck import apify_startup_check
from PIL import Image

# Initialize environment variables (works with or without .env file)
init_environment()

# Validate Apify configuration at startup (warn but don't kill app)
try:
    apify_startup_check()
except RuntimeError as e:
    print(f"[STARTUP WARNING] {e}")
    print("[STARTUP] Apify will be disabled until configuration is fixed")
    # Continue loading - don't kill the UI!

# Initialize authentication
auth = SupabaseAuth()

# Debug: Print loaded environment variables (without exposing full API keys)
# Only print on first load
if 'env_vars_logged' not in st.session_state:
    openai_key = os.environ.get('OPENAI_API_KEY', '')
    slack_webhook = os.environ.get('SLACK_WEBHOOK_URL', '')
    print(f"OPENAI_API_KEY loaded: {'Yes (starts with ' + openai_key[:5] + '...)' if openai_key else 'No'}")
    print(f"SLACK_WEBHOOK_URL loaded: {'Yes' if slack_webhook else 'No'}")
    st.session_state.env_vars_logged = True

# Add the project root to the path
project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

# Helper functions for progress tracking
def update_progress(current, total):
    """Update progress in session state"""
    if 'processing_progress' in st.session_state:
        st.session_state['processing_progress']['current'] = current
        st.session_state['processing_progress']['total'] = total

def format_time(seconds):
    """Format seconds into human-readable time"""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        return f"{int(seconds // 60)}m {int(seconds % 60)}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

# Import local modules
try:
    from src.ingest.ingest_database import run_ingest_database, run_ingest_database_with_filters
    from src.utils.database import get_database
except ImportError:
    # Define a stub for when the module is not yet implemented
    def run_ingest_database(file_path):
        st.error("Database ingest module not implemented yet")
        return False
    def run_ingest_database_with_filters(url, filters):
        st.error("Database ingest with filters module not implemented yet")
        return False
    def get_database():
        st.error("Database module not available")
        return None

# Cache the database connection to prevent multiple connections
@st.cache_resource
def get_cached_database():
    """Get a cached database connection that persists across reruns"""
    return get_database()

@st.cache_data
def load_person_outlets_mapping():
    """Load Person_ID to Media Outlets mapping from JSON file"""
    try:
        mapping_file = os.path.join(project_root, "data", "person_outlets_mapping.json")
        if os.path.exists(mapping_file):
            with open(mapping_file, 'r') as f:
                mapping = json.load(f)
            print(f"✅ Loaded Person_ID mapping with {len(mapping)} unique Person_IDs")
            return mapping
        else:
            print("⚠️ Person_ID mapping file not found")
            return {}
    except Exception as e:
        print(f"❌ Error loading Person_ID mapping: {e}")
        return {}

def get_outlet_options_for_person(person_id, mapping):
    """Get list of outlet names for a given Person_ID"""
    if not person_id or not mapping:
        return []
    
    person_id_str = str(person_id)
    if person_id_str in mapping:
        # FIX: Correctly access the nested 'outlets' list
        person_data = mapping.get(person_id_str, {})
        outlets_list = person_data.get('outlets', [])
        return [outlet['outlet_name'] for outlet in outlets_list]
    return []

def get_full_outlet_data_for_person(person_id, mapping):
    """Get full outlet data (name, id, impressions) for a given Person_ID"""
    if not person_id or not mapping:
        return {}
    
    person_id_str = str(person_id)
    if person_id_str in mapping:
        person_data = mapping.get(person_id_str, {})
        outlets_list = person_data.get('outlets', [])
        # Return a dict mapping outlet_name to full outlet data
        return {outlet['outlet_name']: outlet for outlet in outlets_list}
    return {}

@st.cache_data
def create_reporter_name_to_id_mapping():
    """Create a mapping from Reporter Name to Person_ID for lookups."""
    try:
        mapping_file = os.path.join(project_root, "data", "person_outlets_mapping.csv")
        if not os.path.exists(mapping_file):
            return {}
        
        df = pd.read_csv(mapping_file)
        # Ensure correct types
        df['Reporter_Name'] = df['Reporter_Name'].astype(str)
        df['Person_ID'] = df['Person_ID'].astype(str)
        
        # Create a dictionary from the two columns, dropping duplicates
        # In case a name is associated with multiple IDs, this takes the first one.
        # NORMALIZE names to handle spacing and case variations
        df['Reporter_Name_Normalized'] = df['Reporter_Name'].str.strip().str.replace(r'\s+', ' ', regex=True).str.title()
        name_to_id_map = df.drop_duplicates('Reporter_Name_Normalized').set_index('Reporter_Name_Normalized')['Person_ID'].to_dict()
        print(f"✅ Created Reporter Name to Person_ID mapping for {len(name_to_id_map)} reporters.")
        return name_to_id_map
    except Exception as e:
        print(f"❌ Error creating reporter name to ID mapping: {e}")
        return {}

def load_loans_data_for_filtering(url: str):
    """
    Load loans data from URL for preview and filtering without processing.
    Returns (success: bool, df: pd.DataFrame, data_info: dict)
    """
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        headers = [
            "ActivityID", "Person_ID", "Make", "Model", "WO #", "Office", "To", 
            "Affiliation", "Start Date", "Stop Date", "Model Short Name", "Links"
        ]
        
        csv_content = response.content.decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_content), header=None, names=headers, on_bad_lines='warn')
        
        df.columns = [col.strip() for col in df.columns]

        total_records = len(df)
        offices_count = df['Office'].nunique() if 'Office' in df.columns else 0
        makes_count = df['Make'].nunique() if 'Make' in df.columns else 0

        data_info = {
            "total_records": total_records,
            "offices_count": offices_count,
            "makes_count": makes_count
        }
        return True, df, data_info
    except Exception as e:
        st.error(f"Failed to load loans data: {e}")
        return False, None, {"error": str(e)}

# Load environment variables
def load_env():
    """Get environment variables (no longer loads .env - handled at startup)"""
    # Environment already loaded by init_environment() at startup
    return os.environ.get("STREAMLIT_PASSWORD", "password")

# Helper function to parse URL tracking data
def parse_url_tracking(df_row):
    """Parse URL tracking data from the backend to show multiple URL information"""
    try:
        # Check if URL_Tracking column exists and has data
        if 'URL_Tracking' in df_row and pd.notna(df_row['URL_Tracking']):
            # Try to parse as JSON if it's a string
            if isinstance(df_row['URL_Tracking'], str):
                url_tracking_str = df_row['URL_Tracking']
                
                # Handle Python-style single quotes by converting to valid JSON
                # Replace single quotes with double quotes, but be careful about quotes inside strings
                try:
                    # First try direct JSON parsing (in case it's already valid JSON)
                    url_tracking = json.loads(url_tracking_str)
                except json.JSONDecodeError:
                    # If that fails, try using Python's eval (safe since it's our own data)
                    # This handles Python-style single quotes
                    url_tracking = eval(url_tracking_str)
            else:
                url_tracking = df_row['URL_Tracking']
            
            return url_tracking
        else:
            # Fallback: infer from available data
            urls_processed = df_row.get('URLs_Processed', 1)
            urls_successful = df_row.get('URLs_Successful', 1)
            
            # Create a simple tracking structure
            return [{
                'original_url': df_row.get('Links', df_row.get('Clip URL', '')),
                'actual_url': df_row.get('Clip URL', ''),
                'success': True,
                'relevance_score': df_row.get('Relevance Score', 0),
                'content_type': 'inferred'
            }]
    except Exception as e:
        # If parsing fails, return basic structure
        return [{
            'original_url': df_row.get('Links', df_row.get('Clip URL', '')),
            'actual_url': df_row.get('Clip URL', ''),
            'success': True,
            'relevance_score': df_row.get('Relevance Score', 0),
            'content_type': 'inferred'
        }]


# Custom CSS for black sidebar with logo
def apply_custom_sidebar_styling():
    """Apply custom CSS styling for black sidebar with white logo"""
    st.markdown("""
    <style>
    /* Force light mode appearance */
    :root {
        color-scheme: light !important;
    }
    
    html[data-theme="dark"] {
        color-scheme: light !important;
    }
    
    .stApp {
        background-color: #ffffff !important;
        color: #000000 !important;
    }
    
    /* Force light theme on all elements */
    * {
        color-scheme: light !important;
    }
    
    /* Main content area styling */
    .main .block-container {
        background-color: #ffffff !important;
        color: #000000 !important;
    }
    
    /* Tab styling - Active tab colored, inactive tabs off-white */
    .stTabs {
        background-color: #ffffff !important;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        background-color: transparent !important;
        gap: 4px;
        padding: 0 8px;
    }
    
    /* Base tab styling - All tabs start with off-white */
    .stTabs [data-baseweb="tab"] {
        height: 45px !important;
        padding-left: 24px !important;
        padding-right: 24px !important;
        border-radius: 8px 8px 0 0 !important;
        border: none !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
        margin-bottom: 0 !important;
        background-color: #f5f5f5 !important;
        color: #666666 !important;
    }
    
    /* Hover effect for inactive tabs */
    .stTabs [data-baseweb="tab"]:not([aria-selected="true"]):hover {
        background-color: #e8e8e8 !important;
        color: #333333 !important;
    }
    
    /* Active tab styling - Blue with white text */
    .stTabs [aria-selected="true"] {
        background-color: #1976d2 !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1) !important;
    }
    
    /* Active tab text color */
    .stTabs [aria-selected="true"] p {
        color: #ffffff !important;
    }
    
    /* Inactive tab text color */
    .stTabs [aria-selected="false"] p {
        color: #666666 !important;
    }
    
    /* Remove the default bottom border indicator */
    .stTabs [data-baseweb="tab-highlight"] {
        background-color: transparent !important;
    }
    
    /* Tab text styling */
    .stTabs [data-baseweb="tab"] p {
        font-size: 14px !important;
        margin: 0 !important;
        line-height: 1.5 !important;
    }
    
    /* Ensure all text is black in main content */
    .main .block-container p,
    .main .block-container span,
    .main .block-container div,
    .main .block-container label {
        color: #000000 !important;
    }
    
    /* Headers in main content */
    .main h1, .main h2, .main h3, .main h4, .main h5, .main h6 {
        color: #000000 !important;
    }
    
    /* Black sidebar styling */
    .css-1d391kg, .css-1lcbmhc, .css-17lntkn, .css-1y4p8pa, 
    .stSidebar > div:first-child, .css-12oz5g7, .css-1cypcdb {
        background-color: #000000 !important;
    }
    
    /* Sidebar content styling */
    .stSidebar {
        background-color: #000000 !important;
    }
    
    /* All text in sidebar to white */
    .stSidebar * {
        color: white !important;
    }
    
    /* Sidebar headers and labels */
    .stSidebar .stMarkdown h1,
    .stSidebar .stMarkdown h2, 
    .stSidebar .stMarkdown h3,
    .stSidebar .stMarkdown h4,
    .stSidebar .stMarkdown h5,
    .stSidebar .stMarkdown h6 {
        color: white !important;
        font-weight: 600;
    }
    
    /* Sidebar radio buttons and selectbox text */
    .stSidebar .stRadio label,
    .stSidebar .stSelectbox label,
    .stSidebar .stMultiSelect label,
    .stSidebar .stTextInput label,
    .stSidebar .stNumberInput label {
        color: white !important;
        font-weight: 500;
    }
    
    /* Radio button options */
    .stSidebar .stRadio > div > div > div > label {
        color: white !important;
    }
    
    /* Selectbox dropdown styling - Updated to fix white text issue */
    .stSidebar .stSelectbox > div > div {
        background-color: white !important;
        border: 1px solid #333333 !important;
        color: #000000 !important;
    }
    
    /* Fix input text colors - make text dark in input fields */
    .stSidebar .stTextInput > div > div > input {
        color: #000000 !important;
        background-color: white !important;
    }
    
    .stSidebar .stNumberInput > div > div > input {
        color: #000000 !important;
        background-color: white !important;
    }
    
    /* Selectbox selected value display */
    .stSidebar .stSelectbox > div > div > div {
        color: #000000 !important;
        background-color: white !important;
    }
    
    /* Additional fix for selectbox text visibility */
    .stSidebar [data-baseweb="select"] > div {
        background-color: white !important;
    }
    
    .stSidebar [data-baseweb="select"] > div > div {
        color: #000000 !important;
    }
    
    /* More specific selectors for the displayed selected value */
    .stSidebar [data-baseweb="select"] [data-baseweb="tag"] {
        color: #000000 !important;
        background-color: #f0f0f0 !important;
    }
    
    /* Target the specific div that contains the selected text */
    .stSidebar [data-baseweb="select"] > div > div > div {
        color: #000000 !important;
    }
    
    /* Another approach - target by class if present */
    .stSidebar .css-1wa3eu0-placeholder {
        color: #000000 !important;
    }
    
    .stSidebar div[data-baseweb="select"] span {
        color: #000000 !important;
    }
    
    /* Dropdown options styling */
    .stSidebar .stSelectbox option {
        color: #000000 !important;
        background-color: white !important;
    }
    
    /* Fix all input elements in sidebar */
    .stSidebar input {
        color: #000000 !important;
        background-color: white !important;
    }
    
    /* Simple fix for number input buttons */
    .stSidebar button {
        color: #000000 !important;
        font-weight: bold !important;
    }
    
    /* Fix multiselect and other input types */
    .stSidebar .stMultiSelect > div > div > div {
        color: #000000 !important;
        background-color: white !important;
    }
    
    /* Ensure placeholder text is visible */
    .stSidebar input::placeholder {
        color: #6c757d !important;
    }
    
    /* Fix text area if any */
    .stSidebar textarea {
        color: #2c3e50 !important;
        background-color: white !important;
    }
    
    /* Fix file uploader visibility */
    .stSidebar .stFileUploader {
        color: white !important;
    }
    
    .stSidebar .stFileUploader label {
        color: white !important;
        font-weight: 500;
    }
    
    .stSidebar .stFileUploader > div {
        background-color: white !important;
        border: 2px dashed #000000 !important;
        border-radius: 6px !important;
    }
    
    .stSidebar .stFileUploader > div > div {
        color: #000000 !important;
        font-weight: 600 !important;
    }
    
    .stSidebar .stFileUploader button {
        color: white !important;
        background-color: #000000 !important;
        border: 1px solid #000000 !important;
        font-weight: 600 !important;
        padding: 8px 16px !important;
        border-radius: 4px !important;
    }
    
    .stSidebar .stFileUploader button:hover {
        background-color: #333333 !important;
        border-color: #333333 !important;
    }
    
    /* Fix file uploader text and button visibility - make text darker */
    .stSidebar .stFileUploader span,
    .stSidebar .stFileUploader p,
    .stSidebar .stFileUploader div[data-testid="stFileUploaderDropzone"] {
        color: #000000 !important;
        font-weight: 600 !important;
        font-size: 14px !important;
    }
    
    .stSidebar .stFileUploader div[data-testid="stFileUploaderDropzone"] span {
        color: #000000 !important;
        font-weight: 500 !important;
    }
    
    /* Make drag and drop text more visible */
    .stSidebar .stFileUploader div[data-testid="stFileUploaderDropzone"] p {
        color: #000000 !important;
        font-weight: 600 !important;
        margin: 0 !important;
    }
    
    /* Button styling in sidebar - LIGHT BLUE BUTTONS with DARK TEXT for visibility */
    .stSidebar .stButton > button {
        background-color: #e3f2fd !important;
        color: #000000 !important;
        border: 2px solid #90caf9 !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;
        padding: 12px 16px !important;
        min-height: 44px !important;
        width: 100% !important;
        text-align: center !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }
    
    .stSidebar .stButton > button:hover {
        background-color: #bbdefb !important;
        color: #0d47a1 !important;
        border-color: #64b5f6 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.25) !important;
    }
    
    /* Force dark blue text on all sidebar buttons - override any inherited styles */
    .stSidebar .stButton > button * {
        color: #000000 !important;
    }
    
    .stSidebar .stButton > button span {
        color: #000000 !important;
    }
    
    .stSidebar .stButton > button:hover * {
        color: #0d47a1 !important;
    }
    
    .stSidebar .stButton > button:hover span {
        color: #0d47a1 !important;
    }
    
    /* Logo container styling */
    .sidebar-logo {
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 1rem 0 2rem 0;
        margin-bottom: 1rem;
        border-bottom: 1px solid #333333;
    }
    
    .sidebar-logo img {
        max-width: 180px;
        height: auto;
        filter: brightness(1.1);
    }
    
    /* Style Streamlit's image component in sidebar */
    .stSidebar .stImage {
        display: flex;
        justify-content: center;
        padding: 1rem 0 2rem 0;
        margin-bottom: 1rem;
        border-bottom: 1px solid #333333;
    }
    
    .stSidebar .stImage img {
        filter: brightness(1.1);
        border-radius: 4px;
    }
    
    /* Sidebar metrics styling */
    .stSidebar .metric-container {
        background-color: #1a1a1a !important;
        padding: 0.5rem;
        border-radius: 6px;
        margin: 0.5rem 0;
        border: 1px solid #333333;
    }
    
    /* Fix for any remaining black text */
    .stSidebar p, .stSidebar span, .stSidebar div {
        color: white !important;
    }
    
    /* Tab styling adjustments for better contrast with black sidebar */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #f8f9fa;
        border-radius: 6px 6px 0 0;
        padding: 0.5rem 1rem;
        font-weight: 500;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #000000;
        color: white;
    }
    
    /* 1) Invert / restore visibility of number-input spinners in WebKit browsers */
    .stSidebar input[type="number"]::-webkit-inner-spin-button,
    .stSidebar input[type="number"]::-webkit-outer-spin-button {
        -webkit-appearance: initial !important;
        filter: invert(1) contrast(200%) !important;
        opacity: 1 !important;
    }

    /* 2) Firefox: turn numbers back into textfields (hides native arrows) */
    .stSidebar input[type="number"] {
        -moz-appearance: textfield !important;
    }

    /* 3) Add a visible border around inputs so they don't vanish on white */
    .stSidebar .stTextInput > div > div > input,
    .stSidebar .stNumberInput > div > div > input {
        border: 1px solid #333333 !important;
        border-radius: 4px !important;
        padding: 4px 8px !important;
    }

    /* 4) Make the file-uploader icon and emoji darker so it shows on charcoal */
    .stSidebar .stFileUploader div[data-testid="stFileUploaderDropzone"] svg {
        filter: brightness(0) invert(1) !important;
        width: 1.2rem !important; height: 1.2rem !important;
    }

    /* 5) Tidy up any stray SVGs or emojis inside buttons */
    .stSidebar .stButton > button svg,
    .stSidebar .stButton > button {
        filter: none !important;
        color: white !important;
    }
    
    /* --- Always show +/– on number inputs --- */
    .stSidebar input[type="number"]::-webkit-inner-spin-button,
    .stSidebar input[type="number"]::-webkit-outer-spin-button {
      /* Make sure they're rendered as controls */
      -webkit-appearance: inner-spin-button !important;
      display: block !important;
      opacity: 1 !important;
      visibility: visible !important;
      pointer-events: auto !important;
      width: 1.2em !important;
      height: 1.2em !important;
    }

    /* Firefox fallback — revert to textfield so at least you can type */
    .stSidebar input[type="number"] {
      -moz-appearance: textfield !important;
    }

    /* Tidy up the outline and border so you can see the control region */
    .stSidebar .stNumberInput > div > div > input {
      border: 1px solid #333333 !important;
      padding-right: 1.6em !important;  /* leave room for the spinner */
      border-radius: 4px !important;
      background-clip: padding-box !important;
    }
    
    /* ALWAYS VISIBLE & DARKENED steppers */
    .stSidebar input[type="number"]::-webkit-inner-spin-button,
    .stSidebar input[type="number"]::-webkit-outer-spin-button {
      /* treat them as real controls */
      -webkit-appearance: inner-spin-button !important;
      display: block !important;
      visibility: visible !important;
      opacity: 1 !important;
      pointer-events: auto !important;

      /* tint them dark */
      filter: brightness(0%) contrast(100%) !important;
    }

    /* Firefox fallback — hide arrows, let user type */
    .stSidebar input[type="number"] {
      -moz-appearance: textfield !important;
    }
    
    /* Compact table area */
    .ag-theme-alpine {
        font-size: 0.8rem !important;
    }
    
    /* Sticky Action Bar for Bulk Review */
    .sticky-action-bar {
        position: fixed;
        bottom: 0;
        left: 0;
        right: 0;
        background: linear-gradient(135deg, #000000 0%, #2c3e50 100%);
        padding: 0.8rem 1rem;
        box-shadow: 0 -4px 20px rgba(0,0,0,0.3);
        border-top: 2px solid #333;
        z-index: 1000;
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 1rem;
    }
    
    .sticky-action-bar .stButton > button {
        background: linear-gradient(135deg, #000000 0%, #333333 100%) !important;
        color: white !important;
        border: 2px solid #444 !important;
        border-radius: 8px !important;
        padding: 0.6rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3) !important;
        min-width: 140px !important;
    }
    
    .sticky-action-bar .stButton > button:hover {
        background: linear-gradient(135deg, #333333 0%, #555555 100%) !important;
        border-color: #666 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.4) !important;
    }
    
    /* Add bottom padding to main content to prevent overlap */
    .main .block-container {
        padding-bottom: 5rem !important;
    }
    
    /* Target ALL metric text elements but keep READABLE */
    div[data-testid="metric-container"] > div,
    div[data-testid="metric-container"] span,
    div[data-testid="metric-container"] p,
    [data-testid="metric-container"] * {
        font-size: 0.9rem !important;
        line-height: 1.2 !important;
        margin: 0.1rem 0 !important;
        padding: 0 !important;
    }
    
    /* Force label text to be small but readable */
    div[data-testid="metric-container"] > div:first-child,
    [data-testid="metric-container"] [data-testid="metric-label"],
    .metric-label {
        font-size: 0.75rem !important;
        font-weight: 500 !important;
        color: #6c757d !important;
        margin-bottom: 0.15rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.3px !important;
        line-height: 1.1 !important;
    }
    
    /* Force value text to be bigger and VISIBLE */
    div[data-testid="metric-container"] > div:last-child,
    [data-testid="metric-container"] [data-testid="metric-value"],
    .metric-value {
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        line-height: 1.2 !important;
        color: #2c3e50 !important;
    }
    
    /* Override Streamlit's default metric styling */
    .stMetric {
        padding: 0.15rem !important;
        margin: 0.15rem !important;
        height: auto !important;
        min-height: 2.5rem !important;
    }
    
    .stMetric > div {
        font-size: 0.9rem !important;
        line-height: 1.2 !important;
    }
    
    /* Nuclear option - target by text content if needed */
    div:contains("Total Clips"),
    div:contains("Avg Score"), 
    div:contains("High Quality"),
    div:contains("Approved") {
        font-size: 0.1rem !important;
    }
    
    /* Table row styling */
    .table-row {
        padding: 6px 4px;
        min-height: 1.8rem;
        display: flex;
        align-items: center;
        text-align: center;
        font-size: 0.7rem;
        line-height: 1.2;
    }
    
    /* URL popup styling */
    .url-popup {
        background-color: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    
    .url-item {
        padding: 0.25rem 0;
        border-bottom: 1px solid #e9ecef;
    }
    
    .url-item:last-child {
        border-bottom: none;
    }
    
    /* Better button targeting - all buttons in action columns */
    div[data-testid="column"]:last-child div[data-testid="stButton"]:nth-child(1) button {
        background: #28a745 !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        height: 1.8rem !important;
    }
    
    div[data-testid="column"]:last-child div[data-testid="stButton"]:nth-child(3) button {
        background: #dc3545 !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        height: 1.8rem !important;
    }
    
    /* More aggressive targeting for all checkmark buttons */
    button:contains("✓") {
        background-color: #28a745 !important;
        color: white !important;
    }
    
    /* More aggressive targeting for all X buttons */
    button:contains("✗") {
        background-color: #dc3545 !important;
        color: white !important;
    }
    
    /* Universal button override */
    .stButton button {
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        border: none !important;
        border-radius: 4px !important;
        height: 1.8rem !important;
    }
    
    /* Score color styling */
    .score-high { color: #28a745; font-weight: 600; }
    .score-med { color: #007bff; font-weight: 600; }
    .score-low { color: #ffc107; font-weight: 600; }
    
    /* Scrollable table styling */
    .scrollable-table {
        max-height: 450px;
        overflow-y: auto;
        border: 1px solid #e0e0e0;
        border-radius: 4px;
        background: white;
    }
    
    .fixed-header {
        position: sticky;
        top: 0;
        z-index: 100;
        background: white;
        border-bottom: 2px solid #adb5bd;
        padding: 0.2rem 0;
    }
    
    /* Smooth scrolling */
    .scrollable-table::-webkit-scrollbar {
        width: 8px;
    }
    
    .scrollable-table::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }
    
    .scrollable-table::-webkit-scrollbar-thumb {
        background: #c1c1c1;
        border-radius: 4px;
    }
    
    .scrollable-table::-webkit-scrollbar-thumb:hover {
        background: #a8a8a8;
    }
    
    /* Compact sidebar styling */
    .css-1d391kg {
        width: 250px !important;
        min-width: 250px !important;
    }
    
    /* Make sidebar content more compact */
    .css-1d391kg .stMarkdown p {
        margin: 0.2rem 0 !important;
        font-size: 0.9rem !important;
    }
    
    .css-1d391kg .stButton button {
        padding: 0.3rem 0.5rem !important;
        font-size: 0.8rem !important;
        margin: 0.1rem 0 !important;
    }
    
    .css-1d391kg .stFileUploader {
        margin: 0.2rem 0 !important;
    }
    
    .css-1d391kg .stFileUploader > div {
        padding: 0.3rem !important;
        font-size: 0.8rem !important;
    }
    
    /* Compact success/error messages in sidebar */
    .css-1d391kg .stAlert {
        padding: 0.3rem !important;
        font-size: 0.8rem !important;
        margin: 0.2rem 0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

def create_client_excel_report(df, approved_df=None):
    """Create a professional Excel report for client presentation"""
    
    # Handle empty DataFrame case
    if df is None or df.empty:
        st.warning("⚠️ No data available to create Excel report. Please process some loans first.")
        # Return a minimal workbook with just headers
        wb = Workbook()
        wb.remove(wb.active)
        summary_ws = wb.create_sheet("No Data Available")
        summary_ws.append(["No clips found", "Please process loans first"])
        return wb
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Executive Summary Sheet
    summary_ws = wb.create_sheet("Executive Summary")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Summary metrics (using Bulk Review column names)
    clips_found = len(df)  # This is the successful clips count
    relevance_col = 'Relevance Score' if 'Relevance Score' in df.columns else 'Relevance'
    
    # Try to get total original records by checking for rejected clips file
    try:
        import os
        project_root = Path(__file__).parent.parent.parent
        rejected_file = os.path.join(project_root, "data", "rejected_clips.csv")
        if os.path.exists(rejected_file):
            rejected_df = pd.read_csv(rejected_file)
            clips_not_found = len(rejected_df)
        else:
            rejected_df = None
            clips_not_found = 0
    except:
        rejected_df = None
        clips_not_found = 0
    
    # Calculate total records
    total_records = clips_found + clips_not_found
    
    # Other metrics
    avg_relevance = df[relevance_col].mean() if relevance_col in df.columns and len(df) > 0 else 0
    high_quality = len(df[df[relevance_col] >= 8]) if relevance_col in df.columns else 0
    positive_sentiment = len(df[df['Sentiment'] == 'positive']) if 'Sentiment' in df.columns else 0
    
    # Add summary data
    summary_data = [
        ["DriveShop Media Monitoring Report", ""],
        ["Report Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p")],
        ["", ""],
        ["EXECUTIVE SUMMARY", ""],
        ["Total Records Processed", total_records],
        ["Clips Found", clips_found],
        ["Not Found/Rejected", clips_not_found],
        ["Success Rate", f"{(clips_found/total_records*100):.1f}%" if total_records > 0 else "0%"],
        ["", ""],
        ["QUALITY METRICS", ""],
        ["Average Relevance Score", f"{avg_relevance:.1f}/10"],
        ["High Quality Clips (8+)", high_quality],
        ["Positive Sentiment", positive_sentiment]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Title row
                cell.font = Font(bold=True, size=16, color="366092")
            elif row_idx == 4 or (isinstance(value, str) and value.isupper()):  # Section headers
                cell.font = Font(bold=True, size=12, color="366092")
    
    # Auto-size columns
    for col in summary_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column].width = adjusted_width
    
    # 2. Detailed Results Sheet
    results_ws = wb.create_sheet("Detailed Results")
    
    # Use the same column names as Bulk Review (exclude Approve/Reject columns)
    # Include Activity_ID for approval workflow even though it's not visible in UI
    bulk_review_columns = [
        'Activity_ID', 'Office', 'WO #', 'Make', 'Model', 'Contact', 'Media Outlet', 
        'URL', 'Relevance', 'Sentiment'
    ]
    
    # Map our data columns to Bulk Review column names
    column_mapping = {
        'Activity_ID': 'Activity_ID',  # Activity_ID column exists but may be empty
        'Office': 'Office',
        'WO #': 'WO #',
        'Make': 'Make',
        'Model': 'Model',
        'To': 'Contact',
        'Affiliation': 'Media Outlet',
        'Clip URL': 'URL',  # Add the primary URL from the View column
        'Relevance Score': 'Relevance',
        'Overall Sentiment': 'Sentiment',  # Fix: Use the correct sentiment column
    }
    
    # FIX: Fetch Activity_ID values from external source if they're blank in the data
    activity_id_mapping = {}
    try:
        import requests
        response = requests.get("https://reports.driveshop.com/?report=file:/home/deployer/reports/clips/media_loans_without_clips.rpt&init=csv", timeout=30)
        if response.status_code == 200:
            source_lines = response.text.strip().split('\n')
            for line in source_lines:
                if line.strip() and not line.startswith('"Activity_ID"'):  # Skip header
                    # Parse CSV line properly (handle quoted fields)
                    import csv
                    from io import StringIO
                    reader = csv.reader(StringIO(line))
                    parts = next(reader)
                    if len(parts) >= 5:
                        # Position mapping: Activity_ID(1st), Person_ID(2nd), Make(3rd), Model(4th), WO#(5th), ...
                        activity_id = parts[0].strip()  # Activity_ID is in 1st position
                        wo_number = parts[4].strip()    # WO# is in 5th position
                        activity_id_mapping[wo_number] = activity_id
        print(f"✅ Fetched Activity_ID mapping for {len(activity_id_mapping)} WO# records")
    except Exception as e:
        print(f"Warning: Could not fetch Activity_ID mapping: {e}")
    
    # Create export dataframe with Bulk Review column structure
    export_df = pd.DataFrame()
    
    for bulk_col in bulk_review_columns:
        # Find the corresponding column in our data
        source_col = None
        for data_col, bulk_name in column_mapping.items():
            if bulk_name == bulk_col and data_col in df.columns:
                source_col = data_col
                break
        
        if source_col:
            if bulk_col == 'Activity_ID' and activity_id_mapping:
                # FIX: Populate Activity_ID from external mapping using WO#
                export_df[bulk_col] = df['WO #'].astype(str).map(activity_id_mapping).fillna(df[source_col])
            else:
                export_df[bulk_col] = df[source_col]
        else:
            # Fill with empty if column doesn't exist
            export_df[bulk_col] = ''
    
    # Add header row
    headers = list(export_df.columns)
    results_ws.append(headers)
    
    # Add data rows with clickable URLs and formatted sentiment
    for idx, row in export_df.iterrows():
        row_data = []
        for col_idx, (col_name, value) in enumerate(row.items()):
            # Format sentiment with abbreviations and emojis
            if col_name == 'Sentiment' and value:
                sentiment_map = {
                    'positive': 'POS 😊',
                    'negative': 'NEG 😞',
                    'neutral': 'NEU 😐',
                    'pos': 'POS 😊',
                    'neg': 'NEG 😞', 
                    'neu': 'NEU 😐'
                }
                # Clean the value and try multiple formats
                cleaned_value = str(value).lower().strip()
                formatted_value = sentiment_map.get(cleaned_value, f"{cleaned_value} 😐")
                row_data.append(formatted_value)

            else:
                row_data.append(value)
        results_ws.append(row_data)
    
    # Style the header row
    for cell in results_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Make URLs clickable and style data rows
    from openpyxl.styles import Font as XLFont
    url_font = XLFont(color="0000FF", underline="single")
    
    for row_idx in range(2, results_ws.max_row + 1):  # Start from row 2 (after header)
        for col_idx, col_name in enumerate(headers, 1):
            cell = results_ws.cell(row=row_idx, column=col_idx)
            
            # Make URLs clickable
            if col_name == 'URL' and cell.value and str(cell.value).startswith(('http://', 'https://')):
                cell.hyperlink = str(cell.value)
                cell.font = url_font

    
    # Auto-size columns
    for col in results_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        results_ws.column_dimensions[column].width = adjusted_width
    
    # 3. Rejected Loans Sheet (if rejected data is available)
    if rejected_df is not None and len(rejected_df) > 0:
        rejected_ws = wb.create_sheet("Rejected Loans")
        
        # Define columns for rejected loans report
        rejected_columns = [
            'WO #', 'Model', 'Contact', 'Media Outlet', 'Rejection Reason', 'URLs Searched', 'Details'
        ]
        
        # Create rejected export dataframe
        rejected_export_df = pd.DataFrame()
        
        # Map rejected data to export columns
        rejected_export_df['WO #'] = rejected_df['WO #']
        rejected_export_df['Model'] = rejected_df['Model']
        rejected_export_df['Contact'] = rejected_df['To']
        rejected_export_df['Media Outlet'] = rejected_df['Affiliation']
        rejected_export_df['Rejection Reason'] = rejected_df['Rejection_Reason']
        
        # Extract the actual URLs from the URL_Details field
        def extract_urls_from_details(url_details):
            """Extract just the URLs from the detailed rejection info"""
            if pd.isna(url_details) or not url_details:
                return "No URLs processed"
            
            urls = []
            # Split by semicolon and extract URLs
            for detail in str(url_details).split(';'):
                detail = detail.strip()
                if detail.startswith(('http://', 'https://')):
                    # Find the colon that separates URL from description (not the :// in https)
                    # Look for ": " (colon followed by space) which indicates the description starts
                    if ': ' in detail:
                        url_part = detail.split(': ')[0].strip()
                        urls.append(url_part)
                    else:
                        # If no description separator found, take the whole thing as URL
                        urls.append(detail)
            
            return '\n'.join(urls) if urls else "No URLs found"
        
        rejected_export_df['URLs Searched'] = rejected_df['URL_Details'].apply(extract_urls_from_details)
        rejected_export_df['Details'] = rejected_df['URL_Details']
        
        # Add header row
        rejected_headers = list(rejected_export_df.columns)
        rejected_ws.append(rejected_headers)
        
        # Add data rows
        for idx, row in rejected_export_df.iterrows():
            row_data = []
            for col_name, value in row.items():
                # Handle URL columns by making them clickable if they contain URLs
                if col_name == 'URLs Searched' and value and 'http' in str(value):
                    row_data.append(str(value))
                else:
                    row_data.append(value)
            rejected_ws.append(row_data)
        
        # Style the header row with a red theme for rejected items
        rejected_header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        for cell in rejected_ws[1]:
            cell.font = header_font
            cell.fill = rejected_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Make URLs clickable in the URLs Searched column
        urls_searched_col_idx = rejected_headers.index('URLs Searched') + 1
        for row_idx in range(2, rejected_ws.max_row + 1):
            cell = rejected_ws.cell(row=row_idx, column=urls_searched_col_idx)
            if cell.value and 'http' in str(cell.value):
                urls_text = str(cell.value)
                # Check if there are multiple URLs (newline separated)
                if '\n' in urls_text:
                    # Multiple URLs - make text blue and add line breaks for readability
                    cell.font = XLFont(color="0000FF")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    # For the first URL, we can still make it a hyperlink
                    first_url = urls_text.split('\n')[0].strip()
                    if first_url.startswith(('http://', 'https://')):
                        cell.hyperlink = first_url
                        cell.font = url_font
                else:
                    # Single URL can be a hyperlink
                    url = urls_text.strip()
                    if url.startswith(('http://', 'https://')):
                        cell.hyperlink = url
                        cell.font = url_font
        
        # Auto-size columns for rejected sheet
        for col in rejected_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # Wider for rejection details
            rejected_ws.column_dimensions[column].width = adjusted_width
    
    # 4. Approved Clips Sheet (filter to match current dataset only)
    if approved_df is not None and len(approved_df) > 0:
        # Filter approved clips to only include WO #s that exist in the current detailed results
        current_wo_numbers = set(df['WO #'].astype(str))
        current_approved_df = approved_df[approved_df['WO #'].astype(str).isin(current_wo_numbers)].copy()
        
        if not current_approved_df.empty:
            # FIX: Get actual loan end dates AND Activity_IDs from source data for Excel
            source_mapping = {}
            activity_id_mapping = {}
            try:
                import requests
                response = requests.get("https://reports.driveshop.com/?report=file:/home/deployer/reports/clips/media_loans_without_clips.rpt&init=csv", timeout=30)
                if response.status_code == 200:
                    source_lines = response.text.strip().split('\n')
                    for line in source_lines:
                        if line.strip() and not line.startswith('"Activity_ID"'):  # Skip header
                            # Parse CSV line properly (handle quoted fields)
                            import csv
                            from io import StringIO
                            reader = csv.reader(StringIO(line))
                            parts = next(reader)
                            if len(parts) >= 10:
                                # Position mapping: Activity_ID(1st), Person_ID(2nd), Make(3rd), Model(4th), WO#(5th), ..., Stop_Date(10th)
                                activity_id = parts[0].strip()  # Activity_ID is in 1st position
                                wo_number = parts[4].strip()    # WO# is in 5th position
                                stop_date = parts[9].strip()    # Stop Date is in 10th position
                                source_mapping[wo_number] = stop_date
                                activity_id_mapping[wo_number] = activity_id
                print(f"✅ Fetched source data mapping for {len(source_mapping)} WO# records")
            except Exception as e:
                print(f"Warning: Could not fetch source data for loan end dates and Activity_IDs: {e}")
            
            # FIX: Populate Activity_ID from external source data using WO# mapping
            if activity_id_mapping:
                current_approved_df['Activity_ID'] = current_approved_df['WO #'].astype(str).map(activity_id_mapping)
            elif 'Article_ID' in current_approved_df.columns:
                current_approved_df['Activity_ID'] = current_approved_df['Article_ID']
                current_approved_df.drop('Article_ID', axis=1, inplace=True)
            elif 'Activity_ID' not in current_approved_df.columns:
                # If neither exists, create empty Activity_ID column
                current_approved_df['Activity_ID'] = ''
            
            # FIX: Clean up date columns - keep only ONE published date and add loan end date
            # Remove the confusing 'published_date' column (it's not the loan end date)
            if 'published_date' in current_approved_df.columns:
                current_approved_df.drop('published_date', axis=1, inplace=True)
            
            # Add actual loan end dates from source data
            current_approved_df['Loan End Date'] = current_approved_df['WO #'].astype(str).map(source_mapping).fillna('')
            
            # Rename the Published Date column to be clear about what it is
            column_renames = {}
            if 'Published Date' in current_approved_df.columns:
                column_renames['Published Date'] = 'Article Published Date'
            
            if column_renames:
                current_approved_df.rename(columns=column_renames, inplace=True)
            
            # Move Activity_ID to first column
            if 'Activity_ID' in current_approved_df.columns:
                cols = current_approved_df.columns.tolist()
                cols.remove('Activity_ID')
                current_approved_df = current_approved_df[['Activity_ID'] + cols]
            
            # Create Approved Clips sheet using openpyxl syntax (not xlsxwriter)
            approved_sheet = wb.create_sheet('Approved Clips')
            
            # Write headers with openpyxl formatting (using already imported Font)
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(current_approved_df.columns, 1):
                cell = approved_sheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data using openpyxl
            url_font = Font(color="0000FF", underline="single")
            
            for row_num, (_, row_data) in enumerate(current_approved_df.iterrows(), 2):  # Start from row 2
                for col_num, value in enumerate(row_data, 1):
                    cell = approved_sheet.cell(row=row_num, column=col_num)
                    
                    if pd.isna(value):
                        cell.value = ''
                    elif isinstance(value, (int, float)):
                        cell.value = value
                    else:
                        # Handle URLs with hyperlinks
                        str_value = str(value)
                        if str_value.startswith(('http://', 'https://')):
                            cell.value = str_value
                            cell.hyperlink = str_value
                            cell.font = url_font
                        else:
                            cell.value = str_value
            
            # Auto-size columns
            for col in approved_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                approved_sheet.column_dimensions[column].width = adjusted_width
    
    return wb

# Initialize session state for popup management
if 'show_url_popup' not in st.session_state:
    st.session_state.show_url_popup = False
if 'popup_data' not in st.session_state:
    st.session_state.popup_data = {}

# Page configuration
st.set_page_config(
    page_title="DriveShop Clip Tracking",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Apply custom dark charcoal sidebar styling
apply_custom_sidebar_styling()

# Add logo to sidebar
with st.sidebar:
    # DriveShop Logo at the top - use Streamlit's built-in image display
    try:
        # Debug current working directory
        print(f"Current working directory: {os.getcwd()}")
        print(f"Project root: {project_root}")
        
        # Try multiple possible paths
        possible_paths = [
            os.path.join(project_root, "assets", "DriveShop_WebLogo.png"),
            "assets/DriveShop_WebLogo.png",
            "./assets/DriveShop_WebLogo.png",
            "DriveShop_WebLogo.png"
        ]
        
        logo_loaded = False
        for logo_path in possible_paths:
            if 'logo_logged' not in st.session_state:
                print(f"Trying logo path: {logo_path}")
                print(f"Path exists: {os.path.exists(logo_path)}")
            if os.path.exists(logo_path):
                try:
                    logo = Image.open(logo_path)
                    st.image(logo, width=180)
                    if 'logo_logged' not in st.session_state:
                        print(f"✅ Logo loaded successfully from: {logo_path}")
                        st.session_state.logo_logged = True
                    logo_loaded = True
                    break
                except Exception as img_error:
                    print(f"❌ Failed to load image from {logo_path}: {img_error}")
                    continue
        
        if not logo_loaded:
            print("❌ No logo paths worked, using text fallback")
            st.markdown("**DriveShop**")
            
    except Exception as e:
        print(f"Logo loading error: {e}")
        st.markdown("**DriveShop**")

# Check authentication and refresh session if needed
# Session timeout is configurable via SESSION_TIMEOUT_HOURS environment variable
# Increased default to 48 hours to reduce logout frequency
session_timeout = int(os.environ.get('SESSION_TIMEOUT_HOURS', '48'))
# Only check session periodically, not on every page interaction
if not auth.check_and_refresh_session(session_timeout_hours=session_timeout):
    # Apply dark background for login page
    st.markdown("""
        <style>
        /* Dark background for login page */
        .stApp {
            background-color: #0e1117;
        }
        </style>
        """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        # Display logo
        logo_path = Path(__file__).parent.parent.parent / "docs" / "assets" / "Logo.png"
        if logo_path.exists():
            try:
                logo = Image.open(logo_path)
                st.image(logo, width=300)
            except Exception:
                st.markdown("# DriveShop", unsafe_allow_html=True)
        else:
            st.markdown("# DriveShop", unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Login form with white background and black border
        st.markdown("""
            <div style="background-color: white; padding: 2rem; border-radius: 10px; border: 1px solid black; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);">
                <h3 style="color: black; text-align: center; margin-bottom: 1rem;">Clip Tracking Dashboard</h3>
                <h4 style="color: black; text-align: center; margin-bottom: 2rem;">Sign In</h4>
            </div>
        """, unsafe_allow_html=True)
        
        with st.container():
            st.markdown("""
                <style>
                div[data-testid="stForm"] {
                    background-color: white;
                    padding: 2rem;
                    border-radius: 10px;
                    border: 1px solid black;
                    margin-top: -3rem;
                    padding-top: 3rem;
                }
                </style>
            """, unsafe_allow_html=True)
            
            with st.form("login_form", clear_on_submit=False):
                email = st.text_input("Email", placeholder="admin@driveshop.com")
                password = st.text_input("Password", type="password", placeholder="Enter password")
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Use stylable_container to make the button blue
                with stylable_container(
                    "blue_button",
                    css_styles="""
                    button {
                        background-color: #5A8FDB !important;
                        color: white !important;
                        border: 2px solid #4A7BC7 !important;
                        border-radius: 6px !important;
                        font-weight: 600 !important;
                        transition: all 0.3s ease !important;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;
                    }
                    button:hover {
                        background-color: #4A7BC7 !important;
                        color: white !important;
                        border-color: #3A6BB7 !important;
                        transform: translateY(-1px) !important;
                        box-shadow: 0 4px 8px rgba(0,0,0,0.25) !important;
                    }
                    """
                ):
                    submitted = st.form_submit_button("Login", use_container_width=True, type="primary")
                
                if submitted:
                    if email and password:
                        success, error = auth.login(email, password)
                        if success:
                            st.success("Login successful!")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error(error or "Invalid credentials")
                    else:
                        st.error("Please enter both email and password")
    
    # Stop here if not authenticated
    st.stop()

# Main application (only runs if authenticated)
with st.sidebar:
    if st.button("Logout", type="secondary", use_container_width=True):
        auth.logout()
        st.rerun()
    
    user = auth.get_current_user()
    if user:
        # Persist user email in session state for job attribution
        try:
            user_email = getattr(user, 'email', None) or (user.get('email') if isinstance(user, dict) else None)
        except Exception:
            user_email = None
        if user_email:
            st.session_state['user_email'] = user_email
            st.markdown(f"**Logged in as:** {user_email}")
        else:
            st.markdown("**Logged in**")
    
    st.markdown("---")

st.title("DriveShop Clip Tracking Dashboard")

# Custom CSS for better styling
st.markdown("""
<style>
    /* Force light mode appearance */
    :root {
        color-scheme: light !important;
    }
    
    html[data-theme="dark"] {
        color-scheme: light !important;
    }
    
    .stApp {
        background-color: #ffffff !important;
        color: #000000 !important;
    }
    
    /* Tab styling - Active tab colored, inactive tabs off-white */
    .stTabs [data-baseweb="tab-list"] {
        background-color: transparent !important;
        gap: 4px;
        padding: 0 8px;
    }
    
    /* Base tab styling - All tabs start with off-white */
    .stTabs [data-baseweb="tab"] {
        height: 45px !important;
        padding-left: 24px !important;
        padding-right: 24px !important;
        border-radius: 8px 8px 0 0 !important;
        border: none !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
        margin-bottom: 0 !important;
        background-color: #f5f5f5 !important;
        color: #666666 !important;
    }
    
    /* Hover effect for inactive tabs */
    .stTabs [data-baseweb="tab"]:not([aria-selected="true"]):hover {
        background-color: #e8e8e8 !important;
        color: #333333 !important;
    }
    
    /* Active tab styling - Blue with white text */
    .stTabs [aria-selected="true"] {
        background-color: #1976d2 !important;
        color: #ffffff !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1) !important;
    }
    
    /* Active tab text color */
    .stTabs [aria-selected="true"] p {
        color: #ffffff !important;
    }
    
    /* Inactive tab text color */
    .stTabs [aria-selected="false"] p {
        color: #666666 !important;
    }
    
    /* Remove the default bottom border indicator */
    .stTabs [data-baseweb="tab-highlight"] {
        background-color: transparent !important;
    }
    
    /* Tab text styling */
    .stTabs [data-baseweb="tab"] p {
        font-size: 14px !important;
        margin: 0 !important;
        line-height: 1.5 !important;
    }
    
    /* ULTRA COMPACT layout - maximum table space */
    .main > div {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', 'Helvetica Neue', Arial, sans-serif;
        padding-top: 0.5rem !important;
    }
    
    .block-container {
        padding-top: 0.5rem !important;
        padding-bottom: 0.5rem !important;
        max-width: 100% !important;
    }
    
    /* Compact main title */
    h1 {
        font-size: 1.4rem !important;
        font-weight: 600 !important;
        line-height: 1.2 !important;
        margin-bottom: 0.3rem !important;
        margin-top: 0 !important;
        padding: 0 !important;
        color: #1a1a1a !important;
    }
    
    /* Ensure title container has space */
    .stApp > header {
        background-color: transparent;
    }
    
    /* Compact table header styling */
    .clip-table-header {
        font-size: 0.7rem !important;
        font-weight: 600 !important;
        color: #5a6c7d !important;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        background-color: #f8f9fa;
        padding: 0.4rem 0.3rem;
        border-bottom: 2px solid #adb5bd;
        text-align: center;
    }
    
    /* ULTRA COMPACT metrics */
    div[data-testid="metric-container"] {
        background-color: #f8f9fa !important;
        border: 1px solid #e9ecef !important;
        padding: 0.2rem 0.3rem !important;
        border-radius: 0.2rem !important;
        margin: 0.1rem 0 !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
        height: auto !important;
        min-height: 2rem !important;
        max-height: 2.5rem !important;
    }
    
    /* Compact tabs */
    .stTabs {
        margin: 0.2rem 0 !important;
    }
    
    .stTabs > div > div > div > div {
        padding: 0.3rem 0.8rem !important;
        font-size: 0.9rem !important;
    }
    
    /* Remove excessive margins everywhere */
    .element-container {
        margin: 0.2rem 0 !important;
    }
    
    /* Compact columns */
    .row-widget {
        margin: 0.1rem 0 !important;
    }
    
    /* Remove divider spacing */
    hr {
        margin: 0.3rem 0 !important;
    }
    
    /* Make everything more compact */
    .stMarkdown {
        margin: 0.1rem 0 !important;
    }
    
    /* Compact table area */
    .ag-theme-alpine {
        font-size: 0.8rem !important;
    }
    
    /* Sticky Action Bar for Bulk Review */
    .sticky-action-bar {
        position: fixed;
        bottom: 0;
        left: 0;
        right: 0;
        background: linear-gradient(135deg, #000000 0%, #2c3e50 100%);
        padding: 0.8rem 1rem;
        box-shadow: 0 -4px 20px rgba(0,0,0,0.3);
        border-top: 2px solid #333;
        z-index: 1000;
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 1rem;
    }
    
    .sticky-action-bar .stButton > button {
        background: linear-gradient(135deg, #000000 0%, #333333 100%) !important;
        color: white !important;
        border: 2px solid #444 !important;
        border-radius: 8px !important;
        padding: 0.6rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3) !important;
        min-width: 140px !important;
    }
    
    .sticky-action-bar .stButton > button:hover {
        background: linear-gradient(135deg, #333333 0%, #555555 100%) !important;
        border-color: #666 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.4) !important;
    }
    
    /* Add bottom padding to main content to prevent overlap */
    .main .block-container {
        padding-bottom: 5rem !important;
    }
    
    /* Target ALL metric text elements but keep READABLE */
    div[data-testid="metric-container"] > div,
    div[data-testid="metric-container"] span,
    div[data-testid="metric-container"] p,
    [data-testid="metric-container"] * {
        font-size: 0.9rem !important;
        line-height: 1.2 !important;
        margin: 0.1rem 0 !important;
        padding: 0 !important;
    }
    
    /* Force label text to be small but readable */
    div[data-testid="metric-container"] > div:first-child,
    [data-testid="metric-container"] [data-testid="metric-label"],
    .metric-label {
        font-size: 0.75rem !important;
        font-weight: 500 !important;
        color: #6c757d !important;
        margin-bottom: 0.15rem !important;
        text-transform: uppercase !important;
        letter-spacing: 0.3px !important;
        line-height: 1.1 !important;
    }
    
    /* Force value text to be bigger and VISIBLE */
    div[data-testid="metric-container"] > div:last-child,
    [data-testid="metric-container"] [data-testid="metric-value"],
    .metric-value {
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        line-height: 1.2 !important;
        color: #2c3e50 !important;
    }
    
    /* Override Streamlit's default metric styling */
    .stMetric {
        padding: 0.15rem !important;
        margin: 0.15rem !important;
        height: auto !important;
        min-height: 2.5rem !important;
    }
    
    .stMetric > div {
        font-size: 0.9rem !important;
        line-height: 1.2 !important;
    }
    
    /* Nuclear option - target by text content if needed */
    div:contains("Total Clips"),
    div:contains("Avg Score"), 
    div:contains("High Quality"),
    div:contains("Approved") {
        font-size: 0.1rem !important;
    }
    
    /* Table row styling */
    .table-row {
        padding: 6px 4px;
        min-height: 1.8rem;
        display: flex;
        align-items: center;
        text-align: center;
        font-size: 0.7rem;
        line-height: 1.2;
    }
    
    /* URL popup styling */
    .url-popup {
        background-color: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    
    .url-item {
        padding: 0.25rem 0;
        border-bottom: 1px solid #e9ecef;
    }
    
    .url-item:last-child {
        border-bottom: none;
    }
    
    /* Better button targeting - all buttons in action columns */
    div[data-testid="column"]:last-child div[data-testid="stButton"]:nth-child(1) button {
        background: #28a745 !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        height: 1.8rem !important;
    }
    
    div[data-testid="column"]:last-child div[data-testid="stButton"]:nth-child(3) button {
        background: #dc3545 !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        height: 1.8rem !important;
    }
    
    /* More aggressive targeting for all checkmark buttons */
    button:contains("✓") {
        background-color: #28a745 !important;
        color: white !important;
    }
    
    /* More aggressive targeting for all X buttons */
    button:contains("✗") {
        background-color: #dc3545 !important;
        color: white !important;
    }
    
    /* Universal button override */
    .stButton button {
        font-size: 12px !important;
        font-weight: 600 !important;
        padding: 3px 12px !important;
        border: none !important;
        border-radius: 4px !important;
        height: 1.8rem !important;
    }
    
    /* Score color styling */
    .score-high { color: #28a745; font-weight: 600; }
    .score-med { color: #007bff; font-weight: 600; }
    .score-low { color: #ffc107; font-weight: 600; }
    
    /* Scrollable table styling */
    .scrollable-table {
        max-height: 450px;
        overflow-y: auto;
        border: 1px solid #e0e0e0;
        border-radius: 4px;
        background: white;
    }
    
    .fixed-header {
        position: sticky;
        top: 0;
        z-index: 100;
        background: white;
        border-bottom: 2px solid #adb5bd;
        padding: 0.2rem 0;
    }
    
    /* Smooth scrolling */
    .scrollable-table::-webkit-scrollbar {
        width: 8px;
    }
    
    .scrollable-table::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }
    
    .scrollable-table::-webkit-scrollbar-thumb {
        background: #c1c1c1;
        border-radius: 4px;
    }
    
    .scrollable-table::-webkit-scrollbar-thumb:hover {
        background: #a8a8a8;
    }
    
    /* Compact sidebar styling */
    .css-1d391kg {
        width: 250px !important;
        min-width: 250px !important;
    }
    
    /* Make sidebar content more compact */
    .css-1d391kg .stMarkdown p {
        margin: 0.2rem 0 !important;
        font-size: 0.9rem !important;
    }
    
    .css-1d391kg .stButton button {
        padding: 0.3rem 0.5rem !important;
        font-size: 0.8rem !important;
        margin: 0.1rem 0 !important;
    }
    
    .css-1d391kg .stFileUploader {
        margin: 0.2rem 0 !important;
    }
    
    .css-1d391kg .stFileUploader > div {
        padding: 0.3rem !important;
        font-size: 0.8rem !important;
    }
    
    /* Compact success/error messages in sidebar */
    .css-1d391kg .stAlert {
        padding: 0.3rem !important;
        font-size: 0.8rem !important;
        margin: 0.2rem 0 !important;
    }
</style>
""", unsafe_allow_html=True)

# --- MAPPING UPDATE FEATURE (SIDEBAR) ---
def update_person_outlets_mapping_from_url(url):
    """
    Download the mapping CSV from the given URL, validate, save, and regenerate the JSON mapping file.
    Returns (success: bool, message: str)
    """
    import requests
    import pandas as pd
    import json
    import os
    try:
        # Download CSV
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        csv_content = resp.content.decode('utf-8')

        # Define headers manually, as the source CSV doesn't contain them
        headers = ["Person_ID", "Reporter_Name", "Outlet_ID", "Outlet_Name", "Outlet_URL", "Impressions"]
        
        # Parse CSV without reading a header row and assign our defined headers
        df = pd.read_csv(io.StringIO(csv_content), header=None, names=headers)
        
        # Validate that the DataFrame now has the correct columns
        required_cols = {'Person_ID', 'Reporter_Name', 'Outlet_Name', 'Outlet_URL', 'Outlet_ID', 'Impressions'}
        if not required_cols.issubset(set(df.columns)):
            return False, f"Internal Error: Failed to assign correct columns. Please check the function."

        # Save CSV
        csv_path = os.path.join(project_root, 'data', 'person_outlets_mapping.csv')
        df.to_csv(csv_path, index=False)
        # Regenerate JSON mapping
        person_outlets = {}
        for _, row in df.iterrows():
            person_id = str(row['Person_ID'])
            reporter_name = str(row['Reporter_Name'])
            outlet_info = {
                'outlet_name': row['Outlet_Name'],
                'outlet_url': row['Outlet_URL'],
                'outlet_id': str(row['Outlet_ID']),
                'impressions': row['Impressions']
            }
            if person_id not in person_outlets:
                person_outlets[person_id] = {
                    'reporter_name': reporter_name,
                    'outlets': []
                }
            person_outlets[person_id]['outlets'].append(outlet_info)
        json_path = os.path.join(project_root, 'data', 'person_outlets_mapping.json')
        with open(json_path, 'w') as f:
            json.dump(person_outlets, f, indent=2)
        return True, f"Mapping updated successfully! {len(person_outlets)} Person_IDs, {len(df)} outlet relationships."
    except Exception as e:
        return False, f"Error updating mapping: {e}"

# --- SIDEBAR UI ---
with st.sidebar:
    # Display mapping update message if it exists in session state
    if 'mapping_update_msg' in st.session_state:
        success, msg = st.session_state.mapping_update_msg
        if success:
            st.success(msg)
        else:
            st.error(msg)
        # Clear the message after displaying it
        del st.session_state.mapping_update_msg

    st.markdown("**🔄 Update Person-Outlets Mapping**")
    default_mapping_url = "https://reports.driveshop.com/?report=file:%2Fhome%2Fdeployer%2Freports%2Fclips%2Fmedia_outlet_list.rpt&init=csv&exportreportdataonly=true&columnnames=true"
    mapping_url = st.text_input(
        "Paste mapping CSV URL here:",
        value=default_mapping_url,
        help="Paste the direct link to the latest mapping CSV."
    )
    if st.button("Update Mapping", use_container_width=True):
        with st.spinner("Updating mapping from URL..."):
            success, msg = update_person_outlets_mapping_from_url(mapping_url)
            st.session_state.mapping_update_msg = (success, msg)
            if success:
                st.cache_data.clear()  # Clear cache so new mapping data is loaded
                st.session_state.outlet_data_mapping = {}  # Clear stale outlet lookup data
            st.rerun()

    # Thin separator line
    st.markdown('<hr style="margin: 1rem 0; border: none; height: 1px; background-color: #666666;">', unsafe_allow_html=True)
    
    st.markdown("### Process Loans from Live URL")
    
    loans_url = st.text_input(
        "Live 'Loans without Clips' URL:",
        "https://reports.driveshop.com/?report=file:/home/deployer/reports/clips/media_loans_without_clips.rpt&init=csv"
    )
    
    st.markdown("### Filters")

    # This input is now the first filter for a clearer workflow.
    limit_records = st.number_input(
        "Limit records to process (0 for all):",
        min_value=0,
        value=0,  # Default to 0, which means "All"
        step=1,
        help="Set the maximum number of records to process from the filtered set. Set to 0 to process all."
    )
    
    offices = ['All Offices']
    makes = ['All Makes']
    reporter_names = ['All Reporters']
    outlets = ['All Outlets']
    name_to_id_map = create_reporter_name_to_id_mapping()

    if 'loans_data_loaded' in st.session_state and st.session_state.loans_data_loaded:
        df = st.session_state.get('loaded_loans_df')
        if df is not None:
            if 'Office' in df.columns:
                offices += sorted(df['Office'].dropna().astype(str).unique().tolist())
            if 'Make' in df.columns:
                makes += sorted(df['Make'].dropna().astype(str).unique().tolist())
    
    mapping_file = os.path.join(project_root, "data", "person_outlets_mapping.csv")
    if os.path.exists(mapping_file):
        try:
            mapping_df = pd.read_csv(mapping_file)
            if 'Reporter_Name' in mapping_df.columns:
                reporter_names += sorted(mapping_df['Reporter_Name'].dropna().unique().tolist())
            if 'Outlet_Name' in mapping_df.columns:
                outlets += sorted(mapping_df['Outlet_Name'].dropna().unique().tolist())
        except Exception as e:
            st.warning(f"Could not load reporter/outlet names: {e}")

    selected_office = st.selectbox("Filter by Office:", offices)
    selected_make = st.selectbox("Filter by Make:", makes)
    selected_reporter_name = st.selectbox("Filter by Reporter:", reporter_names)
    selected_outlet = st.selectbox("Filter by Media Outlet:", outlets)
    
    # Add WO # and Activity ID filters with multiple value support
    wo_number_filter = st.text_input(
        "Filter by WO # (optional):",
        value="",
        help="Enter Work Order number(s). Use commas to separate multiple values (e.g., 1182796,1182884,1182887)"
    )
    
    activity_id_filter = st.text_input(
        "Filter by Activity ID (optional):",
        value="",
        help="Enter Activity ID(s). Use commas to separate multiple values (e.g., 1114558,1114646,1114649)"
    )
    
    # Show batch processing info if available
    suggested_value = ""
    if 'batch_info' in st.session_state:
        info = st.session_state.batch_info
        st.success(f"""
        **📊 Last Batch Completed at {info['timestamp']}:**
        - **Last Activity ID:** {info['last_processed_id']}
        - **Records Processed:** {info['records_processed']}
        - **Completed At:** {info['timestamp']}
        """)
        
        # Add a button to auto-fill the next Activity ID
        if st.button("📋 Use Suggested ID for Next Batch", key="use_suggested_id", help="Auto-fill the suggested Activity ID"):
            st.session_state.suggested_id_to_use = info['next_suggested_id']
            st.rerun()
        
        # Check if we should use the suggested value
        if 'suggested_id_to_use' in st.session_state:
            suggested_value = st.session_state.suggested_id_to_use
            del st.session_state.suggested_id_to_use
    
    # Date range filter for Loan Start Date
    st.markdown("**📅 Filter by Loan Start Date Range**")
    date_col1, date_col2 = st.columns(2)
    
    with date_col1:
        start_date_filter = st.date_input(
            "From Date:",
            value=None,
            help="Select start date to filter loans (inclusive)",
            key="loan_start_date_from"
        )
    
    with date_col2:
        end_date_filter = st.date_input(
            "To Date:",
            value=None,
            help="Select end date to filter loans (inclusive)",
            key="loan_start_date_to"
        )
    
    # Position-based filter for batch processing (much simpler!)
    skip_records = st.number_input(
        "Skip first X records (optional):",
        min_value=0,
        value=0,
        step=1,
        help="Enter number of records to skip from the beginning. For example, enter 200 to start processing from record 201."
    )

    if 'loans_data_loaded' in st.session_state and st.session_state.loans_data_loaded:
        filtered_df = st.session_state.loaded_loans_df.copy()
        
        if selected_office != 'All Offices':
            filtered_df = filtered_df[filtered_df['Office'] == selected_office]
        
        if selected_make != 'All Makes':
            filtered_df = filtered_df[filtered_df['Make'] == selected_make]
        
        if selected_reporter_name != 'All Reporters':
            person_id = name_to_id_map.get(selected_reporter_name)
            if person_id:
                # This includes the fix for the Person_ID data type issue
                filtered_df['Person_ID'] = pd.to_numeric(filtered_df['Person_ID'], errors='coerce').astype('Int64').astype(str)
                filtered_df = filtered_df[filtered_df['Person_ID'] == person_id]
        
        # Apply WO # filter if specified (supports multiple comma-separated values)
        if wo_number_filter.strip():
            if 'WO #' in filtered_df.columns:
                # Parse comma-separated values and clean them
                wo_numbers = [wo.strip() for wo in wo_number_filter.split(',') if wo.strip()]
                filtered_df = filtered_df[filtered_df['WO #'].astype(str).isin(wo_numbers)]
                st.info(f"🎯 Filtering by {len(wo_numbers)} WO #(s): {', '.join(wo_numbers)}")
            else:
                st.warning("⚠️ WO # column not found in data")
        
        # Apply Activity ID filter if specified (supports multiple comma-separated values)
        if activity_id_filter.strip():
            if 'Activity_ID' in filtered_df.columns:
                # Parse comma-separated values and clean them
                activity_ids = [aid.strip() for aid in activity_id_filter.split(',') if aid.strip()]
                filtered_df = filtered_df[filtered_df['Activity_ID'].astype(str).isin(activity_ids)]
                st.info(f"🎯 Filtering by {len(activity_ids)} Activity ID(s): {', '.join(activity_ids)}")
            else:
                st.warning("⚠️ Activity_ID column not found in data")
        
        # Apply date range filter for Loan Start Date
        if start_date_filter or end_date_filter:
            if 'Start Date' in filtered_df.columns:
                # Convert Start Date column to datetime
                filtered_df['Start Date'] = pd.to_datetime(filtered_df['Start Date'], errors='coerce')
                
                # Apply start date filter
                if start_date_filter:
                    filtered_df = filtered_df[filtered_df['Start Date'] >= pd.Timestamp(start_date_filter)]
                    
                # Apply end date filter  
                if end_date_filter:
                    filtered_df = filtered_df[filtered_df['Start Date'] <= pd.Timestamp(end_date_filter)]
                
                # Show date range info
                date_range_info = []
                if start_date_filter:
                    date_range_info.append(f"from {start_date_filter}")
                if end_date_filter:
                    date_range_info.append(f"to {end_date_filter}")
                st.info(f"📅 Filtering by Loan Start Date {' '.join(date_range_info)}")
            else:
                st.warning("⚠️ Start Date column not found in data")
        
        # Apply position-based filtering (skip first X records)
        if skip_records > 0:
            if skip_records < len(filtered_df):
                filtered_df = filtered_df.iloc[skip_records:].reset_index(drop=True)
                st.info(f"📍 Skipping first {skip_records} records, starting from position {skip_records + 1}")
            else:
                st.warning(f"Skip value ({skip_records}) is greater than available records ({len(filtered_df)}). Processing all records.")
        
        # Clarify how many records will be processed based on filters and the limit
        num_filtered = len(filtered_df)
        limit_text = f"Up to {limit_records} of these will be processed." if limit_records > 0 else "All of these will be processed."
        st.info(f"**{num_filtered}** records match your filters. {limit_text}")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Load Data", key='load_data_for_filtering'):
            with st.spinner("Loading data from URL..."):
                success, df, data_info = load_loans_data_for_filtering(loans_url)
                if success:
                    st.session_state.loans_data_loaded = True
                    st.session_state.loaded_loans_df = df
                    st.session_state.loans_data_info = data_info
                    st.success(f"✅ Loaded {data_info['total_records']} records. Ready to filter.")
                    st.rerun()
                else:
                    st.error("❌ Failed to load data.")
    
    with col2:
        if st.button("Process Filtered", key='process_from_url_filtered'):
            # Only proceed if data has been loaded and filtered
            if 'filtered_df' in locals() and not filtered_df.empty:
                from datetime import datetime
                
                # Convert filtered dataframe to list of records
                records_to_process = filtered_df.to_dict('records')

                # Remap dataframe columns to the format the backend expects
                remapped_records = []
                for record in records_to_process:
                    urls = []
                    if 'Links' in record and pd.notna(record['Links']):
                        urls = [url.strip() for url in str(record['Links']).split(',') if url.strip()]

                    # Clean up any NaN/Infinity values that can't be serialized to JSON
                    def clean_value(val):
                        if pd.isna(val) or (isinstance(val, float) and (val != val or val == float('inf') or val == float('-inf'))):
                            return None
                        return val

                    remapped_records.append({
                        'work_order': clean_value(record.get('WO #')),
                        'model': clean_value(record.get('Model')),
                        'model_short': clean_value(record.get('Model Short Name')),
                        'to': clean_value(record.get('To')),
                        'affiliation': clean_value(record.get('Affiliation')),
                        'urls': urls,
                        'start_date': str(record.get('Start Date')) if pd.notna(record.get('Start Date')) else None,
                        'make': clean_value(record.get('Make')),
                        'activity_id': clean_value(record.get('ActivityID')),
                        'person_id': clean_value(record.get('Person_ID')),
                        'office': clean_value(record.get('Office'))
                    })
                
                # Get user email for job tracking
                user_email = st.session_state.get('user_email', 'System')
                # Note: Jobs created by System show they were submitted via UI
                # Jobs with 'anonymous' are a bug from worker claiming
                
                # Create job parameters - don't include the full records, just the URL and filters
                job_params = {
                    'url': loans_url,
                    'filters': {
                        'office': selected_office if 'selected_office' in locals() else 'All',
                        'make': selected_make if 'selected_make' in locals() else 'All',
                        'reporter': selected_reporter_name if 'selected_reporter_name' in locals() else 'All',
                        'wo_numbers': wo_number_filter if 'wo_number_filter' in locals() else '',
                        'activity_ids': activity_id_filter if 'activity_id_filter' in locals() else '',
                        'skip_records': skip_records if 'skip_records' in locals() else 0,
                        # Include date range for worker-side filtering
                        'date_from': str(start_date_filter) if 'start_date_filter' in locals() and start_date_filter else None,
                        'date_to': str(end_date_filter) if 'end_date_filter' in locals() and end_date_filter else None
                    },
                    'limit': limit_records if 'limit_records' in locals() else 0,
                    'record_count': len(remapped_records)  # Just store the count for info
                }
                
                # Create run name
                run_name = f"CSV Process - {len(remapped_records)} records - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                
                try:
                    # Submit job to queue
                    job_id = submit_job_to_queue(
                        job_type='csv_upload',
                        job_params=job_params,
                        run_name=run_name,
                        user_email=user_email
                    )
                    
                    st.success(f"""
                    ✅ **Job submitted successfully!**
                    
                    Job ID: `{job_id[:8]}...`
                    
                    Navigate to the **"🚀 Active Jobs"** tab to monitor progress.
                    """)
                            
                except Exception as e:
                    st.error(f"❌ Failed to submit job: {str(e)}")
                    logger.error(f"Job submission failed: {e}", exc_info=True)
            else:
                st.warning("No data loaded or no records match filters. Please load data first.")
    if 'loans_data_loaded' in st.session_state and st.session_state.loans_data_loaded:
        info = st.session_state.get('loans_data_info', {})
        st.markdown(f"📊 Data loaded: **{info.get('total_records', 0)}** total records, **{info.get('offices_count', 0)}** offices, **{info.get('makes_count', 0)}** makes")

    # Thin separator line
    st.markdown('<hr style="margin: 1rem 0; border: none; height: 1px; background-color: #666666;">', unsafe_allow_html=True)
    
    st.markdown("**📁 Process from File Upload**")
    uploaded_file = st.file_uploader("Upload Loans CSV/XLSX", type=['csv', 'xlsx'], label_visibility="collapsed")
    
    if uploaded_file is not None:
        temp_file_path = os.path.join(project_root, "data", "fixtures", "temp_upload.csv")
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        if st.button("Process Uploaded File", use_container_width=True):
            with st.spinner("Processing..."):
                success = run_ingest_database(input_file=temp_file_path)
                if success:
                    st.success("✅ Done!")
                    # Clear cache so Bulk Review shows new clips
                    st.cache_data.clear()
                    st.rerun() # Refresh the page
                else:
                    st.error("❌ Failed")
    
    # Thin separator line
    st.markdown('<hr style="margin: 1rem 0; border: none; height: 1px; background-color: #666666;">', unsafe_allow_html=True)

    if st.button("🔄 Process Default File (for testing)", use_container_width=True):
        with st.spinner("Processing default file..."):
            default_file = os.path.join(project_root, "data", "fixtures", "Loans_without_Clips.csv")
            success = run_ingest_database(input_file=default_file)
            if success:
                st.success("✅ Done!")
                # Clear cache so Bulk Review shows new clips
                st.cache_data.clear()
                st.rerun() # Refresh the page
            else:
                st.error("❌ Failed")

# Create tabs for different user workflows
bulk_review_tab, approved_queue_tab, active_jobs_tab, rejected_tab, analysis_tab, pullthrough_tab, oem_tab, reprocess_tab, cooldown_tab, export_tab = st.tabs([
    "Bulk Review",
    "Approved Queue",
    "Active Jobs",
    "Rejected/Issues",
    "Strategic Intelligence",
    "Message Pull-Through",
    "OEM Messaging",
    "Re-Process Historical",
    "Cooldown Mgmt",
    "Export"
])

# ========== CREATORIQ TAB ========== (REMOVED)

# ========== ACTIVE JOBS TAB ==========
with active_jobs_tab:
    display_active_jobs_tab()

# ========== BULK REVIEW TAB (Compact Interface) ==========
with bulk_review_tab:
    
    # Initialize session state for viewed records tracking
    if 'viewed_records' not in st.session_state:
        st.session_state.viewed_records = set()
    if 'total_records_count' not in st.session_state:
        st.session_state.total_records_count = 0
    
    # Initialize session state for approve/reject tracking (persist across refreshes)
    if 'approved_records' not in st.session_state:
        st.session_state.approved_records = set()
    if 'rejected_records' not in st.session_state:
        st.session_state.rejected_records = set()
    
    # Initialize Media Outlet, Byline, and Published Date tracking
    if 'last_saved_outlets' not in st.session_state:
        st.session_state.last_saved_outlets = {}
    if 'last_saved_bylines' not in st.session_state:
        st.session_state.last_saved_bylines = {}
    if 'last_saved_dates' not in st.session_state:
        st.session_state.last_saved_dates = {}
    # Add tracking for outlet data (id and impressions)
    if 'outlet_data_mapping' not in st.session_state:
        st.session_state.outlet_data_mapping = {}
    # Add tracking for edited URLs
    if 'edited_urls' not in st.session_state:
        st.session_state.edited_urls = {}
    
    # Load saved checkbox states from a temp file EARLY
    import pickle
    temp_file = os.path.join(project_root, "temp", "checkbox_state.pkl")
    
    # Ensure temp directory exists
    os.makedirs(os.path.dirname(temp_file), exist_ok=True)
    
    # Try to load saved state
    if os.path.exists(temp_file):
        try:
            with open(temp_file, 'rb') as f:
                saved_state = pickle.load(f)
                if 'approved' in saved_state:
                    st.session_state.approved_records.update(saved_state['approved'])
                if 'rejected' in saved_state:
                    st.session_state.rejected_records.update(saved_state['rejected'])
                if 'viewed' in saved_state:
                    st.session_state.viewed_records = saved_state['viewed']
                if 'outlets' in saved_state:
                    st.session_state.last_saved_outlets.update(saved_state['outlets'])
                if 'bylines' in saved_state:
                    st.session_state.last_saved_bylines.update(saved_state['bylines'])
                if 'outlet_data' in saved_state:
                    st.session_state.outlet_data_mapping.update(saved_state['outlet_data'])
        except Exception as e:
            print(f"Could not load saved checkbox state: {e}")
    
    # Add manual refresh button to control when data reloads
    col1, col2, col3 = st.columns([1, 1, 8])
    with col1:
        if st.button("🔄 Refresh Data", help="Manually refresh clips data from database"):
            st.cache_data.clear()
            st.session_state.outlet_data_mapping = {}  # Clear stale outlet lookup data
            st.rerun()
    
    # Cache database calls to improve performance
    @st.cache_data(ttl=300)  # Cache for 5 minutes to allow fresh data
    def cached_get_pending_clips():
        db = get_cached_database()
        return db.get_pending_clips()
    
    # Try to load results from database
    try:
        clips_data = cached_get_pending_clips()
        
        if clips_data:
            # Convert database results to DataFrame
            df = pd.DataFrame(clips_data)
            
            # Load UI states from database into session state
            # This ensures persistence across browser refreshes
            for clip in clips_data:
                wo_num = str(clip.get('wo_number', ''))
                if wo_num:
                    # Load viewed state
                    if clip.get('ui_viewed', False):
                        st.session_state.viewed_records.add(wo_num)
                    
                    # Load pending approval state
                    if clip.get('ui_approved_pending', False):
                        st.session_state.approved_records.add(wo_num)
                        st.session_state.selected_for_approval.add(wo_num)
                    
                    # Load pending rejection state
                    if clip.get('ui_rejected_pending', False):
                        st.session_state.rejected_records.add(wo_num)
                        st.session_state.selected_for_rejection.add(wo_num)
            
            # Map database fields to expected CSV format
            df = df.rename(columns={
                'wo_number': 'WO #',
                'make': 'Make',
                'model': 'Model',
                'contact': 'To',
                'office': 'Office',
                'person_id': 'Person_ID',  # Map database person_id to Person_ID for UI
                'clip_url': 'Clip URL',
                'relevance_score': 'Relevance Score',
                'status': 'Status',
                'tier_used': 'Processing Method',
                'published_date': 'Published Date',  # Map database field to expected name
                'media_outlet': 'Media Outlet',  # Map media_outlet to Media Outlet for dropdown
                'byline_author': 'Actual_Byline',  # Map database field to expected name
                'attribution_strength': 'Attribution_Strength'  # Map database field to expected name
            })
            
            # Set default values for missing columns  
            if 'Affiliation' not in df.columns:
                df['Affiliation'] = df.get('Media Outlet', 'N/A')
        else:
            df = pd.DataFrame()  # Empty DataFrame if no clips
    except Exception as e:
        st.error(f"❌ Error loading clips from database: {e}")
        df = pd.DataFrame()  # Empty DataFrame on error
    
    # Process database results
    if len(df) > 0:
        try:
            
            # Ensure WO # is treated as string
            if 'WO #' in df.columns:
                df['WO #'] = df['WO #'].astype(str)
            
            if not df.empty:
                # Update total records count for progress tracking
                st.session_state.total_records_count = len(df)
                
                # Show rejection success message if flagged
                if st.session_state.get('rejection_success', False):
                    st.success("✅ Record successfully rejected and moved to Rejected/Issues tab!")
                    st.session_state.rejection_success = False  # Clear the flag

                # Quick stats overview - only show relevant metrics for pending review
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Clips", len(df))
                with col2:
                    avg_score = df['Relevance Score'].mean() if 'Relevance Score' in df.columns and not df.empty else 0
                    st.metric("Avg Score", f"{avg_score:.1f}/10")
                with col3:
                    high_quality = len(df[df['Relevance Score'] >= 8]) if 'Relevance Score' in df.columns and not df.empty else 0
                    st.metric("High Quality", high_quality)
                
                
                # Display filtered results with AgGrid
                display_df = df.copy()
                
                # Create the table structure with the new column order
                clean_df = pd.DataFrame()
                
                # First add the Mark Viewed column (will be first in display)
                clean_df['👁️ Mark Viewed'] = display_df['WO #'].apply(lambda wo: str(wo) in st.session_state.viewed_records) if 'WO #' in display_df.columns else False
                
                # Then add columns in the requested order
                clean_df['WO #'] = display_df['WO #'] if 'WO #' in display_df.columns else ''
                clean_df['Contact'] = display_df['To'] if 'To' in display_df.columns else ''
                
                # Add hidden columns that are still needed
                clean_df['Office'] = display_df['Office'] if 'Office' in display_df.columns else 'N/A'
                
                # --- FIX: Use a name-to-ID mapping to get the correct numeric Person_ID ---
                reporter_name_to_id_map = create_reporter_name_to_id_mapping()
                
                # Use the 'Contact' column to look up the numeric Person_ID
                # NORMALIZE contact names to match the mapping (handle double spaces, case variations)
                def lookup_person_id(name):
                    if pd.isna(name) or not name:
                        return ''
                    # Normalize: strip, replace multiple spaces with single space, title case
                    normalized_name = str(name).strip().replace(r'\s+', ' ').title()
                    normalized_name = ' '.join(normalized_name.split())  # Extra normalization for multiple spaces
                    return reporter_name_to_id_map.get(normalized_name, '')
                
                # Use database Person_ID if available, otherwise lookup from contact name
                def get_person_id(row):
                    # First try to use the Person_ID from database
                    db_person_id = row.get('Person_ID', '')
                    if db_person_id and str(db_person_id).strip() and str(db_person_id) != 'nan':
                        return str(db_person_id)
                    # Fallback to lookup by contact name
                    return lookup_person_id(row.get('Contact', ''))
                
                clean_df['Person_ID'] = display_df.apply(get_person_id, axis=1)
                
                # Add Media Outlet column right after Contact (replacing Publication)
                # Smart matching: find the correct Outlet_Name from Person_outlets_mapping
                person_outlets_mapping = load_person_outlets_mapping()
                
                def smart_outlet_matching(row):
                    affiliation = str(row.get('Affiliation', ''))
                    person_id = str(row.get('Person_ID', ''))
                    
                    # Handle 'None' string explicitly
                    if not affiliation or not person_id or not person_outlets_mapping or affiliation == 'None' or affiliation == 'nan':
                        return ''  # Return empty for dropdown
                    
                    # Get available outlet options for this person
                    outlet_options = get_outlet_options_for_person(person_id, person_outlets_mapping)
                    if not outlet_options:
                        return ''
                    
                    print(f"🔍 Smart matching '{affiliation}' for Person_ID {person_id}")
                    print(f"   Available options: {outlet_options}")
                    
                    # Try exact match first
                    if affiliation in outlet_options:
                        print(f"✅ Exact match: '{affiliation}'")
                        return affiliation
                    
                    # Try fuzzy matching - check if outlet name is contained in affiliation
                    affiliation_lower = affiliation.lower().strip()
                    for outlet in outlet_options:
                        outlet_lower = outlet.lower().strip()
                        if outlet_lower in affiliation_lower:
                            print(f"🎯 Smart match: '{affiliation}' -> '{outlet}'")
                            return outlet
                    
                    print(f"❌ No match found for '{affiliation}'")
                    return ''  # Return empty if no match
                
                # Use database values if they exist, otherwise use smart matching
                if 'Media Outlet' in display_df.columns:
                    clean_df['Media Outlet'] = display_df.apply(
                        lambda row: row.get('Media Outlet', '') or smart_outlet_matching(row), 
                        axis=1
                    )
                else:
                    clean_df['Media Outlet'] = display_df.apply(smart_outlet_matching, axis=1)
                
                # Override with saved Media Outlet values from session state
                for idx, row in clean_df.iterrows():
                    wo_num = str(row.get('WO #', ''))
                    if wo_num in st.session_state.last_saved_outlets:
                        clean_df.at[idx, 'Media Outlet'] = st.session_state.last_saved_outlets[wo_num]
                
                # Add Make and Model columns
                clean_df['Make'] = display_df['Make'] if 'Make' in display_df.columns else ''
                clean_df['Model'] = display_df['Model'] if 'Model' in display_df.columns else ''
                
                # Format relevance score as "8/10" format
                if 'Relevance Score' in display_df.columns:
                    clean_df['Relevance'] = display_df['Relevance Score'].apply(lambda x: f"{x}/10" if pd.notna(x) and x != 'N/A' else 'N/A')
                else:
                    clean_df['Relevance'] = 'N/A'
                
                # REMOVED: Sentiment column - NEW database system uses cost-optimized GPT (relevance-only)
                # Sentiment analysis is not performed in the NEW system to save costs
                
                # Handle the URL for the View column
                url_column = None
                for col in ['Clip URL', 'url', 'final_url', 'Links']:
                    if col in display_df.columns:
                        url_column = col
                        break
                
                if url_column:
                    # ChatGPT's Alternative Solution: Create separate View column
                    clean_df['Clip URL'] = display_df[url_column]  # Keep raw URLs hidden
                    clean_df['📄 View'] = display_df[url_column]   # Copy URLs for cellRenderer
                else:
                    clean_df['Clip URL'] = 'No URL found'
                    clean_df['📄 View'] = 'No URL found'
                
                # ===== NEW: Add URL tracking columns =====
                
                # Add viewed status column for styling
                clean_df['Viewed'] = clean_df['WO #'].apply(lambda wo: str(wo) in st.session_state.viewed_records)
                
                # Add Published Date column - improved extraction with multiple fallbacks
                def get_published_date(row):
                    """Extract published date with multiple fallback methods"""
                    try:
                        # Method 1: Try Published Date field first
                        raw_date = row.get('Published Date', '')
                        if pd.notna(raw_date) and str(raw_date).strip() and str(raw_date).lower() not in ['nan', 'none', '']:
                            try:
                                import dateutil.parser
                                parsed_date = dateutil.parser.parse(str(raw_date))
                                return parsed_date.strftime('%m/%d/%y')
                            except:
                                pass
                        
                        # Method 2: Try to extract from URL if it contains date patterns
                        url = row.get('Clip URL', '')
                        if url:
                            import re
                            # Look for year/month/day patterns in URL
                            date_patterns = [
                                r'/(\d{4})/(\d{1,2})/(\d{1,2})/',  # /2024/12/25/
                                r'/(\d{4})-(\d{1,2})-(\d{1,2})',    # /2024-12-25
                                r'_(\d{4})(\d{2})(\d{2})',          # _20241225
                            ]
                            for pattern in date_patterns:
                                match = re.search(pattern, url)
                                if match:
                                    try:
                                        year, month, day = match.groups()
                                        from datetime import datetime
                                        date_obj = datetime(int(year), int(month), int(day))
                                        return date_obj.strftime('%m/%d/%Y')
                                    except:
                                        continue
                        
                        # Method 3: Use processed date as fallback if recent
                        processed_date = row.get('Processed Date', '')
                        if processed_date:
                            try:
                                import dateutil.parser
                                from datetime import datetime, timedelta
                                parsed_date = dateutil.parser.parse(str(processed_date))
                                # Only use if within last 30 days (likely recent article)
                                if (datetime.now() - parsed_date).days <= 30:
                                    return parsed_date.strftime('%m/%d/%Y')
                            except:
                                pass
                        
                        return "—"
                    except:
                        return "—"
                
                clean_df['📅 Published Date'] = display_df.apply(get_published_date, axis=1)
                
                # ===== NEW: Add Attribution Information columns =====
                def smart_attribution_analysis(row):
                    """Determine attribution strength using smart comparison logic"""
                    try:
                        contact_person = str(row.get('To', '')).strip()
                        actual_byline = str(row.get('Actual_Byline', '')).strip()
                        affiliation = str(row.get('Affiliation', '')).strip()
                        
                        # Clean up names for comparison (remove titles, normalize spaces)
                        import re
                        def normalize_name(name):
                            if not name or name.lower() in ['nan', 'none', '']:
                                return ''
                            # Remove common titles and normalize
                            name = re.sub(r'\b(Mr|Mrs|Ms|Dr|Prof|Sr|Jr)\.?\b', '', name, flags=re.IGNORECASE)
                            name = re.sub(r'\s+', ' ', name).strip()
                            return name.lower()
                        
                        contact_normalized = normalize_name(contact_person)
                        byline_normalized = normalize_name(actual_byline)
                        
                        # If we have a byline
                        if byline_normalized:
                            # Strong attribution: Contact person matches article byline
                            if contact_normalized:
                                # Check for exact match or partial match (handle middle names, etc.)
                                if contact_normalized == byline_normalized:
                                    return 'strong', actual_byline
                                # Check if one name is contained in the other (e.g., "John Smith" vs "John A. Smith")
                                elif contact_normalized in byline_normalized or byline_normalized in contact_normalized:
                                    return 'strong', actual_byline
                                else:
                                    # Different author - this is delegated
                                    return 'delegated', actual_byline
                            else:
                                # No contact to compare, but we have a byline
                                return 'unknown', actual_byline
                        else:
                            # No byline found
                            return 'unknown', None
                    
                    except Exception as e:
                        print(f"Attribution analysis error: {e}")
                        return 'unknown', None

                def format_attribution_strength(row):
                    """Format attribution strength for display with smart logic"""
                    # First check if we have attribution strength from database
                    db_attribution = str(row.get('Attribution_Strength', '')).strip().lower()
                    if db_attribution in ['strong', 'delegated', 'unknown']:
                        if db_attribution == 'strong':
                            return '✅ Direct'
                        elif db_attribution == 'delegated':
                            return '⚠️ Delegated'
                        else:
                            return '❓ Unknown'
                    
                    # Otherwise calculate it
                    attribution_strength, _ = smart_attribution_analysis(row)
                    
                    if attribution_strength == 'strong':
                        return '✅ Direct'
                    elif attribution_strength == 'delegated':
                        return '⚠️ Delegated'
                    else:
                        return '❓ Unknown'

                def get_actual_byline(row):
                    """Get actual byline author with smart fallbacks"""
                    try:
                        # NEW LOGIC: Check if we have a saved value from user editing first
                        wo_num = str(row.get('WO #', ''))
                        if wo_num in st.session_state.last_saved_bylines:
                            return st.session_state.last_saved_bylines[wo_num]
                        
                        # Check if we have an actual byline from the database
                        actual_byline = str(row.get('Actual_Byline', '')).strip()
                        if actual_byline and actual_byline.lower() not in ['nan', 'none', '', '—', 'null']:
                            # IMPORTANT: Check if this looks like the concatenated problematic string
                            # If it contains "Posted:" and "Author:" pattern, skip it and use fallbacks
                            if 'Posted:' in actual_byline and 'Author:' in actual_byline:
                                pass  # Skip this value, continue to fallbacks
                            else:
                                return actual_byline
                        
                        # NEW LOGIC: Default to Contact field value (user's requirement)
                        contact_person = str(row.get('To', '')).strip()
                        if contact_person and contact_person.lower() not in ['nan', 'none', '']:
                            return contact_person
                        
                        # If no contact name available, return placeholder
                        return '—'
                    
                    except:
                        # If anything fails, try to use contact as fallback
                        try:
                            contact_person = str(row.get('To', '')).strip()
                            if contact_person and contact_person.lower() not in ['nan', 'none', '']:
                                return contact_person
                        except:
                            pass
                        return '—'

                # Attribution logic kept in code but column hidden from UI
                # clean_df['✍️ Attribution'] = display_df.apply(format_attribution_strength, axis=1)
                clean_df['📝 Byline Author'] = display_df.apply(get_actual_byline, axis=1)
                
                # Override with saved Byline Author values from session state
                for idx, row in clean_df.iterrows():
                    wo_num = str(row.get('WO #', ''))
                    if wo_num in st.session_state.last_saved_bylines:
                        clean_df.at[idx, '📝 Byline Author'] = st.session_state.last_saved_bylines[wo_num]
                
                # Store the full URL tracking data for popup (hidden column)
                clean_df['URL_Tracking_Data'] = display_df.apply(lambda row: json.dumps(parse_url_tracking(row)), axis=1)
                
                # Add activity_id as a hidden column for hyperlink functionality
                clean_df['Activity_ID'] = display_df['activity_id'] if 'activity_id' in display_df.columns else ''
                
                # Note: Mark Viewed column already added at the beginning
                
                # Note: Saved checkbox states are already loaded at the beginning of the tab
                
                # Add action columns with session state persistence
                clean_df['✅ Approve'] = clean_df['WO #'].apply(lambda wo: str(wo) in st.session_state.approved_records)
                clean_df['❌ Reject'] = clean_df['WO #'].apply(lambda wo: str(wo) in st.session_state.rejected_records)
                
                # Populate last_saved_outlets and last_saved_bylines with current database values
                # Only populate if not already set (to preserve changes across refreshes)
                for idx, row in clean_df.iterrows():
                    wo_num = str(row.get('WO #', ''))
                    media_outlet = row.get('Media Outlet', '')
                    byline_author = row.get('📝 Byline Author', '')
                    contact = row.get('Contact', '')
                    person_id = row.get('Person_ID', '')
                    
                    # Only set if not already tracked (preserves user changes)
                    if wo_num and media_outlet and wo_num not in st.session_state.last_saved_outlets:
                        st.session_state.last_saved_outlets[wo_num] = media_outlet
                    
                    # For Byline Author: Use Contact as default if no byline_author exists
                    if wo_num and wo_num not in st.session_state.last_saved_bylines:
                        # If we have a byline author, use it; otherwise default to Contact
                        if byline_author and byline_author.strip() and byline_author.strip() not in ['—', 'nan', 'None']:
                            st.session_state.last_saved_bylines[wo_num] = byline_author
                        elif contact and contact.strip():
                            st.session_state.last_saved_bylines[wo_num] = contact
                    
                    # Initialize published dates to prevent false changes on load
                    published_date = row.get('📅 Published Date', '')
                    if wo_num and published_date and wo_num not in st.session_state.last_saved_dates:
                        st.session_state.last_saved_dates[wo_num] = published_date
                    
                    # Populate outlet data mapping for this WO
                    if wo_num and person_id and wo_num not in st.session_state.outlet_data_mapping:
                        full_outlet_data = get_full_outlet_data_for_person(person_id, person_outlets_mapping)
                        st.session_state.outlet_data_mapping[wo_num] = full_outlet_data
                
                # Create editable view renderer with edit icon
                cellRenderer_view = JsCode("""
                class UrlCellRenderer {
                  init(params) {
                    const isViewed = params.data['Viewed'];
                    this.isEditMode = false;
                    this.params = params;
                    
                    this.eGui = document.createElement('div');
                    this.eGui.style.display = 'flex';
                    this.eGui.style.alignItems = 'center';
                    this.eGui.style.gap = '5px';
                    this.eGui.style.width = '100%';
                    
                    // Add checkmark for viewed records
                    if (isViewed) {
                      const checkmark = document.createElement('span');
                      checkmark.innerHTML = '✓ ';
                      checkmark.style.color = '#28a745';
                      checkmark.style.fontWeight = 'bold';
                      checkmark.style.fontSize = '12px';
                      this.eGui.appendChild(checkmark);
                    }
                    
                    // Container for view/edit mode
                    this.contentContainer = document.createElement('div');
                    this.contentContainer.style.display = 'flex';
                    this.contentContainer.style.alignItems = 'center';
                    this.contentContainer.style.gap = '5px';
                    this.contentContainer.style.flex = '1';
                    this.contentContainer.style.minWidth = '0'; // Allow flex shrinking
                    this.contentContainer.style.overflow = 'hidden';
                    
                    // Create the link
                    this.link = document.createElement('a');
                    this.link.innerText = '📄 View';
                    this.link.href = params.data['Clip URL'];
                    this.link.target = '_blank';
                    this.link.style.color = isViewed ? '#6c757d' : '#1f77b4';
                    this.link.style.textDecoration = 'underline';
                    this.link.style.cursor = 'pointer';
                    this.link.style.opacity = isViewed ? '0.7' : '1';
                    this.link.title = params.data['Clip URL']; // Show full URL on hover
                    
                    // Create edit icon
                    this.editIcon = document.createElement('span');
                    this.editIcon.innerHTML = '✏️';
                    this.editIcon.style.cursor = 'pointer';
                    this.editIcon.style.fontSize = '12px';
                    this.editIcon.style.opacity = '0.6';
                    this.editIcon.title = 'Edit URL';
                    
                    // Create input field (hidden initially)
                    this.input = document.createElement('input');
                    this.input.type = 'text';
                    this.input.style.display = 'none';
                    this.input.style.width = 'calc(100% - 60px)'; // Leave room for buttons
                    this.input.style.minWidth = '200px';
                    this.input.style.padding = '2px 4px';
                    this.input.style.fontSize = '12px';
                    this.input.style.border = '1px solid #ccc';
                    this.input.style.borderRadius = '3px';
                    
                    // Create save/cancel buttons (hidden initially)
                    this.saveBtn = document.createElement('button');
                    this.saveBtn.innerHTML = '✓';
                    this.saveBtn.style.display = 'none';
                    this.saveBtn.style.cursor = 'pointer';
                    this.saveBtn.style.padding = '2px 6px';
                    this.saveBtn.style.fontSize = '11px';
                    this.saveBtn.style.backgroundColor = '#28a745';
                    this.saveBtn.style.color = 'white';
                    this.saveBtn.style.border = 'none';
                    this.saveBtn.style.borderRadius = '3px';
                    this.saveBtn.title = 'Save';
                    
                    this.cancelBtn = document.createElement('button');
                    this.cancelBtn.innerHTML = '✕';
                    this.cancelBtn.style.display = 'none';
                    this.cancelBtn.style.cursor = 'pointer';
                    this.cancelBtn.style.padding = '2px 6px';
                    this.cancelBtn.style.fontSize = '11px';
                    this.cancelBtn.style.backgroundColor = '#dc3545';
                    this.cancelBtn.style.color = 'white';
                    this.cancelBtn.style.border = 'none';
                    this.cancelBtn.style.borderRadius = '3px';
                    this.cancelBtn.title = 'Cancel';
                    
                    // Add event listeners
                    this.editIcon.addEventListener('click', () => this.enterEditMode());
                    this.saveBtn.addEventListener('click', () => this.saveUrl());
                    this.cancelBtn.addEventListener('click', () => this.exitEditMode());
                    this.input.addEventListener('keydown', (e) => {
                      if (e.key === 'Enter') this.saveUrl();
                      if (e.key === 'Escape') this.exitEditMode();
                      // Allow Ctrl+V / Cmd+V
                      if ((e.ctrlKey || e.metaKey) && e.key === 'v') {
                        e.stopPropagation();
                      }
                    });
                    
                    // Enable paste functionality
                    this.input.addEventListener('paste', (e) => {
                      e.stopPropagation();
                      // Allow default paste behavior
                    });
                    
                    // Prevent grid from intercepting input events
                    this.input.addEventListener('click', (e) => e.stopPropagation());
                    this.input.addEventListener('dblclick', (e) => e.stopPropagation());
                    this.input.addEventListener('mousedown', (e) => e.stopPropagation());
                    
                    // Add elements to container
                    this.contentContainer.appendChild(this.link);
                    this.contentContainer.appendChild(this.editIcon);
                    this.contentContainer.appendChild(this.input);
                    this.contentContainer.appendChild(this.saveBtn);
                    this.contentContainer.appendChild(this.cancelBtn);
                    
                    this.eGui.appendChild(this.contentContainer);
                  }
                  
                  enterEditMode() {
                    this.isEditMode = true;
                    this.link.style.display = 'none';
                    this.editIcon.style.display = 'none';
                    this.input.style.display = 'block';
                    this.saveBtn.style.display = 'inline-block';
                    this.cancelBtn.style.display = 'inline-block';
                    this.input.value = this.params.data['Clip URL'];
                    
                    // Ensure proper focus and selection
                    setTimeout(() => {
                      this.input.focus();
                      this.input.select();
                      // Enable context menu for right-click paste
                      this.input.addEventListener('contextmenu', (e) => {
                        e.stopPropagation();
                      });
                    }, 10);
                  }
                  
                  exitEditMode() {
                    this.isEditMode = false;
                    this.link.style.display = 'inline';
                    this.editIcon.style.display = 'inline';
                    this.input.style.display = 'none';
                    this.saveBtn.style.display = 'none';
                    this.cancelBtn.style.display = 'none';
                  }
                  
                  saveUrl() {
                    const newUrl = this.input.value.trim();
                    if (newUrl && newUrl !== this.params.data['Clip URL']) {
                      // Update the data
                      this.params.node.setDataValue('Clip URL', newUrl);
                      
                      // Update the link
                      this.link.href = newUrl;
                      
                      // Trigger grid update - this will cause the grid data to be sent to Python
                      this.params.api.refreshCells({
                        force: true,
                        columns: ['📄 View', 'Clip URL'],
                        rowNodes: [this.params.node]
                      });
                    }
                    
                    this.exitEditMode();
                  }

                  getGui() {
                    return this.eGui;
                  }

                  refresh(params) {
                    const isViewed = params.data['Viewed'];
                    this.link.style.color = isViewed ? '#6c757d' : '#1f77b4';
                    this.link.style.opacity = isViewed ? '0.7' : '1';
                    this.link.href = params.data['Clip URL'];
                    this.link.title = params.data['Clip URL']; // Update tooltip
                    return true;
                  }
                }
                """)
                
                # Create checkbox cell renderer for Approve column
                cellRenderer_approve = JsCode("""
                class ApproveCellRenderer {
                  init(params) {
                    this.eGui = document.createElement('div');
                    this.eGui.style.display = 'flex';
                    this.eGui.style.justifyContent = 'center';
                    this.eGui.style.alignItems = 'center';
                    this.eGui.style.height = '100%';
                    this.eGui.style.paddingLeft = '0px';
                    
                    this.checkbox = document.createElement('input');
                    this.checkbox.type = 'checkbox';
                    this.checkbox.checked = params.value === true;
                    this.checkbox.style.cursor = 'pointer';
                    this.checkbox.style.transform = 'scale(1.2)';
                    
                    this.checkbox.addEventListener('change', () => {
                      if (this.checkbox.checked) {
                        // Validate Pub Date, Byline Author, and Media Outlet before allowing approval
                        const pubDate = params.data['📅 Published Date'];
                        const bylineAuthor = params.data['📝 Byline Author'];
                        const mediaOutlet = params.data['Media Outlet'];
                        
                        // Check if Pub Date is valid (not empty, null, or just whitespace)
                        const isValidDate = pubDate && pubDate.trim() !== '' && pubDate.trim() !== '-';
                        
                        // Check if Byline Author is valid (not empty, null, just whitespace, or just "-")
                        const isValidAuthor = bylineAuthor && bylineAuthor.trim() !== '' && bylineAuthor.trim() !== '-' && bylineAuthor.trim() !== '—';
                        
                        // Check if Media Outlet is valid (not empty, null, or just whitespace)
                        const isValidOutlet = mediaOutlet && mediaOutlet.trim() !== '' && mediaOutlet.trim() !== '-';
                        
                        if (!isValidDate || !isValidAuthor || !isValidOutlet) {
                          // Prevent approval and show alert
                          this.checkbox.checked = false;
                          
                          let errorMsg = 'Cannot approve this record:\\n\\n';
                          if (!isValidDate) {
                            errorMsg += '• Published Date is missing or invalid\\n';
                          }
                          if (!isValidAuthor) {
                            errorMsg += '• Byline Author is missing or invalid\\n';
                          }
                          if (!isValidOutlet) {
                            errorMsg += '• Media Outlet is missing or invalid\\n';
                          }
                          errorMsg += '\\nPlease fill in these fields before approving.';
                          
                          alert(errorMsg);
                          return;
                        }
                        
                        // If approve is checked, uncheck reject
                        const rowNode = params.node;
                        rowNode.setDataValue('❌ Reject', false);
                      }
                      // Don't use setValue to avoid triggering grid update
                      params.node.setDataValue('✅ Approve', this.checkbox.checked);
                      
                      params.api.refreshCells({
                        force: true,
                        columns: ['✅ Approve', '❌ Reject'],
                        rowNodes: [params.node]
                      });
                    });
                    
                    this.eGui.appendChild(this.checkbox);
                  }

                  getGui() {
                    return this.eGui;
                  }

                  refresh(params) {
                    this.checkbox.checked = params.value === true;
                    return true;
                  }
                }
                """)
                
                # Create checkbox cell renderer for Reject column
                cellRenderer_reject = JsCode("""
                class RejectCellRenderer {
                  init(params) {
                    this.eGui = document.createElement('div');
                    this.eGui.style.display = 'flex';
                    this.eGui.style.justifyContent = 'center';
                    this.eGui.style.alignItems = 'center';
                    this.eGui.style.height = '100%';
                    this.eGui.style.paddingLeft = '0px';
                    
                    this.checkbox = document.createElement('input');
                    this.checkbox.type = 'checkbox';
                    this.checkbox.checked = params.value === true;
                    this.checkbox.style.cursor = 'pointer';
                    this.checkbox.style.transform = 'scale(1.2)';
                    
                    this.checkbox.addEventListener('change', () => {
                      if (this.checkbox.checked) {
                        // If reject is checked, uncheck approve
                        const rowNode = params.node;
                        rowNode.setDataValue('✅ Approve', false);
                      }
                      // Don't use setValue to avoid triggering grid update
                      params.node.setDataValue('❌ Reject', this.checkbox.checked);
                      
                      params.api.refreshCells({
                        force: true,
                        columns: ['✅ Approve', '❌ Reject'],
                        rowNodes: [params.node]
                      });
                    });
                    
                    this.eGui.appendChild(this.checkbox);
                  }

                  getGui() {
                    return this.eGui;
                  }

                  refresh(params) {
                    this.checkbox.checked = params.value === true;
                    return true;
                  }
                }
                """)
                
                # Create Mark Viewed button renderer
                cellRenderer_mark_viewed = JsCode("""
                class MarkViewedRenderer {
                  init(params) {
                    const isViewed = params.data['Viewed'];
                    
                    this.eGui = document.createElement('div');
                    this.eGui.style.display = 'flex';
                    this.eGui.style.justifyContent = 'center';
                    this.eGui.style.alignItems = 'center';
                    this.eGui.style.height = '100%';
                    
                    this.button = document.createElement('button');
                    this.button.innerHTML = isViewed ? '✓ Viewed' : '👁️ Mark';
                    this.button.style.padding = '4px 8px';
                    this.button.style.fontSize = '11px';
                    this.button.style.border = '1px solid #ccc';
                    this.button.style.borderRadius = '4px';
                    this.button.style.cursor = 'pointer';
                    this.button.style.backgroundColor = isViewed ? '#d4edda' : '#f8f9fa';
                    this.button.style.color = isViewed ? '#155724' : '#495057';
                    
                    this.button.addEventListener('click', () => {
                      const newValue = !params.data['Viewed'];
                      const woNum = params.data['WO #'];
                      
                      // Don't use setValue to avoid triggering grid update
                      params.node.setDataValue('Viewed', newValue);
                      
                      // Also update the 👁️ Mark Viewed column to trigger session state update
                      params.node.setDataValue('👁️ Mark Viewed', newValue);
                      
                      // Update button appearance
                      this.button.innerHTML = newValue ? '✓ Viewed' : '👁️ Mark';
                      this.button.style.backgroundColor = newValue ? '#d4edda' : '#f8f9fa';
                      this.button.style.color = newValue ? '#155724' : '#495057';
                      
                      // Refresh the entire row to update styling
                      params.api.refreshCells({
                        force: true,
                        rowNodes: [params.node]
                      });
                    });
                    
                    this.eGui.appendChild(this.button);
                  }

                  getGui() {
                    return this.eGui;
                  }

                  refresh(params) {
                    const isViewed = params.data['Viewed'];
                    this.button.innerHTML = isViewed ? '✓ Viewed' : '👁️ Mark';
                    this.button.style.backgroundColor = isViewed ? '#d4edda' : '#f8f9fa';
                    this.button.style.color = isViewed ? '#155724' : '#495057';
                    return true;
                  }
                }
                """)
                
                # Create cell renderer for WO # column with hyperlink to FMS activity
                cellRenderer_wo = JsCode("""
                class WoCellRenderer {
                  init(params) {
                    const woNumber = params.value;
                    const activityId = params.data['Activity_ID'];
                    
                    this.eGui = document.createElement('div');
                    
                    if (activityId && activityId !== '') {
                      // Create hyperlink to FMS activity
                      this.link = document.createElement('a');
                      this.link.innerText = woNumber;
                      this.link.href = `https://fms.driveshop.com/activities/edit/${activityId}`;
                      this.link.target = '_blank';
                      this.link.style.color = '#1f77b4';
                      this.link.style.textDecoration = 'underline';
                      this.link.style.cursor = 'pointer';
                      this.eGui.appendChild(this.link);
                    } else {
                      // No activity ID, just show the WO number as plain text
                      this.eGui.innerText = woNumber;
                    }
                  }

                  getGui() {
                    return this.eGui;
                  }

                  refresh(params) {
                    return false;
                  }
                }
                """)


                # Configure the grid
                gb = GridOptionsBuilder.from_dataframe(clean_df)
                
                # *** RESTORE ADVANCED FEATURES WITH SET FILTERS (CHECKBOXES) ***
                gb.configure_side_bar()  # Enable filtering sidebar
                gb.configure_default_column(
                    filter="agSetColumnFilter",  # CHECKBOX FILTERS with search
                    sortable=True,  # Enable sorting
                    resizable=True,  # Enable column resizing
                    editable=False, 
                    groupable=True, 
                    value=True, 
                    enableRowGroup=True, 
                    enablePivot=True, 
                    enableValue=True,
                    filterParams={
                        "buttons": ["reset", "apply"],
                        "closeOnApply": True,
                        "newRowsAction": "keep"
                    }
                )
                
                # Now configure visible columns in the exact order needed
                # 1. Viewed (Mark Viewed button)
                gb.configure_column(
                    "👁️ Mark Viewed",
                    headerName="Viewed",
                    cellRenderer=cellRenderer_mark_viewed,
                    minWidth=80,
                    maxWidth=100,
                    editable=True,
                    sortable=False,
                    filter=False,
                    pinned='left'  # Keep it visible when scrolling
                )
                
                # 2. Work Order #
                gb.configure_column("WO #", minWidth=100, cellRenderer=cellRenderer_wo, headerName="Work Order #")
                
                # 3. Contact
                gb.configure_column("Contact", minWidth=180)
                
                # 4. Media Outlet (configured here but renderer will be set later if dropdown is available)
                gb.configure_column("Media Outlet", minWidth=220, editable=True, sortable=True, filter=True)
                
                # 5. View (clip link)
                gb.configure_column(
                    "📄 View", 
                    cellRenderer=cellRenderer_view,
                    minWidth=80,
                    maxWidth=100,
                    sortable=False,
                    filter=False
                )
                
                # 6. Make
                gb.configure_column("Make", minWidth=120)
                
                # 7. Model
                gb.configure_column("Model", minWidth=150)
                
                # 8. Pub Date (Published Date)
                gb.configure_column(
                    "📅 Published Date",
                    headerName="Pub Date",
                    editable=True,
                    cellEditor="agTextCellEditor",
                    cellEditorParams={
                        "maxLength": 8  # MM/DD/YY is 8 characters
                    },
                    valueParser=JsCode("""
                    function(params) {
                        const value = params.newValue;
                        if (!value || value === '—') return value;
                        
                        // Remove any non-numeric characters except /
                        const cleaned = value.replace(/[^0-9/]/g, '');
                        
                        // Check if it matches MM/DD/YY format
                        const dateRegex = /^(0[1-9]|1[0-2])\/(0[1-9]|[12][0-9]|3[01])\/\d{2}$/;
                        if (dateRegex.test(cleaned)) {
                            return cleaned;
                        }
                        
                        // Try to parse and format the date
                        const parts = cleaned.split('/');
                        if (parts.length === 3) {
                            const month = parts[0].padStart(2, '0');
                            const day = parts[1].padStart(2, '0');
                            const year = parts[2].length === 4 ? parts[2].substring(2) : parts[2].padStart(2, '0');
                            
                            const formatted = month + '/' + day + '/' + year;
                            if (dateRegex.test(formatted)) {
                                return formatted;
                            }
                        }
                        
                        // Return original value if can't parse
                        return params.oldValue;
                    }
                    """),
                    minWidth=100,
                    sortable=True,
                    filter=True,
                    tooltipField="📅 Published Date",
                    tooltipValueGetter=JsCode("""
                    function(params) {
                        return "Enter date as MM/DD/YY";
                    }
                    """)
                )
                
                # 9. Byline Author
                gb.configure_column(
                    "📝 Byline Author",
                    headerName="Byline Author",
                    editable=True,
                    cellEditor="agTextCellEditor",
                    cellEditorParams={
                        "maxLength": 100  # Limit input length
                    },
                    minWidth=180,
                    sortable=True,
                    filter=True
                )
                
                # 10. Score (Relevance)
                gb.configure_column("Relevance", minWidth=80, headerName="Score")

                # Configure selection
                gb.configure_selection(selection_mode="multiple", use_checkbox=False)
                
                # First, configure all hidden columns
                gb.configure_column("Office", hide=True)
                gb.configure_column("Person_ID", hide=True)
                gb.configure_column("Clip URL", hide=True)
                gb.configure_column("URL_Tracking_Data", hide=True)
                gb.configure_column("Viewed", hide=True)  # Hide the viewed status column
                gb.configure_column("Activity_ID", hide=True)  # Hide the activity ID column
                gb.configure_column("Outlet_Options", hide=True)  # Hide the outlet options column
                
                
                # Load Person_ID to Media Outlets mapping for dropdown
                person_outlets_mapping = load_person_outlets_mapping()
                
                # Update Media Outlet column with dropdown if mapping is available
                if person_outlets_mapping:
                    
                    # Create custom cell renderer for dynamic dropdown options
                    cellRenderer_outlet_dropdown = JsCode("""
                    class OutletDropdownRenderer {
                      init(params) {
                        this.eGui = document.createElement('select');
                        this.eGui.style.width = '100%';
                        this.eGui.style.height = '100%';
                        this.eGui.style.border = 'none';
                        this.eGui.style.background = 'transparent';
                        this.eGui.style.fontSize = '12px';
                        
                        // Get Person_ID from the row data
                        const personId = params.data['Person_ID'] || params.data['Contact'];
                        const currentValue = params.value || '';
                        
                        // Debug log to see what value we're working with
                        console.log('Outlet Dropdown - Person:', personId, 'Current Value:', currentValue);
                        
                        // Get outlet options based on Person_ID
                        const outletOptions = params.data['Outlet_Options'] || [];
                        
                        // Add empty option only if no current value is set
                        if (!currentValue) {
                          const emptyOption = document.createElement('option');
                          emptyOption.value = '';
                          emptyOption.text = 'Select outlet...';
                          emptyOption.selected = true;
                          this.eGui.appendChild(emptyOption);
                        }
                        
                        // Add outlet options based on Person_ID (these are the valid Outlet_Names)
                        outletOptions.forEach(outlet => {
                          const option = document.createElement('option');
                          option.value = outlet;
                          option.text = outlet;
                          option.selected = currentValue === outlet;
                          this.eGui.appendChild(option);
                        });
                        
                        // Set the current value (should be pre-selected by backend smart matching)
                        this.eGui.value = currentValue;
                        
                        // Add change event listener
                        this.eGui.addEventListener('change', () => {
                          params.setValue(this.eGui.value);
                        });
                      }

                      getGui() {
                        return this.eGui;
                      }

                      refresh(params) {
                        return false;
                      }
                    }
                    """)
                    
                    # Update the already configured Media Outlet column to add dropdown functionality
                    gb.configure_column(
                        "Media Outlet",
                        cellEditor="agSelectCellEditor",
                        cellEditorParams={
                            "values": []  # Will be populated dynamically per row
                        },
                        cellRenderer=cellRenderer_outlet_dropdown,  # Add the dropdown renderer
                        minWidth=220,
                        editable=True,
                        sortable=True,
                        filter=True
                    )
                    
                    # Add outlet options to each row based on Person_ID
                    def add_outlet_options(row):
                        person_id = row.get('Person_ID')
                        outlet_options = get_outlet_options_for_person(person_id, person_outlets_mapping)
                        # --- FIX: Ensure options are a list of strings ---
                        row['Outlet_Options'] = [str(opt) for opt in outlet_options] if outlet_options else []
                        return row
                    
                    # Apply outlet options to each row
                    clean_df = clean_df.apply(add_outlet_options, axis=1)
                
                # Reorder columns to match the requested sequence
                column_order = [
                    '👁️ Mark Viewed',  # Viewed
                    'WO #',            # Work Order #
                    'Contact',         # Contact
                    'Media Outlet',    # Media Outlet
                    '📄 View',         # View (clip link)
                    'Make',            # Make
                    'Model',           # Model
                    '📅 Published Date',  # Pub Date
                    '📝 Byline Author',   # Byline Author
                    'Relevance',       # Score
                    '✅ Approve',      # Approve
                    '❌ Reject',       # Reject
                    # Hidden columns
                    'Office',
                    'Person_ID',
                    'Clip URL',
                    'URL_Tracking_Data',
                    'Viewed',
                    'Activity_ID',
                    'Outlet_Options'
                ]
                
                # Ensure all columns exist before reordering
                existing_columns = [col for col in column_order if col in clean_df.columns]
                # Add any remaining columns not in the order list
                remaining_columns = [col for col in clean_df.columns if col not in existing_columns]
                clean_df = clean_df[existing_columns + remaining_columns]
                
                # Add row styling for viewed records with better visibility
                gb.configure_grid_options(
                    maintainColumnOrder=True,  # Maintain column order from DataFrame
                    getRowStyle=JsCode("""
                    function(params) {
                        if (params.data.Viewed === true) {
                            return {
                                'background-color': '#e8f5e8',
                                'border-left': '4px solid #28a745',
                                'opacity': '0.85'
                            };
                        }
                        return {};
                    }
                    """)
                )
                
                # 11. Approve
                gb.configure_column(
                    "✅ Approve", 
                    headerName="Approve",
                    cellRenderer=cellRenderer_approve,
                    minWidth=90,
                    editable=True,
                    sortable=False,
                    filter=False
                )
                
                # 12. Reject
                gb.configure_column(
                    "❌ Reject", 
                    headerName="Reject",
                    cellRenderer=cellRenderer_reject,
                    minWidth=90,
                    editable=True,
                    sortable=False,
                    filter=False
                )
                
                # Configure grid auto-sizing
                gb.configure_grid_options(
                    domLayout='normal',
                    maintainColumnOrder=True,
                    onFirstDataRendered=JsCode("""
                    function(params) {
                        params.api.sizeColumnsToFit();
                    }
                    """),
                    onGridSizeChanged=JsCode("""
                    function(params) {
                        params.api.sizeColumnsToFit();
                    }
                    """)
                )
                
                # Build grid options
                grid_options = gb.build()
                
                # Force column definitions in the exact order we want
                grid_options['columnDefs'] = [
                    # Approve and Reject columns first with icon-only headers
                    {'field': '✅ Approve', 'headerName': '✅', 'cellRenderer': cellRenderer_approve, 'minWidth': 80, 'maxWidth': 90, 'width': 85, 'editable': True, 'sortable': False, 'filter': False, 'pinned': 'left', 'cellStyle': {'backgroundColor': '#f0f9f0'}},
                    {'field': '❌ Reject', 'headerName': '❌', 'cellRenderer': cellRenderer_reject, 'minWidth': 80, 'maxWidth': 90, 'width': 85, 'editable': True, 'sortable': False, 'filter': False, 'pinned': 'left', 'cellStyle': {'backgroundColor': '#fef0f0'}},
                    {'field': 'WO #', 'headerName': 'Work Order #', 'cellRenderer': cellRenderer_wo, 'minWidth': 100},
                    {'field': 'Contact', 'minWidth': 180, 'sortable': True, 'sort': 'asc'},
                    {'field': 'Media Outlet', 'cellRenderer': cellRenderer_outlet_dropdown if person_outlets_mapping else None, 'minWidth': 220, 'editable': True, 'sortable': True, 'filter': True},
                    {'field': '📄 View', 'cellRenderer': cellRenderer_view, 'minWidth': 150, 'maxWidth': 600, 'sortable': False, 'filter': False, 'resizable': True},
                    {'field': 'Make', 'minWidth': 120},
                    {'field': 'Model', 'minWidth': 150},
                    {'field': '📅 Published Date', 'headerName': 'Pub Date', 'editable': True, 'minWidth': 100, 'sortable': True, 'filter': True, 'cellEditor': 'agTextCellEditor', 'cellEditorParams': {'maxLength': 8}},
                    {'field': '📝 Byline Author', 'headerName': 'Byline Author', 'editable': True, 'minWidth': 180, 'sortable': True, 'filter': True},
                    {'field': 'Relevance', 'headerName': 'Score', 'minWidth': 80},
                    # Hidden columns
                    {'field': '👁️ Mark Viewed', 'hide': True},  # Hide but keep for functionality
                    {'field': 'Office', 'hide': True},
                    {'field': 'Person_ID', 'hide': True},
                    {'field': 'Clip URL', 'hide': True},
                    {'field': 'URL_Tracking_Data', 'hide': True},
                    {'field': 'Viewed', 'hide': True},
                    {'field': 'Activity_ID', 'hide': True},
                    {'field': 'Outlet_Options', 'hide': True}
                ]
                
                # Add default sorting by Contact field in ascending order
                grid_options['defaultColDef'] = grid_options.get('defaultColDef', {})
                grid_options['defaultColDef']['sortable'] = True
                grid_options['sortModel'] = [
                    {'colId': 'Contact', 'sort': 'asc'}
                ]
                
                # Call AgGrid with Enterprise modules enabled for Set Filters
                selected_rows = AgGrid(
                    clean_df,
                    gridOptions=grid_options,
                    allow_unsafe_jscode=True,
                    update_mode=GridUpdateMode.VALUE_CHANGED,  # Only trigger on cell value changes
                    height=400,  # Reduced height so action buttons are visible without scrolling
                    fit_columns_on_grid_load=True,
                    columns_auto_size_mode='FIT_ALL_COLUMNS_TO_VIEW',  # Auto-size all columns
                    theme="alpine",
                    enable_enterprise_modules=True,  # REQUIRED for Set Filters with checkboxes
                    reload_data=False,  # Prevent automatic data reloading
                    key="bulk_review_grid_stable"  # Stable key to prevent unnecessary reruns
                )
                
                                                # Process grid changes to update session state (debounced to prevent flashing)
                if selected_rows["data"] is not None and not selected_rows["data"].empty:
                    grid_df = selected_rows["data"]
                    
                    # Update session state based on Mark Viewed button states
                    new_viewed_records = set()
                    new_approved_records = set()
                    new_rejected_records = set()
                    
                    for idx, row in grid_df.iterrows():
                        wo_num = str(row.get('WO #', ''))
                        if not wo_num:
                            continue
                            
                        # Track viewed records - check both columns
                        if row.get('Viewed', False) or row.get('👁️ Mark Viewed', False):
                            new_viewed_records.add(wo_num)
                        
                        # Track approved records
                        if row.get('✅ Approve', False):
                            new_approved_records.add(wo_num)
                        
                        # Track rejected records
                        if row.get('❌ Reject', False):
                            new_rejected_records.add(wo_num)
                    
                    # Silently update session state (avoid reruns that cause flashing)
                    st.session_state.viewed_records = new_viewed_records
                    st.session_state.approved_records = new_approved_records
                    st.session_state.rejected_records = new_rejected_records
                    
                    # Also update legacy session state for compatibility
                    st.session_state.selected_for_approval = new_approved_records.copy()
                    st.session_state.selected_for_rejection = new_rejected_records.copy()
                
                # Note: Session state tracking already initialized at the beginning of the tab
                if 'selected_for_approval' not in st.session_state:
                    st.session_state.selected_for_approval = set()
                if 'selected_for_rejection' not in st.session_state:
                    st.session_state.selected_for_rejection = set()
                if 'show_rejection_dialog' not in st.session_state:
                    st.session_state.show_rejection_dialog = False
                
                # Process changes from AgGrid WITHOUT triggering reruns
                if not selected_rows["data"].empty:
                    # Debug: Print current checkbox states
                    approved_rows = selected_rows["data"][selected_rows["data"]["✅ Approve"] == True]
                    rejected_rows = selected_rows["data"][selected_rows["data"]["❌ Reject"] == True]
                    if not approved_rows.empty or not rejected_rows.empty:
                        print(f"🔍 Checkbox changes detected: {len(approved_rows)} approved, {len(rejected_rows)} rejected")
                    
                    # 1. First handle Media Outlet changes (save to database)
                    outlet_changed = False
                    changed_count = 0
                    changed_wos = []
                    
                    for idx, row in selected_rows["data"].iterrows():
                        wo_num = str(row.get('WO #', ''))
                        new_outlet = row.get('Media Outlet', '')
                        
                        if wo_num and new_outlet:
                            # Get the last saved value to avoid duplicate saves
                            last_saved = st.session_state.last_saved_outlets.get(wo_num, '')
                            
                            # Save if different from last saved
                            if new_outlet != last_saved:
                                outlet_changed = True
                                changed_count += 1
                                changed_wos.append(wo_num)
                                st.session_state.last_saved_outlets[wo_num] = new_outlet
                                print(f"💾 Saving Media Outlet change for WO# {wo_num}: → '{new_outlet}'")
                    
                    # Save outlet changes to database (background operation)
                    if outlet_changed:
                        try:
                            # Update the database with new media outlet selections
                            # Since there's no dedicated media_outlet field, we'll store it in the contact field temporarily
                            # or add a new field to the database schema
                            
                            # Update clips in the database
                            for wo_num in changed_wos:
                                new_outlet = st.session_state.last_saved_outlets[wo_num]
                                try:
                                    # Get database connection
                                    db = get_cached_database()
                                    
                                    # Get outlet data for this WO
                                    outlet_id = None
                                    impressions = None
                                    if wo_num in st.session_state.outlet_data_mapping:
                                        outlet_data_dict = st.session_state.outlet_data_mapping[wo_num]
                                        if new_outlet in outlet_data_dict:
                                            outlet_info = outlet_data_dict[new_outlet]
                                            outlet_id = outlet_info.get('outlet_id')
                                            impressions = outlet_info.get('impressions')
                                    
                                    # Update the clip in database with outlet data
                                    success = db.update_clip_media_outlet(wo_num, new_outlet, outlet_id, impressions)
                                    if success:
                                        print(f"✅ Updated WO# {wo_num} media outlet to: {new_outlet} (ID: {outlet_id}, Impressions: {impressions})")
                                    else:
                                        print(f"⚠️ Failed to update WO# {wo_num} in database")
                                except Exception as e:
                                    print(f"❌ Error updating WO# {wo_num}: {e}")
                            
                            # Use session state to show success message without rerun
                            from datetime import datetime
                            timestamp = datetime.now().strftime("%H:%M:%S")
                            if changed_count == 1:
                                st.session_state.outlet_save_message = f"💾 Media Outlet saved for WO# {changed_wos[0]} at {timestamp}"
                            else:
                                st.session_state.outlet_save_message = f"💾 {changed_count} Media Outlet selections saved at {timestamp}"
                            print(f"✅ Updated database with {changed_count} Media Outlet changes")
                        except Exception as e:
                            st.session_state.outlet_save_message = f"❌ Error saving Media Outlet changes: {e}"
                            print(f"❌ Error saving changes: {e}")
                    
                    # 1.5. Handle Byline Author changes (save to database)
                    byline_changed = False
                    byline_changed_count = 0
                    byline_changed_wos = []
                    
                    for idx, row in selected_rows["data"].iterrows():
                        wo_num = str(row.get('WO #', ''))
                        new_byline = row.get('📝 Byline Author', '')
                        
                        if wo_num and new_byline:
                            # Get the last saved value to avoid duplicate saves
                            last_saved_byline = st.session_state.last_saved_bylines.get(wo_num, '')
                            
                            # Save if different from last saved
                            if new_byline != last_saved_byline:
                                byline_changed = True
                                byline_changed_count += 1
                                byline_changed_wos.append(wo_num)
                                st.session_state.last_saved_bylines[wo_num] = new_byline
                                print(f"💾 Saving Byline Author change for WO# {wo_num}: → '{new_byline}'")
                    
                    # Save byline changes to database
                    if byline_changed:
                        try:
                            # Update clips in the database
                            for wo_num in byline_changed_wos:
                                new_byline = st.session_state.last_saved_bylines[wo_num]
                                try:
                                    # Get database connection
                                    db = get_cached_database()
                                    # Update the clip in database using the new method
                                    success = db.update_clip_byline_author(wo_num, new_byline)
                                    if success:
                                        print(f"✅ Updated WO# {wo_num} byline author to: {new_byline}")
                                    else:
                                        print(f"⚠️ Failed to update WO# {wo_num} byline in database")
                                except Exception as e:
                                    print(f"❌ Error updating WO# {wo_num} byline: {e}")
                            
                            # Use session state to show success message
                            from datetime import datetime
                            timestamp = datetime.now().strftime("%H:%M:%S")
                            if byline_changed_count == 1:
                                st.session_state.byline_save_message = f"💾 Byline Author saved for WO# {byline_changed_wos[0]} at {timestamp}"
                            else:
                                st.session_state.byline_save_message = f"💾 {byline_changed_count} Byline Author edits saved at {timestamp}"
                            print(f"✅ Updated database with {byline_changed_count} Byline Author changes")
                        except Exception as e:
                            st.session_state.byline_save_message = f"❌ Error saving Byline Author changes: {e}"
                            print(f"❌ Error saving byline changes: {e}")
                    
                    # 1.6. Handle Published Date changes (save to database)
                    date_changed = False
                    date_changed_count = 0
                    date_changed_wos = []
                    
                    for idx, row in selected_rows["data"].iterrows():
                        wo_num = str(row.get('WO #', ''))
                        new_date = row.get('📅 Published Date', '')
                        
                        if wo_num and new_date:
                            # Get the last saved value to avoid duplicate saves
                            last_saved_date = st.session_state.last_saved_dates.get(wo_num, '')
                            
                            # Save if different from last saved
                            if new_date != last_saved_date:
                                date_changed = True
                                date_changed_count += 1
                                date_changed_wos.append(wo_num)
                                st.session_state.last_saved_dates[wo_num] = new_date
                                print(f"💾 Saving Published Date change for WO# {wo_num}: → '{new_date}'")
                    
                    # Save published date changes to database
                    if date_changed:
                        try:
                            # Update clips in the database
                            for wo_num in date_changed_wos:
                                new_date = st.session_state.last_saved_dates[wo_num]
                                try:
                                    # Get database connection
                                    db = get_cached_database()
                                    # Update the clip in database using the new method
                                    success = db.update_clip_published_date(wo_num, new_date)
                                    if success:
                                        print(f"✅ Updated WO# {wo_num} published date to: {new_date}")
                                    else:
                                        print(f"⚠️ Failed to update WO# {wo_num} published date in database")
                                except Exception as e:
                                    print(f"❌ Error updating WO# {wo_num} published date: {e}")
                            
                            # Use session state to show success message
                            from datetime import datetime
                            timestamp = datetime.now().strftime("%H:%M:%S")
                            if date_changed_count == 1:
                                st.session_state.date_save_message = f"💾 Published Date saved for WO# {date_changed_wos[0]} at {timestamp}"
                            else:
                                st.session_state.date_save_message = f"💾 {date_changed_count} Published Date edits saved at {timestamp}"
                            print(f"✅ Updated database with {date_changed_count} Published Date changes")
                        except Exception as e:
                            st.session_state.date_save_message = f"❌ Error saving Published Date changes: {e}"
                            print(f"❌ Error saving published date changes: {e}")
                    
                    # 1.7. Handle URL changes (save to database)
                    url_changed = False
                    url_changed_count = 0
                    url_changed_wos = []
                    
                    for idx, row in selected_rows["data"].iterrows():
                        wo_num = str(row.get('WO #', ''))
                        new_url = row.get('Clip URL', '')
                        
                        if wo_num and new_url:
                            # Get the last saved value to avoid duplicate saves
                            last_saved_url = st.session_state.edited_urls.get(wo_num, '')
                            
                            # Also check against the original URL in the data
                            original_url = ''
                            for orig_idx, orig_row in df.iterrows():
                                if str(orig_row.get('WO #', '')) == wo_num:
                                    original_url = orig_row.get('Clip URL', '')
                                    break
                            
                            # Save if different from both last saved and original
                            if new_url != last_saved_url and new_url != original_url:
                                url_changed = True
                                url_changed_count += 1
                                url_changed_wos.append(wo_num)
                                st.session_state.edited_urls[wo_num] = new_url
                                print(f"💾 Saving URL change for WO# {wo_num}: → '{new_url}'")
                    
                    # Save URL changes to database
                    if url_changed:
                        try:
                            # Update clips in the database
                            for wo_num in url_changed_wos:
                                new_url = st.session_state.edited_urls[wo_num]
                                try:
                                    # Get database connection
                                    db = get_cached_database()
                                    # Update the clip in database
                                    success = db.update_clip_url(wo_num, new_url)
                                    if success:
                                        print(f"✅ Updated WO# {wo_num} URL to: {new_url}")
                                    else:
                                        print(f"⚠️ Failed to update WO# {wo_num} URL in database")
                                except Exception as e:
                                    print(f"❌ Error updating WO# {wo_num} URL: {e}")
                            
                            # Use session state to show success message
                            from datetime import datetime
                            timestamp = datetime.now().strftime("%H:%M:%S")
                            if url_changed_count == 1:
                                st.session_state.url_save_message = f"💾 URL saved for WO# {url_changed_wos[0]} at {timestamp}"
                            else:
                                st.session_state.url_save_message = f"💾 {url_changed_count} URL edits saved at {timestamp}"
                            print(f"✅ Updated database with {url_changed_count} URL changes")
                        except Exception as e:
                            st.session_state.url_save_message = f"❌ Error saving URL changes: {e}"
                            print(f"❌ Error saving URL changes: {e}")
                    
                    # 2. Then handle approval/rejection checkboxes (stable tracking)
                    approved_rows = selected_rows["data"][selected_rows["data"]["✅ Approve"] == True]
                    rejected_rows = selected_rows["data"][selected_rows["data"]["❌ Reject"] == True]
                    
                    # Get current checkbox states
                    current_approved_wos = set(approved_rows['WO #'].astype(str))
                    current_rejected_wos = set(rejected_rows['WO #'].astype(str))
                    
                    # REPLACE the session state entirely with current checkbox states
                    # This prevents accumulation and refresh issues
                    st.session_state.selected_for_approval = current_approved_wos.copy()
                    st.session_state.selected_for_rejection = current_rejected_wos.copy()
                    
                    # Also sync with persistent session state
                    st.session_state.approved_records = current_approved_wos.copy()
                    st.session_state.rejected_records = current_rejected_wos.copy()
                    
                    # Debug: Print session state updates
                    if current_approved_wos or current_rejected_wos:
                        print(f"📊 Session state updated: {len(current_approved_wos)} approved, {len(current_rejected_wos)} rejected")
                        print(f"   Approved WOs: {list(current_approved_wos)[:5]}...")  # Show first 5
                    
                    # Ensure mutual exclusivity (approve overrides reject)
                    if current_approved_wos:
                        st.session_state.selected_for_rejection -= current_approved_wos
                        st.session_state.rejected_records -= current_approved_wos
                    
                    # Save checkbox state to file for persistence
                    try:
                        import pickle
                        temp_file = os.path.join(project_root, "temp", "checkbox_state.pkl")
                        with open(temp_file, 'wb') as f:
                            pickle.dump({
                                'approved': st.session_state.approved_records,
                                'rejected': st.session_state.rejected_records,
                                'viewed': st.session_state.viewed_records,
                                'outlets': st.session_state.last_saved_outlets,
                                'bylines': st.session_state.last_saved_bylines,
                                'outlet_data': st.session_state.outlet_data_mapping
                            }, f)
                    except Exception as e:
                        print(f"Could not save checkbox state: {e}")
                
                # Display persistent messages
                if hasattr(st.session_state, 'outlet_save_message') and st.session_state.outlet_save_message:
                    if st.session_state.outlet_save_message.startswith("💾"):
                        st.success(st.session_state.outlet_save_message)
                    else:
                        st.error(st.session_state.outlet_save_message)
                    # Clear message after showing
                    st.session_state.outlet_save_message = None
                
                # Display byline save messages
                if hasattr(st.session_state, 'byline_save_message') and st.session_state.byline_save_message:
                    if st.session_state.byline_save_message.startswith("💾"):
                        st.success(st.session_state.byline_save_message)
                    else:
                        st.error(st.session_state.byline_save_message)
                    # Clear message after showing
                    st.session_state.byline_save_message = None
                
                # Display published date save messages
                if hasattr(st.session_state, 'date_save_message') and st.session_state.date_save_message:
                    if st.session_state.date_save_message.startswith("💾"):
                        st.success(st.session_state.date_save_message)
                    else:
                        st.error(st.session_state.date_save_message)
                    # Clear message after showing
                    st.session_state.date_save_message = None
                
                # Display URL save messages
                if hasattr(st.session_state, 'url_save_message') and st.session_state.url_save_message:
                    if st.session_state.url_save_message.startswith("💾"):
                        st.success(st.session_state.url_save_message)
                    else:
                        st.error(st.session_state.url_save_message)
                    # Clear message after showing
                    st.session_state.url_save_message = None
                
                # Show current selection counts
                approved_count = len(st.session_state.selected_for_approval)
                rejected_count = len(st.session_state.selected_for_rejection)
                if approved_count > 0:
                    st.info(f"📋 {approved_count} clips selected for approval")
                if rejected_count > 0:
                    st.info(f"📋 {rejected_count} clips selected for rejection")
                
                # Action buttons below table
                st.markdown("---")
                
                # Create sticky action bar container
                st.markdown('<div class="sticky-action-bar">', unsafe_allow_html=True)
                
                col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                
                with col1:
                    # Submit Approved Clips Button
                    selected_count = len(st.session_state.get('selected_for_approval', set()))
                    if st.button(f"✅ Submit {selected_count} Approved Clips", disabled=selected_count == 0, key="submit_approved_main"):
                        if selected_count > 0:
                            # Show confirmation dialog
                            st.session_state.show_approval_dialog = True
                
                with col2:
                    # Submit Rejected Clips Button (side by side with approved)
                    rejected_count = len(st.session_state.get('selected_for_rejection', set()))
                    if st.button(f"❌ Submit {rejected_count} Rejected Clips", disabled=rejected_count == 0, key="submit_rejected_main"):
                        if rejected_count > 0:
                            # Show rejection confirmation dialog
                            st.session_state.show_rejection_dialog = True
                
                with col3:
                    if st.button("✅ Auto-Approve High Quality (9+)"):
                        high_quality_df = df[df['Relevance Score'] >= 9]
                        if not high_quality_df.empty:
                            # Add to session state selections
                            if 'selected_for_approval' not in st.session_state:
                                st.session_state.selected_for_approval = set()
                            high_quality_wos = set(high_quality_df['WO #'].astype(str))
                            st.session_state.selected_for_approval.update(high_quality_wos)
                            st.success(f"📋 Added {len(high_quality_wos)} high-quality clips to selection!")
                            st.rerun()
                        else:
                            st.info("No high-quality clips (9+) found")
                
                with col4:
                    # Manual Save Progress button
                    if st.button("💾 Save Progress", help="Save all UI selections to database"):
                        try:
                            db = get_cached_database()
                            saved_count = 0
                            
                            # Save viewed states
                            for wo_num in st.session_state.viewed_records:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_viewed': True,
                                        'ui_viewed_at': datetime.now().isoformat()
                                    }).eq('wo_number', wo_num).execute()
                                    saved_count += 1
                                except:
                                    pass
                            
                            # Save approved checkbox states (not submitted, just UI state)
                            for wo_num in st.session_state.approved_records:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_approved_pending': True
                                    }).eq('wo_number', wo_num).execute()
                                    saved_count += 1
                                except:
                                    pass
                            
                            # Save rejected checkbox states (not submitted, just UI state)
                            for wo_num in st.session_state.rejected_records:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_rejected_pending': True
                                    }).eq('wo_number', wo_num).execute()
                                    saved_count += 1
                                except:
                                    pass
                            
                            # Clear any records that were unchecked
                            all_wos = set(str(row.get('WO #', '')) for _, row in df.iterrows() if row.get('WO #'))
                            cleared_approved = all_wos - st.session_state.approved_records
                            cleared_rejected = all_wos - st.session_state.rejected_records
                            cleared_viewed = all_wos - st.session_state.viewed_records
                            
                            for wo_num in cleared_approved:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_approved_pending': False
                                    }).eq('wo_number', wo_num).execute()
                                except:
                                    pass
                            
                            for wo_num in cleared_rejected:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_rejected_pending': False
                                    }).eq('wo_number', wo_num).execute()
                                except:
                                    pass
                            
                            for wo_num in cleared_viewed:
                                try:
                                    db.supabase.table('clips').update({
                                        'ui_viewed': False
                                    }).eq('wo_number', wo_num).execute()
                                except:
                                    pass
                            
                            st.success(f"💾 Saved progress to database!")
                            
                            # Also save to pickle as backup
                            import pickle
                            temp_file = os.path.join(project_root, "temp", "checkbox_state.pkl")
                            with open(temp_file, 'wb') as f:
                                pickle.dump({
                                    'approved': st.session_state.approved_records,
                                    'rejected': st.session_state.rejected_records,
                                    'viewed': st.session_state.viewed_records,
                                    'outlets': st.session_state.last_saved_outlets,
                                    'bylines': st.session_state.last_saved_bylines,
                                    'outlet_data': st.session_state.outlet_data_mapping
                                }, f)
                                
                        except Exception as e:
                            st.error(f"Failed to save: {str(e)}")
                
                # Close sticky action bar container
                st.markdown('</div>', unsafe_allow_html=True)
                
                # Approval confirmation dialog
                if st.session_state.get('show_approval_dialog', False):
                    st.markdown("---")
                    st.warning(f"⚠️ **Approve {selected_count} clips?** This will save them and generate client files.")
                    
                    col_confirm, col_cancel = st.columns(2)
                    with col_confirm:
                        if st.button("✅ Confirm Approval", type="primary", key="confirm_approval_btn"):
                            # Process the approvals - SIMPLIFIED WORKFLOW
                            selected_wos = st.session_state.selected_for_approval
                            if selected_wos:
                                # Update clips in database to approved status (workflow_stage stays 'found' for Approved Queue)
                                try:
                                    # Get database connection
                                    db = get_cached_database()
                                    
                                    # First, approve the clips with Media Outlet data
                                    approved_clips = []
                                    for wo_number in selected_wos:
                                        # Get Media Outlet data from session state
                                        update_data = {
                                            'status': 'approved',
                                            'workflow_stage': 'found'  # Keep as 'found' for now
                                        }
                                        
                                        # Add Media Outlet if selected
                                        if wo_number in st.session_state.last_saved_outlets:
                                            media_outlet = st.session_state.last_saved_outlets[wo_number]
                                            update_data['media_outlet'] = media_outlet
                                            
                                            # Get outlet_id and impressions from mapping
                                            if wo_number in st.session_state.outlet_data_mapping:
                                                outlet_data_dict = st.session_state.outlet_data_mapping[wo_number]
                                                if media_outlet in outlet_data_dict:
                                                    outlet_info = outlet_data_dict[media_outlet]
                                                    update_data['media_outlet_id'] = outlet_info.get('outlet_id', '')
                                                    update_data['impressions'] = outlet_info.get('impressions', 0)
                                        
                                        # Add Byline Author if available
                                        if wo_number in st.session_state.last_saved_bylines:
                                            update_data['byline_author'] = st.session_state.last_saved_bylines[wo_number]
                                        
                                        result = db.supabase.table('clips').update(update_data).eq('wo_number', wo_number).execute()
                                        
                                        if result.data:
                                            approved_clips.extend(result.data)
                                            # Debug: Print the first clip to see what fields we have
                                            if result.data:
                                                logger.info(f"Sample approved clip fields: {list(result.data[0].keys())}")
                                    
                                    logger.info(f"✅ Approved {len(approved_clips)} clips")
                                    
                                    # Extract trim for clips that don't have it
                                    from src.utils.trim_extractor import extract_trim_from_model
                                    
                                    for clip in approved_clips:
                                        if not clip.get('trim'):
                                            make = clip.get('make', '')
                                            model = clip.get('model', '')
                                            if model:
                                                base_model, extracted_trim = extract_trim_from_model(model, make)
                                                if extracted_trim:
                                                    # Update the clip with extracted trim
                                                    try:
                                                        db.supabase.table('clips').update({
                                                            'trim': extracted_trim,
                                                            'model': base_model  # Update to base model without trim
                                                        }).eq('wo_number', clip['wo_number']).execute()
                                                        
                                                        # Update the clip object for sentiment analysis
                                                        clip['trim'] = extracted_trim
                                                        clip['model'] = base_model
                                                        logger.info(f"Extracted trim '{extracted_trim}' for WO# {clip['wo_number']}")
                                                    except Exception as e:
                                                        logger.error(f"Failed to update trim for WO# {clip['wo_number']}: {e}")
                                    
                                    # Show progress bar for sentiment analysis
                                    st.info("🧠 Running sentiment analysis on approved clips...")
                                    progress_bar = st.progress(0)
                                    progress_text = st.empty()
                                    
                                    # Run sentiment analysis on approved clips
                                    
                                    def update_progress(progress, message):
                                        # Progress is already a fraction between 0 and 1
                                        progress_bar.progress(progress)
                                        progress_text.text(message)
                                    
                                    # Check if OpenAI API key is available
                                    if not os.environ.get('OPENAI_API_KEY'):
                                        st.error("❌ OpenAI API key not found. Clips approved but sentiment analysis skipped.")
                                        # Update clips to sentiment_analyzed without sentiment
                                        for clip in approved_clips:
                                            db.supabase.table('clips').update({
                                                'workflow_stage': 'sentiment_analyzed'
                                            }).eq('id', clip['id']).execute()
                                    else:
                                        # Run sentiment analysis
                                        try:
                                            results = run_sentiment_analysis(approved_clips, update_progress)
                                        except Exception as e:
                                            st.error(f"❌ Sentiment analysis error: {str(e)}")
                                            logger.error(f"Sentiment analysis failed: {e}")
                                            # Still move clips to ready to export even if sentiment fails
                                            for clip in approved_clips:
                                                db.supabase.table('clips').update({
                                                    'workflow_stage': 'sentiment_analyzed',
                                                    'sentiment_completed': False
                                                }).eq('id', clip['id']).execute()
                                            results = None
                                        
                                        # Update clips with sentiment results and move to ready_to_export
                                        sentiment_success_count = 0
                                        if results and 'results' in results:
                                            for clip, result in zip(approved_clips, results['results']):
                                                if result.get('sentiment_completed'):
                                                    success = db.update_clip_sentiment(clip['id'], result)
                                                    if success:
                                                        # Update workflow stage to sentiment_analyzed (which means ready to export)
                                                        db.supabase.table('clips').update({
                                                            'workflow_stage': 'sentiment_analyzed'
                                                        }).eq('id', clip['id']).execute()
                                                        sentiment_success_count += 1
                                                else:
                                                    # If sentiment failed, still move to sentiment_analyzed but note the failure
                                                    db.supabase.table('clips').update({
                                                        'workflow_stage': 'sentiment_analyzed',
                                                        'sentiment_completed': False
                                                    }).eq('id', clip['id']).execute()
                                        
                                        progress_bar.progress(1.0)
                                        progress_text.text(f"✅ Sentiment analysis complete! {sentiment_success_count}/{len(approved_clips)} successful")
                                    
                                    # Success message and cleanup
                                    st.success(f"✅ Successfully processed {len(approved_clips)} clips!")
                                    st.info("📋 **Clips are ready for export** in the Approved Queue")
                                    
                                    # Clear selections and dialog
                                    st.session_state.selected_for_approval = set()
                                    st.session_state.approved_records = set()
                                    st.session_state.show_approval_dialog = False
                                    
                                    # Clear saved checkbox state
                                    try:
                                        temp_file = os.path.join(project_root, "temp", "checkbox_state.pkl")
                                        if os.path.exists(temp_file):
                                            os.remove(temp_file)
                                    except Exception as e:
                                        print(f"Could not clear saved checkbox state: {e}")
                                    
                                    # Clear the cached approved queue data so it refreshes
                                    if 'get_approved_queue_data' in st.session_state:
                                        del st.session_state['get_approved_queue_data']
                                    st.cache_data.clear()
                                    
                                    # Refresh the page to update the Bulk Review table
                                    st.rerun()
                                    
                                except Exception as e:
                                    st.error(f"❌ Error approving clips in database: {e}")
                                    logger.error(f"Database approval error: {e}")
                    
                    with col_cancel:
                        if st.button("❌ Cancel", key="cancel_approval_btn"):
                            st.session_state.show_approval_dialog = False
                            st.rerun()
                
                # NEW: Rejection confirmation dialog
                if st.session_state.get('show_rejection_dialog', False):
                    st.markdown("---")
                    rejected_count = len(st.session_state.get('selected_for_rejection', set()))
                    st.error(f"⚠️ **Reject {rejected_count} clips?** This will move them to Rejected/Issues tab.")
                    
                    col_confirm, col_cancel = st.columns(2)
                    with col_confirm:
                        if st.button("❌ Confirm Rejection", type="secondary", key="confirm_rejection_btn"):
                            # Process the rejections
                            selected_rejected_wos = st.session_state.selected_for_rejection
                            if selected_rejected_wos:
                                try:
                                    # Get database instance
                                    db = get_cached_database()
                                    if not db:
                                        st.error("Database connection not available")
                                        st.session_state.show_rejection_dialog = False
                                        st.rerun()
                                        
                                    # Update clips in database to rejected status
                                    rejected_count = 0
                                    for wo_number in selected_rejected_wos:
                                        result = db.supabase.table('clips').update({
                                            'status': 'rejected',
                                            'failure_reason': 'Manual rejection by reviewer'
                                        }).eq('wo_number', wo_number).execute()
                                        
                                        if result.data:
                                            rejected_count += 1
                                            logger.info(f"✅ Rejected clip WO #{wo_number}")
                                        else:
                                            logger.warning(f"⚠️ Could not find clip WO #{wo_number} to reject")
                                    
                                    if rejected_count > 0:
                                        st.success(f"✅ Successfully rejected {rejected_count} clips!")
                                        st.info("📋 **Clips moved to Rejected/Issues tab**")
                                        
                                        # Clear the cache to force reload of updated data
                                        st.cache_data.clear()
                                    else:
                                        st.error("❌ No clips were rejected - they may not exist in the database")
                                    
                                    # Clear selections and dialog (both new and legacy tracking)
                                    st.session_state.selected_for_rejection = set()
                                    st.session_state.rejected_records = set()  # Clear new tracking too
                                    st.session_state.show_rejection_dialog = False
                                    
                                    # Clear saved checkbox state
                                    try:
                                        temp_file = os.path.join(project_root, "temp", "checkbox_state.pkl")
                                        if os.path.exists(temp_file):
                                            os.remove(temp_file)
                                    except Exception as e:
                                        print(f"Could not clear saved checkbox state: {e}")
                                    
                                    # Also clear any approved selections for the rejected items
                                    st.session_state.approved_records -= selected_rejected_wos
                                    st.session_state.selected_for_approval -= selected_rejected_wos
                                    
                                    # Immediate rerun to update the UI
                                    st.rerun()
                                    
                                except Exception as e:
                                    st.error(f"Error processing rejection: {e}")
                                    print(f"Rejection error: {e}")
                                    import traceback
                                    print(f"Full traceback: {traceback.format_exc()}")
                                    # Clear dialog even on error
                                    st.session_state.show_rejection_dialog = False
                                    st.rerun()  # Force rerun even on error
                    
                    with col_cancel:
                        if st.button("❌ Cancel Rejection", key="cancel_rejection_btn"):
                            st.session_state.show_rejection_dialog = False
                            st.rerun()

            else:
                st.info("No clips to review. Process loans first.")
        except Exception as e:
            st.error(f"Error loading clips: {e}")
    else:
        st.info("No results file found. Upload and process loans to begin.")

    # ========== WORKFLOW PROGRESS SECTION ==========
    # Show current workflow status
    if len(df) > 0:
        st.markdown("---")
        st.markdown("### 🔄 Workflow Status")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            pending_count = len(df)
            st.metric("📋 Pending Review", pending_count)
        
        with col2:
            try:
                approved_queue_clips = db.get_approved_queue_clips()
                approved_count = len(approved_queue_clips)
            except:
                approved_count = 0
            st.metric("✅ Approved Queue", approved_count)
        
        with col3:
            # This will be implemented in Phase 2
            st.metric("🧠 Sentiment Ready", "Phase 2")
        
        with col4:
            # This will be implemented in Phase 3
            st.metric("📤 Export Ready", "Phase 3")
        
        if approved_count > 0:
            st.info(f"💡 **Next Step:** Visit the **Approved Queue** tab to manage {approved_count} approved clips")
    
    # Add bottom padding to prevent UI elements from touching the bottom (CORRECTLY PLACED)
    st.markdown('<div style="height: 100px;"></div>', unsafe_allow_html=True)


# ========== APPROVED QUEUE TAB (Enhanced with FMS Export) ==========
with approved_queue_tab:
    st.markdown('<h4 style="margin-top: 0; margin-bottom: 0.5rem; font-size: 1.2rem; font-weight: 600; color: #2c3e50;">✅ Approved Queue</h4>', unsafe_allow_html=True)
    st.markdown('<p style="margin-top: 0; margin-bottom: 1rem; font-size: 0.9rem; color: #6c757d; font-style: italic;">Export clips with completed sentiment analysis to FMS</p>', unsafe_allow_html=True)
    
    # Initialize session state for workflow filtering
    if 'approved_queue_filter' not in st.session_state:
        st.session_state.approved_queue_filter = 'ready_to_export'
    
    # Workflow filtering tabs (updated)
    filter_col1, filter_col2 = st.columns(2)
    
    with filter_col1:
        if st.button("📤 Ready to Export", key="filter_ready_export", 
                    type="primary" if st.session_state.approved_queue_filter == 'ready_to_export' else "secondary"):
            st.session_state.approved_queue_filter = 'ready_to_export'
            st.rerun()
    
    with filter_col2:
        if st.button("✅ Recent Complete", key="filter_complete",
                    type="primary" if st.session_state.approved_queue_filter == 'recent_complete' else "secondary"):
            st.session_state.approved_queue_filter = 'recent_complete'
            st.rerun()
    
    # Add a container div to help with CSS targeting
    st.markdown('<div id="approved-queue-filters"></div>', unsafe_allow_html=True)
    
    # Custom CSS to style the Ready to Export button with light blue instead of red
    st.markdown("""
    <style>
    /* Custom styling for Approved Queue filter buttons */
    /* This targets all primary buttons but we'll make Ready to Export blue */
    
    /* First, override ALL primary buttons in this area to be blue */
    #approved-queue-filters + div button[kind="primary"] {
        background-color: #5b9bd5 !important;  /* Light blue */
        border-color: #5b9bd5 !important;
        color: white !important;
    }
    
    /* Target buttons that contain the export emoji */
    button:has(p:contains("📤")) {
        background-color: #5b9bd5 !important;
        border-color: #5b9bd5 !important;
    }
    
    /* Use the data-testid for columns approach */
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child button {
        background-color: #5b9bd5 !important;
        border-color: #5b9bd5 !important;
        color: white !important;
    }
    
    /* Alternative: Target all buttons and then override */
    .stButton > button {
        transition: all 0.2s ease;
    }
    
    /* Specific override for buttons with Ready to Export text */
    .stButton > button[aria-label*="Ready to Export"] {
        background-color: #5b9bd5 !important;
        border-color: #5b9bd5 !important;
    }
    
    /* Last resort: target by partial text match */
    button {
        position: relative;
    }
    
    button[kind="primary"]:not([aria-label*="Complete"]) {
        background-color: #5b9bd5 !important;
        border-color: #5b9bd5 !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Load clips based on selected filter
    try:
        # Use the global cached database connection
        db = get_cached_database()
        
        # DISABLED: Migration was causing performance issues
        # One-time migration should be run manually if needed, not on every page load
        
        # Cache the approved queue data with TTL of 60 seconds
        @st.cache_data
        def get_approved_queue_data():
            return db.get_approved_queue_clips()
        
        if st.session_state.approved_queue_filter == 'ready_to_export':
            # Get clips that are ready to export (workflow_stage = 'ready_to_export')
            @st.cache_data
            def get_ready_to_export_data():
                # OPTIMIZED: Exclude massive content fields to prevent 820MB data transfer
                needed_columns = 'id,wo_number,office,make,model,contact,media_outlet,activity_id,relevance_score,overall_sentiment,workflow_stage,processed_date,clip_url,published_date,sentiment_completed,sentiment_data_enhanced,marketing_impact_score,executive_summary,brand_narrative,sentiment_version'
                
                # Get clips with workflow_stage = 'sentiment_analyzed' (ready to export)
                result = db.supabase.table('clips').select(needed_columns).eq('workflow_stage', 'sentiment_analyzed').execute()
                
                # Also get any legacy clips that are approved with sentiment completed but not exported
                legacy_result = db.supabase.table('clips').select(needed_columns).eq('status', 'approved').eq('sentiment_completed', True).eq('workflow_stage', 'found').execute()
                
                all_clips = result.data if result.data else []
                if legacy_result.data:
                    all_clips.extend(legacy_result.data)
                return all_clips
            
            clips_data = get_ready_to_export_data()
            tab_title = "📤 Ready to Export"
            tab_description = "Clips with completed sentiment analysis ready for FMS export"
            
        elif st.session_state.approved_queue_filter == 'recent_complete':
            # Get exported clips from the last 30 days - OPTIMIZED
            from datetime import datetime, timedelta
            thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
            
            with st.spinner("Loading recent complete clips..."):
                try:
                    # OPTIMIZATION 1: Select only needed columns, not '*'
                    needed_columns = 'id,wo_number,office,make,model,contact,media_outlet,relevance_score,overall_sentiment,workflow_stage,fms_export_date,processed_date,clip_url,published_date,sentiment_completed,sentiment_data_enhanced'
                    
                    # OPTIMIZATION 2: Use database filtering with proper ordering and smaller limit
                    # First get clips with fms_export_date (most recent exports)
                    recent_exports = db.supabase.table('clips').select(needed_columns).in_('workflow_stage', ['exported', 'complete']).gte('fms_export_date', thirty_days_ago).order('fms_export_date', desc=True).limit(100).execute()
                    
                    clips_data = recent_exports.data if recent_exports.data else []
                    
                    # OPTIMIZATION 3: Only check legacy clips if we need more records
                    if len(clips_data) < 50:  # Only if we have fewer than 50 recent clips
                        # Get some legacy clips without fms_export_date but with recent processed_date
                        legacy_clips = db.supabase.table('clips').select(needed_columns).in_('workflow_stage', ['exported', 'complete']).is_('fms_export_date', 'null').gte('processed_date', thirty_days_ago).order('processed_date', desc=True).limit(50).execute()
                        
                        if legacy_clips.data:
                            clips_data.extend(legacy_clips.data)
                                
                except Exception as e:
                    st.error(f"Error loading recent complete data: {str(e)}")
                    logger.error(f"Recent complete query error: {e}")
                    clips_data = []
            tab_title = "✅ Recent Complete (Last 30 Days)"
            tab_description = "Exported clips from the last 30 days"
        
        # Display current filter info
        st.markdown(f'<h5 style="margin-top: 0.5rem; margin-bottom: 0.3rem; font-size: 1.1rem; font-weight: 600; color: #2c3e50;">{tab_title}</h5>', unsafe_allow_html=True)
        st.markdown(f'<p style="margin-top: 0; margin-bottom: 1rem; font-size: 0.85rem; color: #6c757d; font-style: italic;">{tab_description}</p>', unsafe_allow_html=True)
        
        if clips_data:
            # Convert to DataFrame for display
            approved_df = pd.DataFrame(clips_data)
            
            # Ensure WO # is treated as string
            if 'wo_number' in approved_df.columns:
                approved_df['wo_number'] = approved_df['wo_number'].astype(str)
            
            # Quick stats overview
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Clips", len(approved_df))
            with col2:
                avg_score = approved_df['relevance_score'].mean() if 'relevance_score' in approved_df.columns and not approved_df.empty else 0
                st.metric("Avg Relevance", f"{avg_score:.1f}/10")
            with col3:
                unique_contacts = approved_df['contact'].nunique() if 'contact' in approved_df.columns else 0
                st.metric("Media Contacts", unique_contacts)
            with col4:
                unique_outlets = approved_df['media_outlet'].nunique() if 'media_outlet' in approved_df.columns else 0
                st.metric("Media Outlets", unique_outlets)
            
            # Create display DataFrame
            clean_df = pd.DataFrame()
            # IMPORTANT: Include the database ID for sentiment updates
            clean_df['id'] = approved_df['id'] if 'id' in approved_df.columns else ''
            clean_df['WO #'] = approved_df['wo_number'] if 'wo_number' in approved_df.columns else ''
            clean_df['Office'] = approved_df['office'] if 'office' in approved_df.columns else ''
            clean_df['Make'] = approved_df['make'] if 'make' in approved_df.columns else ''
            clean_df['Model'] = approved_df['model'] if 'model' in approved_df.columns else ''
            clean_df['Contact'] = approved_df['contact'] if 'contact' in approved_df.columns else ''
            clean_df['Media Outlet'] = approved_df['media_outlet'] if 'media_outlet' in approved_df.columns else ''
            clean_df['Relevance'] = approved_df['relevance_score'].apply(lambda x: f"{x}/10" if pd.notna(x) and x != 'N/A' else 'N/A') if 'relevance_score' in approved_df.columns else 'N/A'
            # Show published date (not processed date) - format as MMM DD, YYYY
            if 'published_date' in approved_df.columns:
                clean_df['Date'] = approved_df['published_date'].apply(
                    lambda x: pd.to_datetime(x).strftime('%b %d, %Y') if pd.notna(x) else ''
                )
            else:
                clean_df['Date'] = ''
            
            # Add View column for URLs
            clean_df['Clip URL'] = approved_df['clip_url'] if 'clip_url' in approved_df.columns else ''
            clean_df['📄 View'] = clean_df['Clip URL']
            
            # Add activity_id as a hidden column for hyperlink functionality
            clean_df['Activity_ID'] = approved_df['activity_id'] if 'activity_id' in approved_df.columns else ''
            
            # Add sentiment status indicator - check both old and new sentiment fields
            def get_sentiment_status(row):
                # Check if enhanced sentiment data exists (new method)
                if pd.notna(row.get('sentiment_data_enhanced')) and row.get('sentiment_data_enhanced'):
                    return "✅ Complete"
                # Check legacy sentiment_completed field (old method)
                elif row.get('sentiment_completed', False):
                    return "✅ Complete"
                else:
                    return "⏳ Pending"
            
            clean_df['Sentiment'] = approved_df.apply(get_sentiment_status, axis=1)
            
            # Add export date for Recent Complete tab
            if st.session_state.approved_queue_filter == 'recent_complete' and 'fms_export_date' in approved_df.columns:
                # Format export date with better handling of null values
                def format_export_date(date_val):
                    if pd.isna(date_val) or date_val is None:
                        return "Not recorded"
                    try:
                        return pd.to_datetime(date_val).strftime('%b %d %I:%M %p')
                    except:
                        return "Invalid date"
                
                clean_df['FMS Export Date'] = approved_df['fms_export_date'].apply(format_export_date)
                
                # Add export status indicator
                clean_df['Export Status'] = approved_df.apply(
                    lambda row: "✅ Exported to FMS" if pd.notna(row.get('fms_export_date')) else "❓ Status Unknown",
                    axis=1
                )
            
            # Add workflow status indicator
            clean_df['Stage'] = approved_df['workflow_stage'].apply(
                lambda x: "📤 Ready to Export" if x == 'sentiment_analyzed' else 
                         "✅ Exported" if x == 'exported' else 
                         "📊 Complete" if x == 'complete' else
                         "🧠 Processing" if x == 'found' else
                         f"📋 {x.replace('_', ' ').title()}"
            ) if 'workflow_stage' in approved_df.columns else 'Unknown'
            
            # Add FMS Export Status for Ready to Export tab
            if st.session_state.approved_queue_filter == 'ready_to_export':
                clean_df['Export Status'] = approved_df.apply(
                    lambda row: "✅ Exported to FMS" if pd.notna(row.get('fms_export_date')) else "📤 Ready",
                    axis=1
                )
            
            # Configure ADVANCED AgGrid for approved queue (same as Bulk Review)
            # Lazy render: limit rows initially to speed first paint, allow full view on demand
            max_initial_rows = 300
            initial_df = clean_df.head(max_initial_rows)
            gb = GridOptionsBuilder.from_dataframe(initial_df)
            
            # *** ADVANCED FEATURES WITH SET FILTERS (CHECKBOXES) ***
            gb.configure_side_bar()  # Enable filtering sidebar
            gb.configure_default_column(
                filter="agSetColumnFilter",  # CHECKBOX FILTERS with search
                sortable=True,  # Enable sorting
                resizable=True,  # Enable column resizing
                editable=False, 
                groupable=True, 
                value=True, 
                enableRowGroup=True, 
                enablePivot=True, 
                enableValue=True,
                filterParams={
                    "buttons": ["reset", "apply"],
                    "closeOnApply": True,
                    "newRowsAction": "keep"
                }
            )
            
            # Create cell renderer for WO # column with hyperlink to FMS activity
            cellRenderer_wo = JsCode("""
            class WoCellRenderer {
              init(params) {
                const woNumber = params.value;
                const activityId = params.data['Activity_ID'];
                
                this.eGui = document.createElement('div');
                
                if (activityId && activityId !== '') {
                  // Create hyperlink to FMS activity
                  this.link = document.createElement('a');
                  this.link.innerText = woNumber;
                  this.link.href = `https://fms.driveshop.com/activities/edit/${activityId}`;
                  this.link.target = '_blank';
                  this.link.style.color = '#1f77b4';
                  this.link.style.textDecoration = 'underline';
                  this.link.style.cursor = 'pointer';
                  this.eGui.appendChild(this.link);
                } else {
                  // No activity ID, just show the WO number as plain text
                  this.eGui.innerText = woNumber;
                }
              }

              getGui() {
                return this.eGui;
              }

              refresh(params) {
                return false;
              }
            }
            """)
            
            # Enable selection for batch operations (only for Ready to Export)
            if st.session_state.approved_queue_filter == 'ready_to_export':
                gb.configure_selection('multiple', use_checkbox=True, groupSelectsChildren=True, groupSelectsFiltered=True)
                # Add checkbox selection to first column for Ready to Export
                gb.configure_column("WO #", minWidth=100, width=140, pinned='left', checkboxSelection=True, headerCheckboxSelection=True, cellRenderer=cellRenderer_wo)
            else:
                # Recent Complete is read-only
                gb.configure_selection('single', use_checkbox=False)
                gb.configure_column("WO #", minWidth=100, width=140, pinned='left', cellRenderer=cellRenderer_wo)
            gb.configure_column("Office", minWidth=115)
            gb.configure_column("Make", minWidth=120)
            gb.configure_column("Model", minWidth=150)
            gb.configure_column("Contact", minWidth=180, sortable=True, sort='asc')
            gb.configure_column("Media Outlet", minWidth=220)
            gb.configure_column("Relevance", minWidth=110)
            gb.configure_column("Date", minWidth=100)
            gb.configure_column("Sentiment", minWidth=140)
            gb.configure_column("Stage", minWidth=120)
            
            # Configure Export Status column if present
            if st.session_state.approved_queue_filter == 'ready_to_export' and 'Export Status' in clean_df.columns:
                gb.configure_column("Export Status", minWidth=160)
            
            # Hide raw URL column and database ID
            gb.configure_column("Clip URL", hide=True)
            gb.configure_column("id", hide=True)  # Hide database ID but keep in data
            gb.configure_column("Activity_ID", hide=True)  # Hide the activity ID column
            
            # Configure View column with URL renderer (same as Bulk Review)
            cellRenderer_view = JsCode("""
            class UrlCellRenderer {
              init(params) {
                this.eGui = document.createElement('a');
                this.eGui.innerText = '📄 View';
                this.eGui.href = params.data['Clip URL'];
                this.eGui.target = '_blank';
                this.eGui.style.color = '#1f77b4';
                this.eGui.style.textDecoration = 'underline';
                this.eGui.style.cursor = 'pointer';
              }

              getGui() {
                return this.eGui;
              }

              refresh(params) {
                return false;
              }
            }
            """)
            
            gb.configure_column(
                "📄 View", 
                cellRenderer=cellRenderer_view,
                minWidth=100,
                sortable=False,
                filter=False
            )
            
            # Build and display grid with ADVANCED features
            grid_options = gb.build()
            
            # Force column definitions exactly like Bulk Review
            base_columns = [
                {'field': 'WO #', 'headerName': 'Work Order #', 'cellRenderer': cellRenderer_wo, 'minWidth': 140, 'checkboxSelection': True, 'headerCheckboxSelection': True} if st.session_state.approved_queue_filter == 'ready_to_export' else {'field': 'WO #', 'headerName': 'Work Order #', 'cellRenderer': cellRenderer_wo, 'minWidth': 100},
                {'field': 'Office', 'minWidth': 115},
                {'field': 'Make', 'minWidth': 120},
                {'field': 'Model', 'minWidth': 150},
                {'field': 'Contact', 'minWidth': 180, 'sortable': True},
                {'field': 'Media Outlet', 'minWidth': 220},
                {'field': 'Relevance', 'minWidth': 110},
                {'field': 'Date', 'minWidth': 100},
                {'field': 'Sentiment', 'minWidth': 140},
                {'field': 'Stage', 'minWidth': 120},
                {'field': '📄 View', 'cellRenderer': cellRenderer_view, 'minWidth': 100, 'sortable': False, 'filter': False},
                {'field': 'Clip URL', 'hide': True},
                {'field': 'id', 'hide': True},
                {'field': 'Activity_ID', 'hide': True}
            ]
            
            # Add tab-specific columns
            if st.session_state.approved_queue_filter == 'ready_to_export' and 'Export Status' in clean_df.columns:
                base_columns.append({'field': 'Export Status', 'minWidth': 160})
            elif st.session_state.approved_queue_filter == 'recent_complete' and 'FMS Export Date' in clean_df.columns:
                base_columns.append({'field': 'FMS Export Date', 'minWidth': 140, 'sortable': True})
            
            grid_options['columnDefs'] = base_columns
            
            # Set default sorting based on tab
            grid_options['defaultColDef'] = grid_options.get('defaultColDef', {})
            grid_options['defaultColDef']['sortable'] = True
            if st.session_state.approved_queue_filter == 'recent_complete':
                # Sort by FMS Export Date descending for Recent Complete
                grid_options['sortModel'] = [
                    {'colId': 'FMS Export Date', 'sort': 'desc'}
                ]
            else:
                # Sort by Contact ascending for Ready to Export
                grid_options['sortModel'] = [
                    {'colId': 'Contact', 'sort': 'asc'}
                ]
            
            selected_clips = AgGrid(
                clean_df,
                gridOptions=grid_options,
                allow_unsafe_jscode=True,
                update_mode=GridUpdateMode.SELECTION_CHANGED,
                height=400,  # Match Bulk Review height
                fit_columns_on_grid_load=True,
                columns_auto_size_mode='FIT_ALL_COLUMNS_TO_VIEW',  # Auto-size all columns
                theme="alpine",
                enable_enterprise_modules=True  # REQUIRED for Set Filters with checkboxes
            )
            
            # Action buttons based on current filter
            st.markdown("---")
            
            if st.session_state.approved_queue_filter == 'ready_to_export':
                # Actions for Ready to Export
                # Count selected rows from AgGrid
                selected_count = len(selected_clips.selected_rows) if hasattr(selected_clips, 'selected_rows') and selected_clips.selected_rows is not None else 0
                
                col1, col2, col3, col4 = st.columns([1, 1.5, 1.5, 1])
                
                with col1:
                    st.metric("Selected", f"{selected_count}/{len(clean_df)}")
                
                with col2:
                    if st.button(f"📤 Export Options ({selected_count})", 
                                 disabled=selected_count == 0, 
                                 help="Export selected clips - Download JSON or Send to FMS",
                                 use_container_width=True,
                                 type="primary"):
                        # Handle FMS Export
                        # Get selected rows from AgGrid response
                        selected_rows = []
                        if hasattr(selected_clips, 'selected_rows'):
                            selected_data = selected_clips.selected_rows
                            if selected_data is not None:
                                if hasattr(selected_data, 'to_dict'):
                                    selected_rows = selected_data.to_dict('records')
                                elif isinstance(selected_data, list):
                                    selected_rows = selected_data
                                else:
                                    selected_rows = []
                        
                        if selected_rows and len(selected_rows) > 0:
                            try:
                                clips_to_export = []
                                
                                # Gather full clip data for export
                                for row in selected_rows:
                                    clip_id = row.get('id')
                                    if clip_id:
                                        # Get full clip data from database using ID
                                        clip_data = next((clip for clip in clips_data if clip['id'] == clip_id), None)
                                        if clip_data:
                                            clips_to_export.append(clip_data)
                                    else:
                                        # Fallback to WO number if ID not available
                                        wo_number = str(row.get('WO #', ''))
                                        if wo_number:
                                            clip_data = next((clip for clip in clips_data if str(clip['wo_number']) == wo_number), None)
                                            if clip_data:
                                                clips_to_export.append(clip_data)
                                
                                if clips_to_export:
                                    # Create FMS export data with all fields including sentiment
                                    fms_export_data = []
                                    export_timestamp = datetime.now().isoformat()
                                    
                                    # Get export data from clips_export view with client field names
                                    clip_ids = [clip['id'] for clip in clips_to_export if clip.get('id')]
                                    
                                    if clip_ids:
                                        # Query the export view for these specific clips to get original flat fields
                                        export_result = db.supabase.table('clips_export').select('*').in_('activity_id', [clip['activity_id'] for clip in clips_to_export if clip.get('activity_id')]).execute()
                                        
                                        if export_result.data:
                                            # Get enhanced sentiment data for the same clips
                                            clips_result = db.supabase.table('clips').select('wo_number, sentiment_data_enhanced').in_('id', clip_ids).execute()
                                            
                                            # Create a lookup for enhanced sentiment data by wo_number
                                            enhanced_data_lookup = {}
                                            if clips_result.data:
                                                for clip_data in clips_result.data:
                                                    enhanced_sentiment = {}
                                                    if clip_data.get('sentiment_data_enhanced'):
                                                        try:
                                                            if isinstance(clip_data['sentiment_data_enhanced'], str):
                                                                enhanced_sentiment = json.loads(clip_data['sentiment_data_enhanced'])
                                                            else:
                                                                enhanced_sentiment = clip_data['sentiment_data_enhanced']
                                                        except (json.JSONDecodeError, TypeError):
                                                            enhanced_sentiment = {}
                                                    
                                                    # Filter to only include the fields specified by client
                                                    allowed_fields = {
                                                        'sentiment_classification', 'key_features_mentioned', 'brand_attributes_captured',
                                                        'purchase_drivers', 'competitive_context', 'trim_level_mentioned', 'trim_impact_score',
                                                        'trim_highlights', 'vehicle_identifier', 'content_type', 'overall_sentiment',
                                                        'relevance_score', 'summary', 'brand_alignment'
                                                    }
                                                    
                                                    filtered_sentiment = {k: v for k, v in enhanced_sentiment.items() if k in allowed_fields}
                                                    enhanced_data_lookup[clip_data.get('wo_number')] = filtered_sentiment

                                            # Create lookup from clips_to_export for media_outlet (VIEW doesn't have this column)
                                            media_outlet_lookup = {clip['activity_id']: clip.get('media_outlet') for clip in clips_to_export if clip.get('activity_id')}

                                            # Combine original flat data with enhanced sentiment data
                                            for export_record in export_result.data:
                                                wo_number = export_record.get('wo_number')

                                                # Fix publication field using clips_to_export lookup
                                                if 'publication' not in export_record or not export_record.get('publication'):
                                                    activity_id = export_record.get('activity_id')
                                                    if activity_id in media_outlet_lookup:
                                                        export_record['publication'] = media_outlet_lookup[activity_id]

                                                if wo_number in enhanced_data_lookup:
                                                    export_record['sentiment_data_enhanced'] = enhanced_data_lookup[wo_number]
                                                else:
                                                    # Add empty enhanced sentiment for clips that failed analysis
                                                    export_record['sentiment_data_enhanced'] = {
                                                        "sentiment_classification": {"overall": "neutral", "confidence": 0.0, "rationale": "Analysis failed"},
                                                        "key_features_mentioned": [],
                                                        "brand_attributes_captured": [],
                                                        "purchase_drivers": [],
                                                        "competitive_context": {"direct_comparisons": [], "market_positioning": ""},
                                                        "trim_level_mentioned": False,
                                                        "trim_impact_score": 0.0,
                                                        "trim_highlights": None,
                                                        "vehicle_identifier": f"{export_record.get('make', '')} {export_record.get('model', '')}".strip(),
                                                        "content_type": "Web Article",
                                                        "overall_sentiment": "neutral",
                                                        "relevance_score": 5,
                                                        "summary": "Enhanced sentiment analysis failed for this clip",
                                                        "brand_alignment": False
                                                    }
                                                fms_export_data.append(export_record)
                                        else:
                                            # Fallback to manual mapping if view query fails
                                            for clip in clips_to_export:
                                                # Parse the enhanced sentiment data
                                                enhanced_sentiment = {}
                                                if clip.get('sentiment_data_enhanced'):
                                                    try:
                                                        if isinstance(clip['sentiment_data_enhanced'], str):
                                                            enhanced_sentiment = json.loads(clip['sentiment_data_enhanced'])
                                                        else:
                                                            enhanced_sentiment = clip['sentiment_data_enhanced']
                                                    except (json.JSONDecodeError, TypeError):
                                                        enhanced_sentiment = {}
                                                
                                                # Filter to only include the fields specified by client
                                                allowed_fields = {
                                                    'sentiment_classification', 'key_features_mentioned', 'brand_attributes_captured',
                                                    'purchase_drivers', 'competitive_context', 'trim_level_mentioned', 'trim_impact_score',
                                                    'trim_highlights', 'vehicle_identifier', 'content_type', 'overall_sentiment',
                                                    'relevance_score', 'summary', 'brand_alignment'
                                                }
                                                
                                                filtered_sentiment = {k: v for k, v in enhanced_sentiment.items() if k in allowed_fields}
                                                
                                                # Create export record with original flat fields PLUS enhanced sentiment
                                                # Use placeholder data if enhanced sentiment failed
                                                if not filtered_sentiment:
                                                    filtered_sentiment = {
                                                        "sentiment_classification": {"overall": "neutral", "confidence": 0.0, "rationale": "Analysis failed"},
                                                        "key_features_mentioned": [],
                                                        "brand_attributes_captured": [],
                                                        "purchase_drivers": [],
                                                        "competitive_context": {"direct_comparisons": [], "market_positioning": ""},
                                                        "trim_level_mentioned": False,
                                                        "trim_impact_score": 0.0,
                                                        "trim_highlights": None,
                                                        "vehicle_identifier": f"{clip.get('make', '')} {clip.get('model', '')}".strip(),
                                                        "content_type": "Web Article",
                                                        "overall_sentiment": "neutral", 
                                                        "relevance_score": 5,
                                                        "summary": "Enhanced sentiment analysis failed for this clip",
                                                        "brand_alignment": False
                                                    }
                                                
                                                export_record = {
                                                    # Client-requested fields with their preferred names (original flat structure)
                                                    "activity_id": clip.get('activity_id'),
                                                    "brand_fit": clip.get('brand_narrative'),
                                                    "byline": clip.get('byline_author'),
                                                    "link": clip.get('clip_url'),
                                                    "cons": clip.get('cons'),
                                                    "impressions": clip.get('impressions'),
                                                    "publication": clip.get('media_outlet'),
                                                    "publication_id": clip.get('media_outlet_id'),
                                                    "overall_score": clip.get('overall_score'),
                                                    "sentiment": clip.get('overall_sentiment'),
                                                    "pros": clip.get('pros'),
                                                    "date": clip.get('published_date'),
                                                    "relevance_score": clip.get('relevance_score'),
                                                    "ai_summary": clip.get('summary'),
                                                    
                                                    # NEW: Enhanced sentiment data as additional field
                                                    "sentiment_data_enhanced": filtered_sentiment
                                                }
                                                fms_export_data.append(export_record)
                                    
                                    # Create JSON export
                                    import json
                                    export_json = json.dumps(fms_export_data, indent=2, default=str)
                                    filename = f"fms_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                                    
                                    # Store export data in session state
                                    st.session_state.fms_export_ready = True
                                    st.session_state.fms_export_json = export_json
                                    st.session_state.fms_export_filename = filename
                                    st.session_state.fms_clips_to_export = clips_to_export
                                    st.session_state.fms_export_timestamp = export_timestamp
                                    
                                    st.success(f"✅ Export prepared for {len(clips_to_export)} clips!")
                                    st.rerun()
                                    
                                else:
                                    st.error("❌ No valid clips found for export")
                            except Exception as e:
                                st.error(f"❌ Error during FMS export: {e}")
                                logger.error(f"FMS export error: {e}")
                        else:
                            st.warning("Please select clips for FMS export")
                
                with col3:
                    if st.button(f"↩️ Move to Bulk Review ({selected_count})",
                                 disabled=selected_count == 0,
                                 help="Move selected clips back to Bulk Review",
                                 use_container_width=True):
                        # Handle moving clips back to Bulk Review
                        selected_rows = []
                        if hasattr(selected_clips, 'selected_rows'):
                            selected_data = selected_clips.selected_rows
                            if selected_data is not None:
                                if hasattr(selected_data, 'to_dict'):
                                    selected_rows = selected_data.to_dict('records')
                                elif isinstance(selected_data, list):
                                    selected_rows = selected_data
                                else:
                                    selected_rows = []
                        
                        if selected_rows and len(selected_rows) > 0:
                            try:
                                moved_count = 0
                                for row in selected_rows:
                                    clip_id = row.get('id')
                                    if clip_id:
                                        # Update workflow stage back to 'found'
                                        result = db.supabase.table('clips').update({
                                            'workflow_stage': 'found',  # Use valid workflow stage
                                            'status': 'pending_review',  # Use valid status value
                                            'overall_sentiment': None,
                                            'pros': None,
                                            'cons': None,
                                            'overall_score': None,
                                            'relevance_score': None,
                                            'brand_narrative': None,
                                            'summary': None
                                        }).eq('id', clip_id).execute()
                                        
                                        if result.data:
                                            moved_count += 1
                                
                                if moved_count > 0:
                                    st.success(f"✅ Moved {moved_count} clips back to Bulk Review!")
                                    st.cache_data.clear()
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to move clips")
                            except Exception as e:
                                st.error(f"❌ Error moving clips: {e}")
                                logger.error(f"Move to bulk review error: {e}")
                        else:
                            st.warning("Please select clips to move")
                
                with col4:
                    # Count clips with pending sentiment (sentiment_completed = False)
                    pending_sentiment_count = len([clip for clip in clips_data if not clip.get('sentiment_completed', False)])
                    
                    # Count selected clips with pending sentiment
                    selected_pending_count = 0
                    if selected_count > 0 and hasattr(selected_clips, 'selected_rows'):
                        selected_data = selected_clips.selected_rows
                        if selected_data is not None:
                            if hasattr(selected_data, 'to_dict'):
                                selected_rows = selected_data.to_dict('records')
                            elif isinstance(selected_data, list):
                                selected_rows = selected_data
                            else:
                                selected_rows = []
                            
                            # Count selected clips with pending sentiment
                            for row in selected_rows:
                                if row.get('Sentiment') == "⏳ Pending":
                                    selected_pending_count += 1
                    
                    # Show button with count
                    button_text = f"🧠 Run Sentiment ({selected_pending_count})" if selected_pending_count > 0 else f"🧠 Run Sentiment ({pending_sentiment_count} pending)"
                    button_disabled = pending_sentiment_count == 0 or (selected_count > 0 and selected_pending_count == 0)
                    
                    if st.button(button_text,
                                 disabled=button_disabled,
                                 help="Run sentiment analysis on clips with pending sentiment",
                                 use_container_width=True):
                        # Get clips to analyze
                        clips_to_analyze = []
                        
                        if selected_count > 0 and selected_pending_count > 0:
                            # Analyze only selected clips with pending sentiment
                            for row in selected_rows:
                                if row.get('Sentiment') == "⏳ Pending":
                                    clip_id = row.get('id')
                                    if clip_id:
                                        clip_data = next((clip for clip in clips_data if clip['id'] == clip_id), None)
                                        if clip_data and not clip_data.get('sentiment_completed', False):
                                            clips_to_analyze.append(clip_data)
                        else:
                            # Analyze all clips with pending sentiment
                            clips_to_analyze = [clip for clip in clips_data if not clip.get('sentiment_completed', False)]
                        
                        if clips_to_analyze:
                            # Show progress bar for sentiment analysis
                            st.info(f"🧠 Running sentiment analysis on {len(clips_to_analyze)} clips...")
                            progress_bar = st.progress(0)
                            progress_text = st.empty()
                            
                            def update_progress(progress, message):
                                progress_bar.progress(progress)
                                progress_text.text(message)
                            
                            # Check if OpenAI API key is available
                            if not os.environ.get('OPENAI_API_KEY'):
                                st.error("❌ OpenAI API key not found. Clips cannot be analyzed without API key.")
                            else:
                                # Run sentiment analysis
                                try:
                                    results = run_sentiment_analysis(clips_to_analyze, update_progress)
                                    
                                    # Process results
                                    success_count = 0
                                    if results and 'results' in results:
                                        for clip, result in zip(clips_to_analyze, results['results']):
                                            if result.get('sentiment_completed'):
                                                success = db.update_clip_sentiment(clip['id'], result)
                                                if success:
                                                    success_count += 1
                                    
                                    progress_bar.progress(1.0)
                                    progress_text.text(f"✅ Sentiment analysis complete! {success_count}/{len(clips_to_analyze)} successful")
                                    
                                    if success_count > 0:
                                        st.success(f"✅ Successfully analyzed {success_count} clips!")
                                        # Clear cache to refresh the table
                                        st.cache_data.clear()
                                        time.sleep(2)
                                        st.rerun()
                                    else:
                                        st.error("❌ No clips were successfully analyzed")
                                        
                                except Exception as e:
                                    st.error(f"❌ Sentiment analysis error: {str(e)}")
                                    logger.error(f"Sentiment analysis failed: {e}")
                        else:
                            st.warning("No clips with pending sentiment found")
            
            # Show download button if export is ready
            if st.session_state.get('fms_export_ready', False):
                st.markdown("---")
                
                # Show export options
                st.markdown("### 📤 Export Options")
                
                col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
                
                with col1:
                    if st.download_button(
                        label="📥 Download JSON",
                        data=st.session_state.fms_export_json,
                        file_name=st.session_state.fms_export_filename,
                        mime="application/json",
                        key="download_fms_json",
                        help="Download the JSON file to your computer",
                        use_container_width=True
                    ):
                        # Just show success message - don't move to exported
                        st.success("✅ JSON file downloaded successfully!")
                        st.info("💡 Clips remain in Ready to Export. Use 'Mark as Complete' when finished.")
                        
                with col2:
                    if st.button("🚀 Send to FMS", key="send_to_fms_api", help="Send clips directly to FMS API", type="primary", use_container_width=True):
                        print("🔥 BUTTON CLICKED: Send to FMS button was pressed!")
                        try:
                            print("🔥 INITIALIZING: Creating FMS API client...")
                            # Initialize FMS API client
                            fms_client = FMSAPIClient()
                            print(f"🔥 CLIENT CREATED: FMS client initialized for {fms_client.environment}")
                            
                            # Parse the JSON data
                            print("🔥 PARSING: Getting JSON data from session state...")
                            clips_data_raw = json.loads(st.session_state.fms_export_json)
                            print(f"🔥 RAW DATA TYPE: {type(clips_data_raw)}")
                            print(f"🔥 RAW DATA SAMPLE: {str(clips_data_raw)[:200]}...")
                            
                            # Handle both formats: list or dict with 'clips' key
                            if isinstance(clips_data_raw, list):
                                clips_data = clips_data_raw  # Already a list of clips
                                print(f"🔥 PARSED: Got {len(clips_data)} clips from list")
                            elif isinstance(clips_data_raw, dict) and 'clips' in clips_data_raw:
                                clips_data = clips_data_raw['clips']  # Extract clips from dict
                                print(f"🔥 PARSED: Got {len(clips_data)} clips from dict")
                            else:
                                print(f"🔥 ERROR: Unexpected data format: {type(clips_data_raw)}")
                                st.error("Invalid clips data format")
                                clips_data = []
                            
                            # Send to FMS API with automatic token rotation
                            print("🔥 SENDING: About to call FMS API...")
                            with st.spinner("Sending clips to FMS API..."):
                                result = fms_client.send_clips_with_retry(clips_data)
                            print(f"🔥 RESULT: FMS API returned: {result}")
                            
                            if result["success"]:
                                # Check if token was rotated during the process
                                if result.get("token_rotated"):
                                    st.warning(f"🔑 Token was automatically rotated during export!")
                                    st.info(f"**Action Required:** Update Render environment variable:\n`FMS_API_TOKEN={result.get('new_token', 'ERROR_GETTING_TOKEN')}`")
                                
                                # Mark clips as sent to FMS based on actual API response
                                clips_to_export = st.session_state.fms_clips_to_export
                                export_timestamp = st.session_state.fms_export_timestamp
                                sent_count = result['sent_count']
                                total_clips = len(clips_to_export)
                                all_clips_sent = result.get('all_clips_sent', sent_count == total_clips)
                                
                                successfully_sent_clips = []
                                marked_count = 0
                                
                                if all_clips_sent:
                                    # All clips were processed successfully by FMS
                                    for clip in clips_to_export:
                                        update_result = db.supabase.table('clips').update({
                                            'fms_export_date': export_timestamp,
                                            'workflow_stage': 'sentiment_analyzed'  # Keep in Ready to Export
                                        }).eq('id', clip['id']).execute()
                                        
                                        if update_result.data:
                                            marked_count += 1
                                            successfully_sent_clips.append(clip)
                                            
                                    st.success(f"✅ Successfully sent all {sent_count} clips to FMS API!")
                                    
                                else:
                                    # Partial success - only mark the number that FMS actually processed
                                    # Since we don't know which specific clips failed, mark the first N as successful
                                    st.warning(f"⚠️ FMS API processed {sent_count} of {total_clips} clips. {total_clips - sent_count} clips may have failed validation.")
                                    
                                    for i, clip in enumerate(clips_to_export):
                                        if i < sent_count:  # Mark first N clips as sent
                                            update_result = db.supabase.table('clips').update({
                                                'fms_export_date': export_timestamp,
                                                'workflow_stage': 'sentiment_analyzed'
                                            }).eq('id', clip['id']).execute()
                                            
                                            if update_result.data:
                                                marked_count += 1
                                                successfully_sent_clips.append(clip)
                                    
                                    st.warning(f"⚠️ Only {sent_count} of {total_clips} clips were successfully sent to FMS!")
                                    st.info(f"💡 The remaining {total_clips - sent_count} clips remain in Ready to Export for retry.")
                                
                                # Store only successfully processed clips for Mark Complete
                                st.session_state.fms_successfully_sent_clips = successfully_sent_clips
                                
                                # Response details removed for cleaner UI
                                        
                                if marked_count > 0:
                                    st.info(f"📊 {marked_count} clips marked as 'Sent to FMS' - they remain in Ready to Export until marked complete")
                                
                                # Store success state for confirmation dialog
                                st.session_state.fms_send_successful = True
                                st.session_state.fms_result = result
                                
                                # Clear cache to refresh the table
                                st.cache_data.clear()
                            else:
                                st.error(f"❌ Failed to send to FMS API: {result.get('error', 'Unknown error')}")
                                if result.get('validation_errors'):
                                    st.error("Validation errors:")
                                    for error in result['validation_errors']:
                                        st.error(f"  • {error}")
                                        
                        except Exception as e:
                            print(f"🔥 EXCEPTION: Error in FMS send: {str(e)}")
                            st.error(f"❌ Error sending to FMS API: {str(e)}")
                            logger.error(f"FMS API send error: {e}", exc_info=True)
                            
                with col3:
                    if st.button("❌ Cancel", key="cancel_export", use_container_width=True):
                        # Clear session state without updating clips
                        st.session_state.fms_export_ready = False
                        st.session_state.fms_export_json = None
                        st.session_state.fms_export_filename = None
                        st.session_state.fms_clips_to_export = None
                        st.session_state.fms_export_timestamp = None
                        st.info("Export cancelled")
                        st.rerun()
                        
                # Add Mark as Complete button
                with col4:
                    # Only show Mark Complete if clips were successfully sent
                    clips_to_complete = st.session_state.get('fms_successfully_sent_clips', [])
                    mark_complete_disabled = len(clips_to_complete) == 0
                    
                    if st.button("✅ Mark as Complete", key="mark_complete", use_container_width=True,
                                 disabled=mark_complete_disabled,
                                 help="Move successfully exported clips to Recent Complete"):
                        # Update only successfully sent clips to exported status
                        export_timestamp = st.session_state.fms_export_timestamp
                        exported_count = 0
                        
                        for clip in clips_to_complete:
                            result = db.supabase.table('clips').update({
                                'workflow_stage': 'exported',
                                'fms_export_date': export_timestamp
                            }).eq('id', clip['id']).execute()
                            
                            if result.data:
                                exported_count += 1
                        
                        # Clear session state
                        st.session_state.fms_export_ready = False
                        st.session_state.fms_export_json = None
                        st.session_state.fms_export_filename = None
                        st.session_state.fms_clips_to_export = None
                        st.session_state.fms_export_timestamp = None
                        st.session_state.fms_send_successful = False
                        st.session_state.fms_result = None
                        
                        # Show success and refresh
                        st.success(f"✅ Moved {exported_count} clips to Recent Complete!")
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                        
                # Show API environment info
                st.markdown("---")
                api_env = os.getenv("FMS_API_ENVIRONMENT", "staging")
                if api_env == "staging":
                    st.info(f"🔧 FMS API Environment: **{api_env.upper()}** (Testing)")
                else:
                    st.warning(f"🚨 FMS API Environment: **{api_env.upper()}** (Live System)")
                    
                # Show confirmation dialog if FMS send was successful
                if st.session_state.get('fms_send_successful', False):
                    st.markdown("---")
                    st.markdown("### 🎉 FMS Send Successful!")
                    st.info("Would you like to mark these clips as complete?")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("Yes, Mark Complete", key="confirm_complete", type="primary", use_container_width=True):
                            # Update only successfully sent clips to exported status
                            clips_to_complete = st.session_state.get('fms_successfully_sent_clips', [])
                            export_timestamp = st.session_state.fms_export_timestamp
                            exported_count = 0
                            
                            for clip in clips_to_complete:
                                result = db.supabase.table('clips').update({
                                    'workflow_stage': 'exported',
                                    'fms_export_date': export_timestamp
                                }).eq('id', clip['id']).execute()
                                
                                if result.data:
                                    exported_count += 1
                            
                            # Clear session state
                            st.session_state.fms_export_ready = False
                            st.session_state.fms_export_json = None
                            st.session_state.fms_export_filename = None
                            st.session_state.fms_clips_to_export = None
                            st.session_state.fms_export_timestamp = None
                            st.session_state.fms_send_successful = False
                            st.session_state.fms_result = None
                            
                            # Show success and refresh
                            st.success(f"✅ Moved {exported_count} clips to Recent Complete!")
                            st.cache_data.clear()
                            time.sleep(1)
                            st.rerun()
                            
                    with col2:
                        if st.button("No, Keep in Export", key="keep_export", use_container_width=True):
                            st.session_state.fms_send_successful = False
                            st.info("Clips remain in Ready to Export")
                            st.rerun()
        
        else:
            # No clips found for current filter
            if st.session_state.approved_queue_filter == 'ready_to_export':
                st.info("📤 No clips ready for export. Approve clips in Bulk Review to see them here after sentiment analysis.")
            elif st.session_state.approved_queue_filter == 'recent_complete':
                st.info("✅ No exported clips in the last 30 days.")
            
            # Show helpful instructions
            st.markdown("""
            **Updated Workflow:**
            1. **📋 Bulk Review** → Select and submit approved clips
            2. **🧠 Automatic Sentiment** → Runs immediately after approval
            3. **📤 Ready to Export** → Select clips and export to FMS JSON
            4. **✅ Recent Complete** → View exported clips from last 30 days
            
            **Key Features:**
            - Sentiment analysis runs automatically on approval
            - All fields including strategic intelligence are exported
            - Exported clips automatically move to Recent Complete
            """)
    
    except Exception as e:
        st.error(f"❌ Error loading approved queue: {e}")
        import traceback
        st.error(f"Full error: {traceback.format_exc()}")
    
    # Add bottom padding to prevent UI elements from being cut off
    st.markdown('<div style="height: 150px;"></div>', unsafe_allow_html=True)


# ========== REJECTED/ISSUES TAB (Enhanced with Current Run + Historical + Date Range) ==========
with rejected_tab:
    # Initialize session state for filtering mode
    if 'rejected_view_mode' not in st.session_state:
        st.session_state.rejected_view_mode = 'current_run'  # Default to current run
    
    # Compact Filtering Controls - Single Row
    col_toggle, col_dates, col_info = st.columns([1, 2, 2])
    
    with col_toggle:
        if st.session_state.rejected_view_mode == 'current_run':
            if st.button("📊 Show Historical", key="show_historical", help="View all historical failed attempts"):
                st.session_state.rejected_view_mode = 'historical'
                st.rerun()
        else:
            if st.button("🔄 Current Run", key="show_current", help="View only the most recent processing run"):
                st.session_state.rejected_view_mode = 'current_run'
                st.rerun()
    
    # Date range filtering (only shown in historical mode)
    start_date = None
    end_date = None
    
    with col_dates:
        if st.session_state.rejected_view_mode == 'historical':
            col_start, col_end = st.columns(2)
            with col_start:
                # Default to 30 days ago
                from datetime import datetime, timedelta
                default_start = datetime.now() - timedelta(days=30)
                start_date = st.date_input("From", value=default_start, key="rejected_start_date")
            with col_end:
                # Default to today
                default_end = datetime.now()
                end_date = st.date_input("To", value=default_end, key="rejected_end_date")
    
    with col_info:
        if st.session_state.rejected_view_mode == 'current_run':
            st.caption("🔄 Current run failures only")
        else:
            date_range = ""
            if start_date and end_date:
                date_range = f" ({start_date} to {end_date})"
            st.caption(f"📊 Historical view{date_range}")
    
    # Load rejected clips and failed processing attempts from database
    try:
        # Get cached database connection
        db = get_cached_database()
        
        # Cache the current run data to avoid duplicate calls
        @st.cache_data(ttl=300)  # Cache for 5 minutes
        def get_current_run_data():
            """Get current run data with caching to prevent duplicate DB calls
            
            Args:
                none
            """
            local_db = get_cached_database()
            latest_id = local_db.get_latest_processing_run_id()
            failed_clips = local_db.get_current_run_failed_clips()
            return latest_id, failed_clips
        
        # Choose data source based on mode
        if st.session_state.rejected_view_mode == 'current_run':
            # Current run mode - get only the most recent processing run
            latest_run_id, current_run_failed_clips = get_current_run_data()
            
            # Also get clips that were skipped in the current run
            current_run_skipped_clips = []
            if latest_run_id:
                @st.cache_data(ttl=300)
                def get_skipped_clips_by_run(run_id: int):
                    # Conservative projection (avoid non-existent columns)
                    projection = 'wo_number, office, make, model, contact, media_outlet, processed_date, original_urls, urls_attempted, failure_reason, last_skip_run_id'
                    return (
                        db.supabase
                        .table('clips')
                        .select(projection)
                        .eq('last_skip_run_id', run_id)
                        .execute()
                    ).data or []
                current_run_skipped_clips = get_skipped_clips_by_run(latest_run_id)
            
            # Get run info for display
            if (current_run_failed_clips or current_run_skipped_clips) and latest_run_id:
                run_info = db.get_processing_run_info(latest_run_id)
                
                if run_info:
                    run_name = run_info.get('run_name', 'Unknown')
                    run_date = run_info.get('start_time', 'Unknown')[:19] if run_info.get('start_time') else 'Unknown'  # Truncate timestamp
                    st.caption(f"🔄 **{run_name}** - {run_date}")
            
            # Convert to combined issues format
            combined_issues = []
            for clip in current_run_failed_clips:
                # Add retry status to rejection reason if in cooldown
                rejection_reason = 'No Content Found' if clip['status'] == 'no_content_found' else 'Processing Failed'
                # Some deployments may not have retry_status/ retry_after columns
                if clip.get('retry_status') == 'in_cooldown':
                    retry_date = clip.get('retry_after', '')
                    if retry_date:
                        # Format the retry date for display
                        try:
                            from datetime import datetime
                            retry_dt = datetime.fromisoformat(retry_date.replace('Z', '+00:00'))
                            retry_str = retry_dt.strftime('%Y-%m-%d %H:%M')
                            rejection_reason += f" (Retry after {retry_str})"
                        except:
                            rejection_reason += " (In retry cooldown)"
                
                combined_issues.append({
                    'WO #': clip['wo_number'],
                    'Office': clip.get('office', ''),
                    'Make': clip.get('make', ''),
                    'Model': clip.get('model', ''),
                    'To': clip.get('contact', ''),
                    'Affiliation': clip.get('media_outlet', ''),
                    'Rejection_Reason': rejection_reason,
                    'URL_Details': f"Processed with {clip.get('tier_used', 'Unknown')}",
                    'Processed_Date': clip.get('processed_date', ''),
                    'Type': 'No Content Found' if clip['status'] == 'no_content_found' else 'Processing Failed',
                    'original_urls': clip.get('original_urls', ''),
                    'urls_attempted': clip.get('urls_attempted', 0),
                    'failure_reason': clip.get('failure_reason', ''),
                    # These fields may not exist in all schemas; keep optional
                    'retry_status': clip.get('retry_status', ''),
                    'attempt_count': clip.get('attempt_count', 1)
                })
            
            # Add skipped clips from current run
            for clip in current_run_skipped_clips:
                skip_reason = clip.get('skip_reason', 'unknown')
                rejection_reason = {
                    'already_approved': 'Already Approved',
                    'already_pending_review': 'Already in Bulk Review',
                    'already_rejected': 'Already User Rejected',
                    'retry_cooldown': 'In Retry Cooldown',
                    'unknown': 'Unknown Skip Reason'
                }.get(skip_reason, skip_reason)
                
                combined_issues.append({
                    'WO #': clip['wo_number'],
                    'Office': clip.get('office', ''),
                    'Make': clip.get('make', ''),
                    'Model': clip.get('model', ''),
                    'To': clip.get('contact', ''),
                    'Affiliation': clip.get('media_outlet', ''),
                    'Rejection_Reason': f'Skipped: {rejection_reason}',
                    'URL_Details': f"Skipped in current run",
                    'Processed_Date': clip.get('processed_date', ''),
                    'Type': 'Skipped',
                    'original_urls': clip.get('original_urls', ''),
                    'urls_attempted': clip.get('urls_attempted', 0),
                    'failure_reason': skip_reason,
                    'retry_status': 'skipped',
                    'attempt_count': clip.get('attempt_count', 0)
                })
            
        else:
            # Historical mode - get all failed clips with optional date filtering
            start_date_str = start_date.strftime('%Y-%m-%d') if start_date else None
            end_date_str = end_date.strftime('%Y-%m-%d') if end_date else None
            
            @st.cache_data(ttl=300)
            def cached_get_all_failed_clips(start_date: str|None, end_date: str|None):
                local_db = get_cached_database()
                return local_db.get_all_failed_clips(start_date=start_date, end_date=end_date)

            all_failed_clips = cached_get_all_failed_clips(
                start_date=start_date_str,
                end_date=end_date_str
            )
            
            # Display date range info compactly
            if start_date_str and end_date_str:
                st.caption(f"📊 **Historical:** {start_date_str} to {end_date_str}")
            elif start_date_str:
                st.caption(f"📊 **Historical:** From {start_date_str}")
            elif end_date_str:
                st.caption(f"📊 **Historical:** Until {end_date_str}")
            else:
                st.caption("📊 **Historical:** All Time")
            
            # Convert to combined issues format
            combined_issues = []
            for clip in all_failed_clips:
                combined_issues.append({
                    'WO #': clip['wo_number'],
                    'Office': clip.get('office', ''),
                    'Make': clip.get('make', ''),
                    'Model': clip.get('model', ''),
                    'To': clip.get('contact', ''),
                    'Affiliation': clip.get('media_outlet', ''),
                    'Rejection_Reason': 'No Content Found' if clip['status'] == 'no_content_found' else 'Processing Failed',
                    'URL_Details': f"Processed with {clip.get('tier_used', 'Unknown')}",
                    'Processed_Date': clip.get('processed_date', ''),
                    'Type': 'No Content Found' if clip['status'] == 'no_content_found' else 'Processing Failed',
                    'original_urls': clip.get('original_urls', ''),
                    'urls_attempted': clip.get('urls_attempted', 0),
                    'failure_reason': clip.get('failure_reason', '')
                })
        
        # Add manually rejected clips based on view mode
        if st.session_state.rejected_view_mode == 'current_run':
            # For current run, only include rejected clips from the current run
            if latest_run_id:
                @st.cache_data
                def cached_get_rejected_clips_by_run(run_id):
                    db = get_cached_database()
                    return db.get_rejected_clips(run_id=run_id)
                
                rejected_clips = cached_get_rejected_clips_by_run(latest_run_id)
            else:
                rejected_clips = []
        else:
            # For historical view, get all rejected clips (with optional date filtering)
            @st.cache_data(ttl=300)
            def cached_get_rejected_clips_hist():
                local_db = get_cached_database()
                return local_db.get_rejected_clips()
            
            rejected_clips = cached_get_rejected_clips_hist()
        
        for clip in rejected_clips:
                combined_issues.append({
                'WO #': clip['wo_number'],
                'Office': clip.get('office', ''),
                'Make': clip.get('make', ''),
                'Model': clip.get('model', ''),
                    'To': clip.get('contact', ''),
                    'Affiliation': clip.get('media_outlet', ''),
                'Rejection_Reason': 'User Rejected Clip',
                'URL_Details': clip.get('clip_url', ''),
                'Processed_Date': clip.get('processed_date', ''),
                'Type': 'Rejected Clip',
                'original_urls': '',  # Rejected clips don't have original URLs
                'urls_attempted': 0,
                'failure_reason': ''
            })
        
        # Create DataFrame from combined issues
        if combined_issues:
            rejected_df = pd.DataFrame(combined_issues)
        else:
            rejected_df = pd.DataFrame()
            
    except Exception as e:
        st.error(f"❌ Error loading rejected clips from database: {e}")
        rejected_df = pd.DataFrame()  # Empty DataFrame on error
    
    # Process results
    if len(rejected_df) > 0:
        # Ensure WO # is treated as string for consistency
        if 'WO #' in rejected_df.columns:
            rejected_df['WO #'] = rejected_df['WO #'].astype(str)
        
        if not rejected_df.empty:
            # Compact metrics in a single row
            st.markdown(f"""
            <div style="background-color: #f0f2f6; padding: 8px; border-radius: 4px; margin: 8px 0;">
                <small>
                    📝 <strong>{len(rejected_df)}</strong> rejected  •  
                    🚫 <strong>{rejected_df['Rejection_Reason'].value_counts().index[0] if 'Rejection_Reason' in rejected_df.columns and len(rejected_df) > 0 else 'N/A'}</strong> top issue  •  
                    ⚡ <strong>{len(rejected_df[rejected_df['Rejection_Reason'].str.contains('No Content Found|Processing Failed', case=False, na=False)] if 'Rejection_Reason' in rejected_df.columns else [])}/{len(rejected_df)}</strong> technical failures  •  
                    ⏭️ <strong>{len(rejected_df[rejected_df['Type'] == 'Skipped'] if 'Type' in rejected_df.columns else [])}</strong> skipped
                </small>
            </div>
            """, unsafe_allow_html=True)
            
            # Create AgGrid table (same format as bulk review but for rejected records)
            clean_df = rejected_df.copy()
            
            # Prepare columns for display
            if 'WO #' in clean_df.columns:
                clean_df['WO #'] = clean_df['WO #'].astype(str)
            
            # Add Office column if it exists
            if 'Office' in rejected_df.columns:
                pass  # Keep original name
            
            # Extract the searched URL from database fields for the View column  
            def extract_searched_url(row):
                """
                Extract the original source URL(s) from database fields for the View link.
                Handles both single and multiple URLs per WO# intelligently.
                """
                # NEW: Try the database original_urls field first
                original_urls = row.get('original_urls', '')
                if original_urls and not pd.isna(original_urls):
                    original_urls = str(original_urls).strip()
                    if original_urls:
                        # Handle multiple URLs separated by semicolon
                        urls = [url.strip() for url in original_urls.split(';') if url.strip()]
                        if urls:
                            # If multiple URLs, return the first one (most common case)
                            # We'll show a count indicator if there are multiple
                            first_url = urls[0]
                            return first_url
                
                # FALLBACK: Try the old URL_Details field for backward compatibility
                url_details = row.get('URL_Details', '')
                if pd.isna(url_details) or not url_details:
                    return ""
                
                url_details_str = str(url_details).strip()
                
                # First try JSON parsing (new format)
                try:
                    import json
                    url_data = json.loads(url_details_str)
                    
                    if isinstance(url_data, list) and len(url_data) > 0:
                        # For rejected records, we want the ORIGINAL source URL (not the found clip)
                        first_entry = url_data[0]
                        if isinstance(first_entry, dict) and 'original_url' in first_entry:
                            original_url = first_entry['original_url']
                            return original_url
                except (json.JSONDecodeError, KeyError, TypeError):
                    # Not JSON format, continue to old string parsing
                    pass
                
                # Handle OLD string format used by rejected records
                # Format: "https://example.com: status; https://example2.com: status"
                
                # Split by semicolon to get individual URL entries
                url_entries = url_details_str.split(';')
                
                for entry in url_entries:
                    entry = entry.strip()
                    if entry:
                        # Split by first colon to separate URL from status
                        if ':' in entry:
                            # Find the URL part (everything before " :")
                            colon_pos = entry.find(': ')
                            if colon_pos > 0:
                                url_part = entry[:colon_pos].strip()
                                # Validate it looks like a URL
                                if url_part.startswith(('http://', 'https://')):
                                    return url_part
                        else:
                            # No colon, treat entire entry as URL
                            if entry.startswith(('http://', 'https://')):
                                return entry
                
                # Final fallback - return empty string if no parsing worked
                return ""
            
            # Add the View column with the searched URL (using row-based function)
            clean_df['Searched URL'] = clean_df.apply(extract_searched_url, axis=1)
            clean_df['📄 View'] = clean_df['Searched URL']  # Create View column for cellRenderer
            
            # Rename columns for better display
            column_mapping = {
                'Office': 'Office',  # Add office column first
                'WO #': 'WO #',
                'Model': 'Model', 
                'To': 'Media Contact',
                'Affiliation': 'Publication',
                '📄 View': '📄 View',  # Add View column
                'Rejection_Reason': '⚠️ Rejection Reason',
                'URL_Details': '📋 Details',
                'Processed_Date': '📅 Processed',
                # Include new fields but don't display them (for JavaScript access)
                'urls_attempted': 'urls_attempted',
                'failure_reason': 'failure_reason'
            }
            
            # Only keep columns that exist and separate display vs hidden columns
            display_columns = []
            hidden_columns = ['urls_attempted', 'failure_reason']  # Keep these for JavaScript but don't display
            
            for old_col, new_col in column_mapping.items():
                if old_col in clean_df.columns:
                    if old_col != new_col:
                        clean_df = clean_df.rename(columns={old_col: new_col})
                    if new_col not in hidden_columns:
                        display_columns.append(new_col)
            
            # Create cellRenderer for View column with multiple URL indicator
            cellRenderer_view = JsCode("""
            class UrlCellRenderer {
              init(params) {
                const urls_attempted = params.data['urls_attempted'] || 0;
                
                this.eGui = document.createElement('div');
                this.eGui.style.display = 'flex';
                this.eGui.style.alignItems = 'center';
                this.eGui.style.gap = '5px';
                
                // Create the link
                this.link = document.createElement('a');
                this.link.innerText = '📄 View';
                this.link.href = params.value;
                this.link.target = '_blank';
                this.link.style.color = '#1f77b4';
                this.link.style.textDecoration = 'underline';
                this.link.style.cursor = 'pointer';
                
                this.eGui.appendChild(this.link);
                
                // Add multiple URL indicator if more than 1 URL was attempted
                if (urls_attempted > 1) {
                  const indicator = document.createElement('span');
                  indicator.innerText = `(+${urls_attempted - 1})`;
                  indicator.style.fontSize = '11px';
                  indicator.style.color = '#6c757d';
                  indicator.style.fontStyle = 'italic';
                  indicator.title = `${urls_attempted} URLs were searched for this WO#`;
                  this.eGui.appendChild(indicator);
                }
              }

              getGui() {
                return this.eGui;
              }

              refresh(params) {
                return false;
              }
            }
            """)
            

            # Configure AgGrid for rejected records (no pagination for full transparency)
            gb = GridOptionsBuilder.from_dataframe(clean_df)
            
            # Hide columns that shouldn't be displayed
            for hidden_col in hidden_columns:
                if hidden_col in clean_df.columns:
                    gb.configure_column(hidden_col, hide=True)
            # Enable row selection for moving records back to Bulk Review
            gb.configure_selection('multiple', use_checkbox=True, groupSelectsChildren=True, groupSelectsFiltered=True)
            # Removed pagination to show all rejected records at once
            gb.configure_side_bar()
            gb.configure_default_column(
                filter="agSetColumnFilter",  # CHECKBOX FILTERS with search
                sortable=True,  # Enable sorting
                resizable=True,  # Enable column resizing
                editable=False, 
                groupable=True, 
                value=True, 
                enableRowGroup=True, 
                enablePivot=True, 
                enableValue=True,
                filterParams={
                    "buttons": ["reset", "apply"],
                    "closeOnApply": True,
                    "newRowsAction": "keep"
                }
            )
            
            # Configure columns with auto-sizing (no fixed widths) for balanced layout
            if "Office" in display_columns:
                gb.configure_column("Office", pinned='left')  # Keep pinned but auto-size
            
            # Configure WO # column with proper filter for string type
            gb.configure_column(
                "WO #", 
                pinned='left',
                filter='agTextColumnFilter',  # Use text filter for string columns
                filterParams={
                    'filterOptions': ['contains', 'notContains', 'equals', 'notEqual', 'startsWith', 'endsWith'],
                    'defaultOption': 'contains',
                    'caseSensitive': False
                }
            )
            
            # Configure the View column with the custom renderer (same as bulk review) 
            if "📄 View" in display_columns:
                gb.configure_column(
                    "📄 View", 
                    cellRenderer=cellRenderer_view,
                    minWidth=80,
                    maxWidth=100,
                    sortable=False,
                    filter=False
                )
                # Hide the Searched URL column since it's only used for the cellRenderer
                gb.configure_column("Searched URL", hide=True)
            
            # Configure text columns with wrapping but no fixed width
            gb.configure_column("⚠️ Rejection Reason", wrapText=True, autoHeight=True)
            gb.configure_column("📋 Details", wrapText=True, autoHeight=True)
            
            # Enable auto-sizing for all columns
            gb.configure_grid_options(
                autoSizeStrategy={
                    'type': 'fitGridWidth',
                    'defaultMinWidth': 100,
                    'columnLimits': [
                        {'key': 'Office', 'minWidth': 80},
                        {'key': 'WO #', 'minWidth': 80},
                        {'key': 'Model', 'minWidth': 120},
                        {'key': 'Media Contact', 'minWidth': 120},
                        {'key': 'Publication', 'minWidth': 120},
                        {'key': '📄 View', 'minWidth': 80, 'maxWidth': 100},
                        {'key': '⚠️ Rejection Reason', 'minWidth': 150},
                        {'key': '📋 Details', 'minWidth': 200},
                        {'key': '📅 Processed', 'minWidth': 100}
                    ]
                }
            )
            
            # Build grid options
            grid_options = gb.build()
            
            # Implement lazy loading - start with first 50 rows for performance
            max_initial_rows = 50
            initial_df = clean_df.head(max_initial_rows) if len(clean_df) > max_initial_rows else clean_df
            
            # Display AgGrid table for rejected records
            st.markdown("**📋 Rejected Records**")
            selected_rejected = AgGrid(
                initial_df,
                gridOptions=grid_options,
                height=400,
                width='100%',
                data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                update_mode=GridUpdateMode.SELECTION_CHANGED,
                fit_columns_on_grid_load=True,  # Enable auto-sizing on load
                theme='streamlit',
                enable_enterprise_modules=True,
                allow_unsafe_jscode=True,
                reload_data=False,  # Disable auto-reload for performance
                columns_auto_size_mode='FIT_ALL_COLUMNS_TO_VIEW'  # Enable auto-sizing
            )

            # Option to load all rows (may be heavy)
            if len(clean_df) > max_initial_rows:
                if st.button(f"Load all {len(clean_df)} records", key="load_all_rejected"):
                    gb_full = GridOptionsBuilder.from_dataframe(clean_df)
                    # Reuse same column configurations
                    for hidden_col in hidden_columns:
                        if hidden_col in clean_df.columns:
                            gb_full.configure_column(hidden_col, hide=True)
                    gb_full.configure_selection('multiple', use_checkbox=True, groupSelectsChildren=True, groupSelectsFiltered=True)
                    gb_full.configure_side_bar()
                    grid_options_full = gb_full.build()
                    AgGrid(
                        clean_df,
                        gridOptions=grid_options_full,
                        height=500,
                        width='100%',
                        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                        update_mode=GridUpdateMode.SELECTION_CHANGED,
                        fit_columns_on_grid_load=True,
                        theme='streamlit',
                        enable_enterprise_modules=True,
                        allow_unsafe_jscode=True,
                        reload_data=True,
                        columns_auto_size_mode='FIT_ALL_COLUMNS_TO_VIEW'
                    )
            
            # Optional: Add functionality to move selected rejected records back to Bulk Review
            if st.button("🔄 Move Selected to Bulk Review", key="move_to_bulk_review"):
                selected_rows = selected_rejected.get('selected_rows', [])
                if selected_rows is not None and len(selected_rows) > 0:
                    # Convert to list if it's a DataFrame
                    if hasattr(selected_rows, 'to_dict'):
                        selected_rows = selected_rows.to_dict('records')
                    
                    # Get database instance
                    db = get_cached_database()
                    if not db:
                        st.error("Database connection not available")
                    else:
                        moved_count = 0
                        for row in selected_rows:
                            wo_number = str(row.get('WO #', ''))
                            if wo_number:
                                # Update status back to pending_review
                                result = db.supabase.table('clips').update({
                                    'status': 'pending_review',
                                    'failure_reason': None  # Clear the rejection reason
                                }).eq('wo_number', wo_number).execute()
                                
                                if result.data:
                                    moved_count += 1
                                    logger.info(f"✅ Moved WO #{wo_number} back to pending review")
                        
                        if moved_count > 0:
                            st.success(f"✅ Moved {moved_count} clips back to Bulk Review")
                            # Clear the cache to force reload of updated data
                            st.cache_data.clear()
                            # Use a small delay before rerun to ensure success message is visible
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("❌ Could not move any clips - they may not exist in the database")
                else:
                    st.warning("No records selected")
            
            # Add Excel Export functionality
            st.markdown("---")
            col_export1, col_export2 = st.columns([1, 3])
            
            with col_export1:
                # Dynamic button text based on mode
                if st.session_state.rejected_view_mode == 'historical' and start_date and end_date:
                    button_text = f"📥 Export {start_date.strftime('%m/%d')} - {end_date.strftime('%m/%d')}"
                    button_help = f"Export records from {start_date} to {end_date} to Excel with clickable URLs"
                else:
                    button_text = "📥 Export to Excel"
                    button_help = "Export current rejected records to Excel with clickable URLs"
                
                if st.button(button_text, key="export_rejected_excel", help=button_help):
                    # Create Excel file with rejected records
                    try:
                        # Get the data to export
                        export_df = rejected_df.copy()
                        
                        # Create a workbook and worksheet
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Rejected Records"
                        
                        # Define header style
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Define the columns to export
                        export_columns = ['WO #', 'Make', 'Model', 'To', 'Office', 'Rejection_Reason', 'Processed_Date']
                        
                        # Write headers
                        for col_idx, col_name in enumerate(export_columns, 1):
                            cell = ws.cell(row=1, column=col_idx, value=col_name)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Write data rows using DataFrame values directly
                        for row_idx in range(len(export_df)):
                            for col_idx, col_name in enumerate(export_columns, 1):
                                value = export_df.iloc[row_idx][col_name] if col_name in export_df.columns else ''
                                ws.cell(row=row_idx + 2, column=col_idx, value=str(value) if value else '')
                        
                        # After writing the main columns, add URL columns
                        # Get unique URLs for each WO
                        url_start_col = len(export_columns) + 1
                        
                        # Add URL headers
                        ws.cell(row=1, column=url_start_col, value="URLs").font = header_font
                        ws.cell(row=1, column=url_start_col, value="URLs").fill = header_fill
                        ws.cell(row=1, column=url_start_col, value="URLs").alignment = header_alignment
                        
                        # Process URLs for each row
                        for row_idx in range(len(export_df)):
                            # Get original URLs from the row
                            original_urls = export_df.iloc[row_idx].get('original_urls', '') if 'original_urls' in export_df.columns else ''
                            if original_urls and not pd.isna(original_urls):
                                urls = [url.strip() for url in str(original_urls).split(';') if url.strip()]
                                
                                # Write each URL as a hyperlink in consecutive columns
                                for url_idx, url in enumerate(urls):
                                    col_idx = url_start_col + url_idx
                                    
                                    # Ensure header exists for this URL column
                                    if row_idx == 0:  # First data row
                                        header_cell = ws.cell(row=1, column=col_idx, value=f"URL {url_idx + 1}")
                                        header_cell.font = header_font
                                        header_cell.fill = header_fill
                                        header_cell.alignment = header_alignment
                                    
                                    # Create hyperlink
                                    cell = ws.cell(row=row_idx + 2, column=col_idx)
                                    cell.value = f"Link {url_idx + 1}"
                                    cell.hyperlink = url
                                    cell.font = Font(color="0000FF", underline="single")
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        # Save to BytesIO
                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_buffer.seek(0)
                        
                        # Generate filename with date range
                        if st.session_state.rejected_view_mode == 'historical' and start_date and end_date:
                            filename = f"rejected_records_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
                        else:
                            filename = f"rejected_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        
                        # Provide download button
                        st.download_button(
                            label="📥 Download Excel File",
                            data=excel_buffer.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.success(f"✅ Excel file ready for download with {len(export_df)} records")
                        
                    except Exception as e:
                        st.error(f"❌ Error creating Excel file: {str(e)}")
                        import traceback
                        st.error(traceback.format_exc())
            
            with col_export2:
                if st.session_state.rejected_view_mode == 'historical' and start_date and end_date:
                    st.info(f"📅 Export will include records from {start_date} to {end_date}")
                else:
                    st.info("📋 Export will include all visible rejected records")
                    
    else:
        st.info("📊 No rejected records found")

# ========== STRATEGIC INTELLIGENCE TAB (Single Column Layout) ==========
with analysis_tab:
    # Single column layout for Strategic Intelligence
    try:
        # Use the already cached database instance instead of getting a new one
        # db is already available from the main cached database instance

        # Cached lightweight fetch (projection + limit) to speed initial render
        @st.cache_data(ttl=300, show_spinner=False)
        def load_sentiment_clips():
            projection = 'wo_number, make, model, contact, media_outlet, overall_sentiment, published_date, sentiment_data_enhanced, clip_url'
            result = (
                db.supabase
                .table('clips')
                .select(projection)
                .eq('sentiment_completed', True)
                .order('published_date', desc=True)
                .limit(500)
                .execute()
            )
            return result.data if result.data else []

        # Optional refresh to bust the cache explicitly
        refresh_col1, _ = st.columns([1, 5])
        with refresh_col1:
            if st.button('🔄 Refresh Data', key='refresh_strategic_cache'):
                load_sentiment_clips.clear()
                st.rerun()

        with st.spinner('Loading clips with sentiment...'):
            sentiment_clips = load_sentiment_clips()
        
        # Use the new single column display
        display_strategic_intelligence_tab(sentiment_clips)
        
    except Exception as e:
        st.error(f"Error loading Strategic Intelligence: {e}")
        logger.error(f"Strategic Intelligence tab error: {e}")

# Remove all old Strategic Intelligence code - it's now handled by display_strategic_intelligence_tab
# The old code from here to the export tab is no longer needed

# ========== MESSAGE PULL-THROUGH TAB ==========
with pullthrough_tab:
    display_pullthrough_analysis_tab()

# ========== OEM MESSAGING TAB ==========
with oem_tab:
    display_oem_messaging_tab()

# ========== HISTORICAL RE-PROCESSING TAB ==========
with reprocess_tab:
    # The function handles its own database connection
    display_historical_reprocessing_tab()

# ========== COOLDOWN MANAGEMENT TAB ==========
with cooldown_tab:
    # Display the cooldown management interface
    display_cooldown_management_tab()

# ========== EXPORT TAB ==========
with export_tab:
    st.markdown('<h4 style="margin-top: 0; margin-bottom: 0.5rem; font-size: 1.2rem; font-weight: 600; color: #2c3e50;">📊 Export Dashboard</h4>', unsafe_allow_html=True)
    st.markdown('<p style="margin-top: 0; margin-bottom: 1rem; font-size: 0.9rem; color: #6c757d; font-style: italic;">Export clips to Excel with custom filters and date ranges</p>', unsafe_allow_html=True)
    
    try:
        # Use the global cached database connection
        db = get_cached_database()
        
        # Create filter columns
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col1:
            # Date range picker
            st.markdown("**📅 Date Range**")
            date_range = st.date_input(
                "Select date range",
                value=(datetime.now() - timedelta(days=30), datetime.now()),
                key="export_date_range",
                help="Filter clips by processed date"
            )
            
            if len(date_range) == 2:
                start_date, end_date = date_range
            else:
                start_date = end_date = date_range[0]
        
        with col2:
            # Status filter
            st.markdown("**📋 Status Filter**")
            status_options = ["All", "approved", "exported", "found", "sentiment_analyzed"]
            selected_status = st.selectbox(
                "Select status",
                options=status_options,
                key="export_status_filter"
            )
        
        with col3:
            # Workflow stage filter
            st.markdown("**🔄 Workflow Stage**")
            workflow_options = ["All", "approved", "exported", "found", "sentiment_analyzed", "complete"]
            selected_workflow = st.selectbox(
                "Select workflow stage",
                options=workflow_options,
                key="export_workflow_filter"
            )
        
        # Advanced filters in expander
        with st.expander("🔧 Advanced Filters", expanded=False):
            adv_col1, adv_col2, adv_col3 = st.columns(3)
            
            with adv_col1:
                min_relevance = st.number_input(
                    "Min Relevance Score",
                    min_value=0,
                    max_value=10,
                    value=0,
                    key="export_min_relevance"
                )
            
            with adv_col2:
                sentiment_filter = st.selectbox(
                    "Sentiment",
                    options=["All", "POS", "NEU", "NEG"],
                    key="export_sentiment_filter"
                )
            
            with adv_col3:
                office_filter = st.text_input(
                    "Office (comma-separated)",
                    placeholder="e.g., San Francisco, Dallas",
                    key="export_office_filter"
                )
        
        # Query button
        if st.button("🔍 Query Database", type="primary", key="export_query_btn"):
            with st.spinner("Querying database..."):
                # Build query
                query = db.supabase.table('clips').select('*')
                
                # Apply date range filter
                start_datetime = datetime.combine(start_date, datetime.min.time()).isoformat()
                end_datetime = datetime.combine(end_date, datetime.max.time()).isoformat()
                query = query.gte('processed_date', start_datetime).lte('processed_date', end_datetime)
                
                # Apply status filter
                if selected_status != "All":
                    query = query.eq('status', selected_status)
                
                # Apply workflow filter
                if selected_workflow != "All":
                    query = query.eq('workflow_stage', selected_workflow)
                
                # Apply relevance filter
                if min_relevance > 0:
                    query = query.gte('relevance_score', min_relevance)
                
                # Apply sentiment filter
                if sentiment_filter != "All":
                    query = query.eq('overall_sentiment', sentiment_filter)
                
                # Apply office filter
                if office_filter.strip():
                    offices = [o.strip() for o in office_filter.split(',')]
                    query = query.in_('office', offices)
                
                # Execute query
                result = query.order('processed_date', desc=True).execute()
                
                if result.data:
                    st.session_state.export_query_results = result.data
                    st.success(f"✅ Found {len(result.data)} clips matching your criteria")
                else:
                    st.session_state.export_query_results = []
                    st.warning("No clips found matching your criteria")
        
        # Display results and export options
        if 'export_query_results' in st.session_state and st.session_state.export_query_results:
            clips_data = st.session_state.export_query_results
            
            # Create DataFrame for display
            df = pd.DataFrame(clips_data)
            
            # Quick stats
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Clips", len(df))
            with col2:
                avg_score = df['relevance_score'].mean() if 'relevance_score' in df.columns else 0
                st.metric("Avg Relevance", f"{avg_score:.1f}")
            with col3:
                sentiment_counts = df['overall_sentiment'].value_counts() if 'overall_sentiment' in df.columns else {}
                top_sentiment = sentiment_counts.index[0] if len(sentiment_counts) > 0 else "N/A"
                st.metric("Top Sentiment", top_sentiment)
            with col4:
                exported_count = len(df[df['workflow_stage'] == 'exported']) if 'workflow_stage' in df.columns else 0
                st.metric("Exported", exported_count)
            
            # Show preview
            st.markdown("### 📋 Preview (First 10 rows)")
            
            # Create display dataframe with selected columns
            display_columns = ['wo_number', 'office', 'make', 'model', 'contact', 'media_outlet', 
                             'relevance_score', 'overall_sentiment', 'processed_date']
            available_columns = [col for col in display_columns if col in df.columns]
            preview_df = df[available_columns].head(10)
            
            # Format dates
            if 'processed_date' in preview_df.columns:
                preview_df['processed_date'] = pd.to_datetime(preview_df['processed_date']).dt.strftime('%Y-%m-%d %H:%M')
            
            st.dataframe(preview_df, use_container_width=True)
            
            # Export options
            st.markdown("### 📥 Export Options")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Excel export with hyperlinks
                if st.button("📊 Generate Excel Report", type="primary", key="export_excel_btn"):
                    with st.spinner("Generating Excel report..."):
                        # Create Excel file with formatting
                        output = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Clip Export"
                        
                        # Define headers based on your screenshot
                        headers = ['Activity_ID', 'Office', 'WO#', 'Make', 'Model', 'Contact', 
                                 'Media Outlet', 'URL', 'Relevance', 'Sentiment']
                        
                        # Header styling
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill("solid", fgColor="366092")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Write headers
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Write data with hyperlinks
                        for row_idx, clip in enumerate(clips_data, 2):
                            ws.cell(row=row_idx, column=1, value=clip.get('activity_id', ''))
                            ws.cell(row=row_idx, column=2, value=clip.get('office', ''))
                            ws.cell(row=row_idx, column=3, value=clip.get('wo_number', ''))
                            ws.cell(row=row_idx, column=4, value=clip.get('make', ''))
                            ws.cell(row=row_idx, column=5, value=clip.get('model', ''))
                            ws.cell(row=row_idx, column=6, value=clip.get('contact', ''))
                            ws.cell(row=row_idx, column=7, value=clip.get('media_outlet', ''))
                            
                            # Add hyperlink to URL
                            url = clip.get('clip_url', '')
                            if url:
                                ws.cell(row=row_idx, column=8, value=url).hyperlink = url
                                ws.cell(row=row_idx, column=8).font = Font(color="0563C1", underline="single")
                            else:
                                ws.cell(row=row_idx, column=8, value='')
                            
                            # Add relevance score with color coding
                            relevance = clip.get('relevance_score', 0)
                            cell = ws.cell(row=row_idx, column=9, value=relevance)
                            if relevance >= 8:
                                cell.font = Font(color="28a745", bold=True)
                            elif relevance >= 5:
                                cell.font = Font(color="007bff")
                            else:
                                cell.font = Font(color="ffc107")
                            
                            # Add sentiment with emoji
                            sentiment = clip.get('overall_sentiment', '')
                            sentiment_display = {
                                'POS': '😊 POS',
                                'NEU': '😐 NEU', 
                                'NEG': '😟 NEG'
                            }.get(sentiment, sentiment)
                            ws.cell(row=row_idx, column=10, value=sentiment_display)
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        # Add borders
                        border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in ws.iter_rows(min_row=1, max_row=len(clips_data)+1, min_col=1, max_col=10):
                            for cell in row:
                                cell.border = border
                        
                        # Save workbook
                        wb.save(output)
                        output.seek(0)
                        
                        # Generate filename
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"clip_export_{timestamp}.xlsx"
                        
                        st.session_state.export_excel_data = output.getvalue()
                        st.session_state.export_excel_filename = filename
                        st.success("✅ Excel report generated!")
            
            with col2:
                # CSV export
                csv_data = df.to_csv(index=False).encode('utf-8')
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="📄 Download CSV",
                    data=csv_data,
                    file_name=f"clip_export_{timestamp}.csv",
                    mime="text/csv"
                )
            
            with col3:
                # JSON export (same as FMS export)
                json_data = json.dumps(clips_data, indent=2, default=str)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="📋 Download JSON",
                    data=json_data,
                    file_name=f"clip_export_{timestamp}.json",
                    mime="application/json"
                )
            
            # Show Excel download button if generated
            if 'export_excel_data' in st.session_state and st.session_state.export_excel_data:
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=st.session_state.export_excel_data,
                        file_name=st.session_state.export_excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_export_excel"
                    )
                    st.info("💡 Excel file includes clickable hyperlinks in the URL column!")
    
    except Exception as e:
        st.error(f"❌ Error in Export tab: {e}")
        import traceback
        st.error(f"Full error: {traceback.format_exc()}")

