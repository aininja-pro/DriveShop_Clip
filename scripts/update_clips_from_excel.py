#!/usr/bin/env python3
"""
Update clips in Supabase from Excel spreadsheet.

Handles Office, Media Outlet, Media Outlet ID, and Circulation updates for existing clips.
Supports dry-run mode, validation, and comprehensive reporting.

Usage:
    # Dry-run (default, no database writes)
    python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx" --dry-run

    # Update with limit
    python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx" --limit 50

    # Full update
    python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx"
"""

import sys
import os
import json
import time
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional

# Setup path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from src.utils.database import DatabaseManager
from src.utils.logger import setup_logger

logger = setup_logger(__name__)


def load_and_validate_excel(file_path: str) -> pd.DataFrame:
    """Load Excel and validate required columns exist.

    Args:
        file_path: Path to Excel file

    Returns:
        DataFrame with loaded data

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If required columns are missing
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Load Excel with openpyxl engine (supports .xlsx)
    df = pd.read_excel(file_path, engine='openpyxl')

    if df.empty:
        raise ValueError("Excel file is empty")

    # Check for required columns
    required_columns = ['work_order_number']
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")

    return df


def resolve_formulas(file_path: str, df: pd.DataFrame) -> pd.DataFrame:
    """Handle VLOOKUP formulas in spreadsheet by loading with data_only=True.

    Args:
        file_path: Path to Excel file
        df: DataFrame loaded with pandas

    Returns:
        DataFrame with formulas resolved
    """
    # Try to load with openpyxl data_only mode to evaluate formulas
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Convert to list of dicts
        data = []
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        wb.close()

        # Create new DataFrame from evaluated values
        df_resolved = pd.DataFrame(data)

        # If we got different data, use it; otherwise stick with original
        if len(df_resolved) > 0:
            return df_resolved
        else:
            return df

    except Exception as e:
        logger.warning(f"Could not resolve formulas with openpyxl: {e}. Using pandas-loaded data.")
        return df


def validate_data(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """Validate data types and return clean DataFrame + errors.

    Args:
        df: Input DataFrame

    Returns:
        Tuple of (clean_df, validation_errors)
    """
    errors = []
    valid_rows = []

    for idx, row in df.iterrows():
        # Extract and validate WO#
        wo_number = row.get('work_order_number')

        if pd.isna(wo_number) or str(wo_number).strip() == '':
            errors.append(f"Row {idx + 2}: Missing WO number")
            continue

        # Check if at least one update field is present
        has_office = pd.notna(row.get('Office'))
        has_media_outlet = pd.notna(row.get('Media Outlet'))
        has_outlet_id = pd.notna(row.get('Media Outlet ID'))
        has_circulation = pd.notna(row.get('Circulation'))

        if not (has_office or has_media_outlet or has_outlet_id or has_circulation):
            # No data to update, skip
            continue

        valid_rows.append(idx)

    # Filter to valid rows
    clean_df = df.loc[valid_rows].copy()

    return clean_df, errors


def detect_database_schema(db: DatabaseManager) -> Dict[str, str]:
    """Query database to detect actual field names for outlet_id field.

    Args:
        db: DatabaseManager instance

    Returns:
        Dict with field name mappings
    """
    # Try to query a single clip to see what fields exist
    try:
        result = db.supabase.table('clips').select('*').limit(1).execute()

        if result.data and len(result.data) > 0:
            fields = result.data[0].keys()

            # Detect which outlet ID field exists
            if 'media_outlet_id' in fields:
                outlet_id_field = 'media_outlet_id'
            elif 'outlet_id' in fields:
                outlet_id_field = 'outlet_id'
            else:
                # Default to media_outlet_id
                outlet_id_field = 'media_outlet_id'
                logger.warning("Could not detect outlet_id field name, defaulting to 'media_outlet_id'")

            return {
                'outlet_id_field': outlet_id_field
            }
        else:
            logger.warning("No clips in database to detect schema. Using default field names.")
            return {'outlet_id_field': 'media_outlet_id'}

    except Exception as e:
        logger.error(f"Failed to detect database schema: {e}")
        return {'outlet_id_field': 'media_outlet_id'}


def process_batch(batch_df: pd.DataFrame, db: DatabaseManager,
                 schema_map: Dict[str, str], dry_run: bool = True) -> Dict[str, Any]:
    """Process a batch of records and return results.

    Args:
        batch_df: DataFrame batch to process
        db: DatabaseManager instance
        schema_map: Field name mappings from detect_database_schema()
        dry_run: If True, don't actually update database

    Returns:
        Dict with results: {'updated': [...], 'not_found': [...], 'failed': [...]}
    """
    results = {
        'updated': [],
        'not_found': [],
        'failed': []
    }

    outlet_id_field = schema_map.get('outlet_id_field', 'media_outlet_id')

    for idx, row in batch_df.iterrows():
        # Extract values
        wo_number = str(row['work_order_number']).strip()
        office = str(row['Office']) if pd.notna(row.get('Office')) else None
        media_outlet = str(row['Media Outlet']) if pd.notna(row.get('Media Outlet')) else None

        # Convert numeric fields
        try:
            outlet_id = int(row['Media Outlet ID']) if pd.notna(row.get('Media Outlet ID')) else None
        except (ValueError, TypeError):
            outlet_id = None

        try:
            circulation = int(row['Circulation']) if pd.notna(row.get('Circulation')) else None
        except (ValueError, TypeError):
            circulation = None

        # Build update dictionary
        update_data = {}
        fields_updated = []

        if office:
            update_data['office'] = office
            fields_updated.append('office')

        if media_outlet:
            update_data['media_outlet'] = media_outlet
            fields_updated.append('media_outlet')

        if outlet_id is not None:
            update_data[outlet_id_field] = outlet_id
            fields_updated.append(outlet_id_field)

        if circulation is not None:
            update_data['impressions'] = circulation
            fields_updated.append('impressions')

        if not update_data:
            # No data to update
            continue

        try:
            # First, check if WO exists in database
            check_result = db.supabase.table('clips').select('wo_number').eq('wo_number', wo_number).execute()

            if not check_result.data or len(check_result.data) == 0:
                # WO not found in database
                results['not_found'].append(wo_number)
                logger.debug(f"WO# {wo_number} not found in database")
                continue

            # WO exists, proceed with update
            if dry_run:
                # Dry run - just log what would be updated
                results['updated'].append({
                    'wo_number': wo_number,
                    'fields_updated': fields_updated,
                    'update_data': update_data
                })
                logger.debug(f"[DRY RUN] Would update WO# {wo_number}: {fields_updated}")
            else:
                # Actual update with retry logic
                max_retries = 3
                retry_delay = 1  # seconds

                for attempt in range(max_retries):
                    try:
                        result = db.supabase.table('clips').update(update_data).eq('wo_number', wo_number).execute()

                        if result.data:
                            results['updated'].append({
                                'wo_number': wo_number,
                                'fields_updated': fields_updated
                            })
                            logger.debug(f"Updated WO# {wo_number}: {fields_updated}")
                            break
                        else:
                            if attempt < max_retries - 1:
                                time.sleep(retry_delay * (2 ** attempt))  # Exponential backoff
                                continue
                            else:
                                raise Exception("Update returned no data")

                    except Exception as e:
                        if attempt < max_retries - 1:
                            logger.warning(f"Retry {attempt + 1}/{max_retries} for WO# {wo_number}: {e}")
                            time.sleep(retry_delay * (2 ** attempt))
                        else:
                            raise

        except Exception as e:
            results['failed'].append({
                'wo_number': wo_number,
                'error': str(e)
            })
            logger.error(f"Failed to update WO# {wo_number}: {e}")

    return results


def generate_reports(results: Dict[str, Any], run_metadata: Dict[str, Any],
                    validation_summary: Dict[str, Any], output_dir: str):
    """Generate JSON and text reports.

    Args:
        results: Processing results dict
        run_metadata: Metadata about the run
        validation_summary: Summary of validation results
        output_dir: Directory to save reports
    """
    # Create reports directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Prepare processing summary
    processing_summary = {
        'total_processed': len(results['updated']) + len(results['not_found']) + len(results['failed']),
        'successful_updates': len(results['updated']),
        'not_found_in_db': len(results['not_found']),
        'failed_updates': len(results['failed'])
    }

    # Generate JSON report
    json_report = {
        'run_metadata': run_metadata,
        'validation_summary': validation_summary,
        'processing_summary': processing_summary,
        'detailed_results': results
    }

    json_path = os.path.join(output_dir, f'update_{timestamp}.json')
    with open(json_path, 'w') as f:
        json.dump(json_report, f, indent=2, default=str)

    print(f"  ✓ Saved: {json_path}")

    # Generate text file for not-found WO#s
    if results['not_found']:
        txt_path = os.path.join(output_dir, f'not_found_{timestamp}.txt')
        with open(txt_path, 'w') as f:
            f.write(f"WO#s Not Found in Database ({len(results['not_found'])} total):\n")
            for wo in results['not_found']:
                f.write(f"{wo}\n")

        print(f"  ✓ Saved: {txt_path}")

    return json_path


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description='Update clips in Supabase from Excel spreadsheet',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Dry-run mode (default)
  python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx" --dry-run

  # Update first 50 records
  python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx" --limit 50

  # Full update
  python scripts/update_clips_from_excel.py --input "data/query_clips_highlighted updated.xlsx"
        """
    )

    parser.add_argument('--input', type=str, required=True,
                       help='Path to Excel file')
    parser.add_argument('--dry-run', action='store_true',
                       help='Preview changes without updating database (default: False)')
    parser.add_argument('--limit', type=int, default=None,
                       help='Limit number of records to process')
    parser.add_argument('--batch-size', type=int, default=100,
                       help='Batch size for processing (default: 100)')
    parser.add_argument('--reports-dir', type=str, default='reports',
                       help='Directory for output reports (default: reports)')

    args = parser.parse_args()

    # Print header
    print("=" * 80)
    print("EXCEL TO SUPABASE CLIP UPDATER")
    print("=" * 80)

    # Stage 1: Load Excel
    print("\n[1/6] Loading Excel file...")
    try:
        df = load_and_validate_excel(args.input)
        print(f"  ✓ Loaded {len(df)} rows from {os.path.basename(args.input)}")
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return 1

    # Stage 2: Resolve formulas
    print("\n[2/6] Resolving formulas...")
    try:
        df = resolve_formulas(args.input, df)
        print(f"  ✓ Formula evaluation complete")
    except Exception as e:
        print(f"  ⚠ Warning: {e}")

    # Stage 3: Validate data
    print("\n[3/6] Validating data...")
    clean_df, validation_errors = validate_data(df)

    skipped_no_updates = len(df) - len(clean_df) - len(validation_errors)

    print(f"  ✓ {len(clean_df)} valid records")
    if validation_errors:
        print(f"  ⚠ {len(validation_errors)} validation errors")
    if skipped_no_updates > 0:
        print(f"  ⚠ {skipped_no_updates} skipped (no updates)")

    if clean_df.empty:
        print("\n  ✗ No valid records to process")
        return 1

    # Apply limit if specified
    if args.limit:
        clean_df = clean_df.head(args.limit)
        print(f"  ℹ Limited to first {len(clean_df)} records")

    # Stage 4: Detect database schema
    print("\n[4/6] Detecting database schema...")
    try:
        db = DatabaseManager()
        schema_map = detect_database_schema(db)
        print(f"  ✓ Field mapping: outlet_id → {schema_map['outlet_id_field']}")
    except Exception as e:
        print(f"  ✗ Error connecting to database: {e}")
        return 1

    # Stage 5: Process records
    mode_text = "DRY RUN MODE" if args.dry_run else "UPDATE MODE"
    print(f"\n[5/6] Processing records ({mode_text})...")

    all_results = {
        'updated': [],
        'not_found': [],
        'failed': []
    }

    # Process in batches with progress bar
    total_records = len(clean_df)
    batch_size = args.batch_size

    with tqdm(total=total_records, desc="  Processing", unit=" records") as pbar:
        for start_idx in range(0, total_records, batch_size):
            end_idx = min(start_idx + batch_size, total_records)
            batch_df = clean_df.iloc[start_idx:end_idx]

            batch_results = process_batch(batch_df, db, schema_map, dry_run=args.dry_run)

            # Merge results
            all_results['updated'].extend(batch_results['updated'])
            all_results['not_found'].extend(batch_results['not_found'])
            all_results['failed'].extend(batch_results['failed'])

            pbar.update(len(batch_df))

    # Stage 6: Generate reports
    print("\n[6/6] Generating reports...")

    run_metadata = {
        'timestamp': datetime.now().isoformat(),
        'input_file': args.input,
        'mode': 'dry_run' if args.dry_run else 'update',
        'total_rows_in_file': len(df),
        'limit_applied': args.limit
    }

    validation_summary = {
        'valid_records': len(clean_df),
        'validation_errors': len(validation_errors),
        'skipped_no_updates': skipped_no_updates
    }

    report_path = generate_reports(all_results, run_metadata, validation_summary, args.reports_dir)

    # Print summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total processed:       {len(all_results['updated']) + len(all_results['not_found']) + len(all_results['failed'])}")
    print(f"Successful updates:    {len(all_results['updated'])}")
    print(f"Not found in DB:       {len(all_results['not_found'])}")
    print(f"Failed updates:        {len(all_results['failed'])}")
    print("=" * 80)

    if args.dry_run:
        print("\n⚠  DRY RUN MODE - No changes made to database")
        print("Run without --dry-run to apply updates")
    else:
        print("\n✅ Updates applied to database")

    print("=" * 80)

    return 0


if __name__ == "__main__":
    sys.exit(main())
